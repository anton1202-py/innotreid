import base64
import glob
import io
import json
import math
import os
import shutil
import time
from collections import Counter
from contextlib import closing
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

import dropbox
import openpyxl
import pandas as pd
import psycopg2
import pythoncom
import requests
from dotenv import load_dotenv
from helpers import (merge_barcode_for_ozon_two_on_two,
                     new_data_for_ozon_ticket, print_barcode_to_pdf2,
                     qrcode_print_for_products, supply_qrcode_to_standart_view)
from openpyxl import Workbook, load_workbook
from openpyxl.drawing import image
from openpyxl.styles import Alignment, Border, Font, Side
from PIL import Image, ImageDraw, ImageFont
from sqlalchemy import create_engine
from win32com.client import DispatchEx

version = 'w1.0'

# Загрузка переменных окружения из файла .env
dotenv_path = os.path.join(os.path.dirname(
    __file__), '..', 'web_barcode', '.env')
load_dotenv(dotenv_path)

REFRESH_TOKEN_DB = os.getenv('REFRESH_TOKEN_DB')
APP_KEY_DB = os.getenv('APP_KEY_DB')
APP_SECRET_DB = os.getenv('APP_SECRET_DB')
API_KEY_WB_IP = os.getenv('API_KEY_WB_IP')

wb_headers_karavaev = {
    'Content-Type': 'application/json',
    'Authorization': API_KEY_WB_IP
}

ozon_headers_karavaev = {
    'Api-Key': os.getenv('API_KEY_OZON_KARAVAEV'),
    'Content-Type': 'application/json',
    'Client-Id': os.getenv('CLIENT_ID_OZON_KARAVAEV')
}

dbx_db = dropbox.Dropbox(oauth2_refresh_token=REFRESH_TOKEN_DB,
                         app_key=APP_KEY_DB,
                         app_secret=APP_SECRET_DB)

delivery_date = datetime.today().strftime("%d.%m.%Y %H-%M-%S")


def stream_dropbox_file(path):
    _, res = dbx_db.files_download(path)
    with closing(res) as result:
        byte_data = result.content
        return io.BytesIO(byte_data)


def clearning_folders():
    dir = 'web_barcode/fbs_mode/data_for_barcodes/'
    for file_name in os.listdir(dir):
        file_path = os.path.join(dir, file_name)
        if os.path.isfile(file_path):
            os.unlink(file_path)

    dirs = ['web_barcode/fbs_mode/data_for_barcodes/cache_dir',
            'web_barcode/fbs_mode/data_for_barcodes/done_data',
            'web_barcode/fbs_mode/data_for_barcodes/pivot_excel',
            'web_barcode/fbs_mode/data_for_barcodes/qrcode_folder/cache_dir_3',
            'web_barcode/fbs_mode/data_for_barcodes/qrcode_folder',
            'web_barcode/fbs_mode/data_for_barcodes/qrcode_supply',
            'web_barcode/fbs_mode/data_for_barcodes/package_tickets/done',
            'web_barcode/fbs_mode/data_for_barcodes/package_tickets',
            'web_barcode/fbs_mode/data_for_barcodes/ozon_docs',
            'web_barcode/fbs_mode/data_for_barcodes/ozon_delivery_barcode'
            ]
    for dir in dirs:
        for filename in glob.glob(os.path.join(dir, "*")):
            file_path = os.path.join(dir, filename)
            try:
                if os.path.isfile(filename) or os.path.islink(filename):
                    os.unlink(filename)
                elif os.path.isdir(filename):
                    shutil.rmtree(filename)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (filename, e))


class WildberriesFbsMode():
    """Класс отвечает за работу с заказами Wildberries"""

    def __init__(self):
        """Основные данные класса"""
        self.amount_articles = {}

    def article_data_for_tickets(self):
        """
        Функция обрабатывает новые сборочные задания.
        Выделяет артикулы продавца светильников, их баркоды и наименования.
        Создает словарь с данными каждого артикулы и словарь с количеством каждого
        артикула.
        """
        url = "https://suppliers-api.wildberries.ru/api/v3/orders/new"

        # Список с ID соборочных заданий
        orders_id_list = []
        # Список с артикулами_продавца соборочных заданий
        order_articles_list = []

        # Словарь с данными {id_задания: артикул_продавца}
        self.article_id_dict = {}

        response = requests.request("GET", url, headers=wb_headers_karavaev)
        orders_data = json.loads(response.text)['orders']

        for order in orders_data:
            order_articles_list.append(order['article'])

        # Словарь с данными артикула: {артикул_продавца: [баркод, наименование]}
        self.data_article_info_dict = {}
        url_data = "https://suppliers-api.wildberries.ru/content/v1/cards/cursor/list"

        # Список только для артикулов ночников. Остальные отфильтровывает
        self.clear_article_list = []

        for article in order_articles_list:
            payload = json.dumps({
                "sort": {
                    "cursor": {
                        "limit": 1
                    },
                    "filter": {
                        "textSearch": article,
                        "withPhoto": -1
                    }
                }
            })
            response_data = requests.request(
                "POST", url_data, headers=wb_headers_karavaev, data=payload)
            print(json.loads(response_data.text))
            if json.loads(response_data.text)[
                    'data']['cards'][0]['object'] == "Ночники":
                self.clear_article_list.append(article)

                # Достаем баркод артикула (первый из списка, если их несколько)
                barcode = json.loads(response_data.text)[
                    'data']['cards'][0]['sizes'][0]['skus'][0]
                # Достаем название артикула
                title = json.loads(response_data.text)[
                    'data']['cards'][0]['title']
                self.data_article_info_dict[article] = [title, barcode]
        # Словарь с данными: {артикул_продавца: количество}
        self.amount_articles = dict(Counter(self.clear_article_list))
        print('self.amount_articles', self.amount_articles)

        for order in orders_data:
            if order['article'] in self.clear_article_list:
                self.article_id_dict[order['id']] = order['article']
                orders_id_list.append(order['id'])

        # Словарь для данных листа подбора.
        self.selection_dict = {}
        # Собирам данные для Листа подбора
        for order_id in self.article_id_dict.keys():
            payload_order = json.dumps({
                "sort": {
                    "cursor": {
                        "limit": 1
                    },
                    "filter": {
                        "textSearch": self.article_id_dict[order_id],
                        "withPhoto": -1
                    }
                }
            })

            response_order = requests.request(
                "POST", url_data, headers=wb_headers_karavaev, data=payload_order)
            photo = json.loads(response_order.text)[
                'data']['cards'][0]['mediaFiles'][0]
            brand = json.loads(response_order.text)[
                'data']['cards'][0]['brand']
            title_article = json.loads(response_order.text)[
                'data']['cards'][0]['title']
            print('title_article', title_article)
            seller_article = self.article_id_dict[order_id]
            # Заполняем словарь данными для Листа подбора
            self.selection_dict[order_id] = [
                photo, brand, title_article, seller_article]
        return self.amount_articles

    def create_delivery(self):
        """Создание поставки"""
        url_data = 'https://suppliers-api.wildberries.ru/api/v3/supplies'
        hour = datetime.now().hour
        delivery_name = f"Поставка {delivery_date}"
        if hour >= 18 or hour <= 6:
            delivery_name = f'Производство ночь {delivery_date}'
        else:
            delivery_name = f'Производство день {delivery_date}'

        payload = json.dumps(
            {
                "name": delivery_name
            }
        )
        # Из этой переменной достать ID поставки
        response_data = requests.request(
            "POST", url_data, headers=wb_headers_karavaev, data=payload)
        # print(response_data)
        self.supply_id = json.loads(response_data.text)['id']

    def qrcode_order(self):
        """
        Функция добавляет сборочные задания по их id
        в созданную поставку и получает qr стикер каждого
        задания и сохраняет его в папку
        """
        # Вызываем эндпоинт для создания поставки и определения ее delivery_id
        for order in self.article_id_dict.keys():
            add_url = f'https://suppliers-api.wildberries.ru/api/v3/supplies/{self.supply_id}/orders/{order}'

            response = requests.request(
                "PATCH", add_url, headers=wb_headers_karavaev)

        for order in self.article_id_dict.keys():
            ticket_url = 'https://suppliers-api.wildberries.ru/api/v3/orders/stickers?type=png&width=58&height=40'
            payload_ticket = json.dumps({"orders": [order]})
            response_ticket = requests.request(
                "POST", ticket_url, headers=wb_headers_karavaev, data=payload_ticket)

            # Расшифровываю ответ, чтобы сохранить файл этикетки задания
            ticket_data = json.loads(response_ticket.text)[
                "stickers"][0]["file"]

            # Узнаю стикер сборочного задания и помещаю его в словарь с данными для
            # листа подбора
            sticker_code_first_part = json.loads(response_ticket.text)[
                "stickers"][0]["partA"]
            sticker_code_second_part = json.loads(response_ticket.text)[
                "stickers"][0]["partB"]
            sticker_code = f'{sticker_code_first_part} {sticker_code_second_part}'
            self.selection_dict[order].append(sticker_code)

            # декодируем строку из base64 в бинарные данные
            binary_data = base64.b64decode(ticket_data)

            # создаем объект изображения из бинарных данных
            img = Image.open(BytesIO(binary_data))

            # сохраняем изображение в файл

            img.save(
                f"web_barcode/fbs_mode/data_for_barcodes/qrcode_folder/{order} {self.article_id_dict[order]}.png")

    def create_selection_list(self):
        """Создает лист сборки"""
        # создаем новую книгу Excel
        selection_file = Workbook()
        COUNT_HELPER = 2
        # выбираем лист Sheet1
        create = selection_file.create_sheet(title='pivot_list', index=0)
        sheet = selection_file['pivot_list']

        # Установка параметров печати
        create.page_setup.paperSize = create.PAPERSIZE_A4
        create.page_setup.orientation = create.ORIENTATION_PORTRAIT
        create.page_margins.left = 0.25
        create.page_margins.right = 0.25
        create.page_margins.top = 0.25
        create.page_margins.bottom = 0.25
        create.page_margins.header = 0.3
        create.page_margins.footer = 0.3

        sheet['A1'] = '№ Задания'
        sheet['B1'] = 'Фото'
        sheet['C1'] = 'Бренд'
        sheet['D1'] = 'Наименование'
        sheet['E1'] = 'Артикул продавца'
        sheet['F1'] = 'Стикер'

        for key, value in self.selection_dict.items():
            # # загружаем изображение
            response = requests.get(value[0])
            img = image.Image(io.BytesIO(response.content))
            # задаем размеры изображения
            img.width = 30
            img.height = 50

            create.cell(row=COUNT_HELPER, column=1).value = key
            # вставляем изображение в Столбец В
            sheet.add_image(img, f'B{COUNT_HELPER}')
            # create.cell(row=COUNT_HELPER, column=2).value = value[0]
            create.cell(row=COUNT_HELPER, column=3).value = value[1]
            create.cell(row=COUNT_HELPER, column=4).value = value[2]
            create.cell(row=COUNT_HELPER, column=5).value = value[3]
            create.cell(row=COUNT_HELPER, column=6).value = value[4]
            COUNT_HELPER += 1
        name_selection_file = 'web_barcode/fbs_mode/data_for_barcodes/pivot_excel/Лист подбора.xlsx'
        path_file = os.path.abspath(name_selection_file)

        selection_file.save(name_selection_file)

        w_b2 = load_workbook(name_selection_file)
        source_page2 = w_b2.active

        al = Alignment(horizontal="center", vertical="center")
        al_left = Alignment(horizontal="left",
                            vertical="center", wrapText=True)

        source_page2.column_dimensions['A'].width = 10  # Номер задания
        source_page2.column_dimensions['B'].width = 5  # Картинка
        source_page2.column_dimensions['C'].width = 15  # Бренд
        source_page2.column_dimensions['D'].width = 25  # Наименование
        source_page2.column_dimensions['E'].width = 16  # Артикул продавца
        source_page2.column_dimensions['F'].width = 16  # Стикер

        thin = Side(border_style="thin", color="000000")
        for i in range(len(self.selection_dict)+1):
            for c in source_page2[f'A{i+1}:F{i+1}']:
                c[0].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[0].font = Font(size=9)
                c[0].alignment = al

                c[1].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[1].font = Font(size=9)
                c[1].alignment = al

                c[2].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[2].font = Font(size=9)
                c[2].alignment = al_left

                c[3].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[3].font = Font(size=9)
                c[3].alignment = al_left

                c[4].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[4].font = Font(size=9)
                c[4].alignment = al_left

                c[5].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[5].font = Font(size=9)
                c[5].alignment = al_left

        # Увеличиваем высоту строки
        for i in range(2, len(self.selection_dict) + 2):
            source_page2.row_dimensions[i].height = 40

        w_b2.save(name_selection_file)

        xl = DispatchEx("Excel.Application")
        xl.DisplayAlerts = False
        folder_path = os.path.dirname(os.path.abspath(path_file))
        name_for_file = f'Лист подбора {delivery_date}'
        name_xls_dropbox = f'Лист подбора {delivery_date}'
        wb = xl.Workbooks.Open(path_file)
        xl.CalculateFull()
        pythoncom.PumpWaitingMessages()
        try:
            wb.ExportAsFixedFormat(
                0, f'{folder_path}/{name_for_file}.pdf')
        except Exception as e:
            print("Failed to convert in PDF format.Please confirm environment meets all the requirements  and try again")
        finally:
            wb.Close()

        # Сохраняем на DROPBOX
        with open(f'{folder_path}/{name_for_file}.pdf', 'rb') as f:
            dbx_db.files_upload(
                f.read(), f'/DATABASE/beta/{name_for_file}.pdf')

    def qrcode_supply(self):
        """
        Функция добавляет поставку в доставку, получает QR код поставки
        и преобразует этот QR код в необходимый формат.
        """
        # Переводим поставку в доставку
        url_to_supply = f'https://suppliers-api.wildberries.ru/api/v3/supplies/{self.supply_id}/deliver'
        response_to_supply = requests.request(
            "PATCH", url_to_supply, headers=wb_headers_karavaev)

        # Получаем QR код поставки:
        url_supply_qrcode = f"https://suppliers-api.wildberries.ru/api/v3/supplies/{self.supply_id}/barcode?type=png"
        response_supply_qrcode = requests.request(
            "GET", url_supply_qrcode, headers=wb_headers_karavaev)

        # Создаем QR код поставки
        qrcode_base64_data = json.loads(response_supply_qrcode.text)["file"]

        # декодируем строку из base64 в бинарные данные
        binary_data = base64.b64decode(qrcode_base64_data)
        # создаем объект изображения из бинарных данных
        img = Image.open(BytesIO(binary_data))
        # сохраняем изображение в файл
        img.save(
            f"web_barcode/fbs_mode/data_for_barcodes/qrcode_supply/{self.supply_id}.png")

    def list_for_print_create(self):
        """
        Функция создает список с полными именами файлов, которые нужно объединить
        amount_articles - словарь с данными {артикул_продавца: количество}
        """
        qrcode_list = qrcode_print_for_products()
        pdf_filenames = glob.glob(
            'web_barcode/fbs_mode/data_for_barcodes/cache_dir/*.pdf')
        list_pdf_file_ticket_for_complect = []
        if self.amount_articles:
            for j in pdf_filenames:
                while self.amount_articles[str(Path(j).stem)] > 0:
                    list_pdf_file_ticket_for_complect.append(j)
                    self.amount_articles[str(Path(j).stem)] -= 1
            for file in qrcode_list:
                list_pdf_file_ticket_for_complect.append(file)
            # Определяем число qr кодов для поставки.
            amount_of_supply_qrcode = math.ceil(
                len(list_pdf_file_ticket_for_complect)/20)
            print('list_pdf_file_ticket_for_complect',
                  list_pdf_file_ticket_for_complect)
            print('amount_of_supply_qrcode', amount_of_supply_qrcode)
            outer_list = []  # Внешний список для процесса сортировки
            for i in list_pdf_file_ticket_for_complect:
                # Разделяю полное название файла на путь к файлу и имя файла
                # Оказывается в python знаком \ отделяется последняя папка перед файлом
                # А все внешние отделяются знаком /
                new_name = i.split('\\')
                full_new_name = []  # Список с полным именени файла после разделения
                # Имена QR кодов у меня составные. Состоят из нескольких слов с пробелами
                # В этом цикле разделяю имена из предыдущих списков по пробелу.
                for j in new_name:
                    split_name = j.split(' ')
                    full_new_name.append(split_name)
                outer_list.append(full_new_name)
            # Сортирую самый внешний список по последнему элемену самого внутреннего списка
            sorted_list = sorted(outer_list, key=lambda x: x[-1][-1])

            # Далее идет обратный процесс - процесс объединения элементов списка
            # в первоначальные имена файлов, но уже отсортированные
            new_sort = []
            for i in sorted_list:
                inner_new_sort = []
                for j in i:
                    j = ' '.join(j)
                    inner_new_sort.append(j)
                new_sort.append(inner_new_sort)

            last_sorted_list = []
            for i in new_sort:
                i = '/'.join(i)
                last_sorted_list.append(i)

            list_pdf_file_ticket_for_complect = last_sorted_list

            qrcode_supply_amount = supply_qrcode_to_standart_view()[0]

            while amount_of_supply_qrcode > 0:
                list_pdf_file_ticket_for_complect.append(qrcode_supply_amount)
                amount_of_supply_qrcode -= 1

            file_name = (f'web_barcode/fbs_mode/data_for_barcodes/done_data/Наклейки для комплектовщиков '
                         f'{time.strftime("%Y-%m-%d %H-%M")}.pdf')
            print_barcode_to_pdf2(list_pdf_file_ticket_for_complect, file_name)


class OzonFbsMode():
    """Класс отвечает за работу с заказами ОЗОН"""

    def __init__(self):
        # Данные отправлений для FBS OZON
        self.posting_data = []
        self.warehouse_list = [1020000089903000, 22655170176000]
        self.date_for_files = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
        self.main_save_folder_server = 'web_barcode/fbs_mode/data_for_barcodes'
        self.dropbox_main_fbs_folder = '/DATABASE/beta'
        self.date_before = datetime.now() - timedelta(days=20)
        self.tomorrow_date = datetime.now() + timedelta(days=1)
        self.future_date = datetime.now() + timedelta(days=20)

        self.since_date_for_filter = self.date_before.strftime(
            '%Y-%m-%d') + 'T08:00:00Z'

        # Словарь с данными {'артикул продавца': 'количество'}. Для сводной таблицы по FBS
        self.ozon_article_amount = {}

    def awaiting_packaging_orders(self):
        """
        Функция собирает новые заказы и преобразует данные
        от заказа в словарь для дальнейшей работы с ним
        """
        # Метод для просмотра новых заказов
        url = "https://api-seller.ozon.ru/v3/posting/fbs/unfulfilled/list"

        future_date_for_filter = self.future_date.strftime(
            '%Y-%m-%d') + 'T08:00:00Z'

        payload = json.dumps({
            "dir": "ASC",
            "filter": {
                "cutoff_from": self.since_date_for_filter,
                "cutoff_to": future_date_for_filter,
                "delivery_method_id": [],
                "provider_id": [],
                "status": "awaiting_packaging",
                "warehouse_id": []
            },
            "limit": 100,
            "offset": 0,
            "with": {
                "analytics_data": True,
                "barcodes": True,
                "financial_data": True,
                "translit": True
            }
        })

        response = requests.request(
            "POST", url, headers=ozon_headers_karavaev, data=payload)
        # Заполняю словарь данными
        for i in json.loads(response.text)['result']['postings']:
            inner_dict_data = {}
            inner_dict_data['posting_number'] = i['posting_number']
            inner_dict_data['order_id'] = i['order_id']
            inner_dict_data['delivery_method'] = i['delivery_method']
            inner_dict_data['in_process_at'] = i['in_process_at']
            inner_dict_data['shipment_date'] = i['shipment_date']
            inner_dict_data['products'] = i['products']
            self.posting_data.append(inner_dict_data)

        # После отработки кода удалить цикл и вывод длины списка
        for j in self.posting_data:
            print(j)
            print('***************')
        print(len(self.posting_data))

    def awaiting_deliver_orders(self):
        """
        Делит заказ на отправления и переводит его в статус awaiting_deliver.
        Каждый элемент в packages может содержать несколько элементов products или отправлений.
        Каждый элемент в products — это товар, включённый в данное отправление.
        """
        url = 'https://api-seller.ozon.ru/v4/posting/fbs/ship'
        for data_dict in self.posting_data:
            # print('posting_number', posting_number)
            # print('data_list', data_list)
            for sku_data in data_dict['products']:
                # print('sku_data[sku]', sku_data['sku'])
                # print('sku_data[quantity]', sku_data['quantity'])
                payload = json.dumps({
                    "packages": [
                        {
                            "products": [
                                {
                                    "product_id": sku_data['sku'],
                                    "quantity": sku_data['quantity']
                                }
                            ]
                        }
                    ],
                    "posting_number": data_dict['posting_number'],
                    "with": {
                        "additional_data": True
                    }
                })
                response = requests.request(
                    "POST", url, headers=ozon_headers_karavaev, data=payload)

    def check_folder_exists(self, path):
        try:
            dbx_db.files_list_folder(path)
            return True
        except dropbox.exceptions.ApiError as e:
            return False

    def prepare_data_for_confirm_delivery(self):
        """Подготовка данных для подтверждения отгрузки"""

        hour = datetime.now().hour
        date_folder = datetime.today().strftime('%Y-%m-%d')

        to_date_for_filter = self.tomorrow_date.strftime(
            '%Y-%m-%d') + 'T23:59:00Z'
        date_confirm_delivery = self.tomorrow_date.strftime('%Y-%m-%d')

        # Проверяем все отгрузки, которые буду завтра
        url = 'https://api-seller.ozon.ru/v3/posting/fbs/list'

        # Словарь с данными {'Номер склада': {quantity: 'Количество артикулов', containers_count: 'количество коробок'}}
        self.ware_house_amount_dict = {}
        # Словарь с данными {'номер отправления': {'артикул продавца': 'количество'}}
        self.fbs_ozon_common_data = {}

        if hour >= 18 or hour <= 6:
            self.delivary_method_id = 22655170176000
            self.dropbox_current_assembling_folder = f'{self.dropbox_main_fbs_folder}/НОЧЬ СБОРКА ФБС/{date_folder}'

        else:
            self.delivary_method_id = 1020000089903000
            self.dropbox_current_assembling_folder = f'{self.dropbox_main_fbs_folder}/ДЕНЬ СБОРКА ФБС/{date_folder}'
        self.departure_date = date_confirm_delivery + 'T10:00:00Z'

        # Создаем папку на dropbox, если ее еще нет
        if self.check_folder_exists(self.dropbox_current_assembling_folder) == False:
            dbx_db.files_create_folder_v2(
                self.dropbox_current_assembling_folder)

        amount_products = 0
        payload = json.dumps(
            {
                "dir": "asc",
                "filter": {
                    "delivery_method_id": [self.delivary_method_id],
                    "provider_id": [24],
                    "since": self.since_date_for_filter,
                    "status": "awaiting_deliver",
                    "to": to_date_for_filter,
                    "warehouse_id": [self.delivary_method_id]
                },
                "limit": 999,
                "offset": 0,
                "with": {
                    "analytics_data": True,
                    "barcodes": True,
                    "financial_data": True,
                    "translit": True
                }
            }
        )
        response = requests.request(
            "POST", url, headers=ozon_headers_karavaev, data=payload)
        for data in json.loads(response.text)['result']['postings']:
            inner_article_amount_dict = {}

            if self.tomorrow_date.strftime('%Y-%m-%d') in data["shipment_date"]:
                # if '29' in data["shipment_date"]:
                for product in data['products']:
                    amount_products += product['quantity']
                    inner_article_amount_dict[product['offer_id']
                                              ] = product['quantity']
                    if product['offer_id'] not in self.ozon_article_amount.keys():
                        self.ozon_article_amount[product['offer_id']] = int(
                            product['quantity'])
                    else:
                        self.ozon_article_amount[product['offer_id']
                                                 ] = self.ozon_article_amount[product['offer_id']] + int(product['quantity'])
                self.fbs_ozon_common_data[data['posting_number']
                                          ] = inner_article_amount_dict
        containers_count = math.ceil(amount_products/20)
        self.ware_house_amount_dict[self.delivary_method_id] = {
            'quantity': amount_products, 'containers_count': containers_count}

        # print(self.delivary_method_id)
        # print(self.departure_date)
        print('self.ware_house_amount_dict', self.ware_house_amount_dict)
        print('self.fbs_ozon_common_data', self.fbs_ozon_common_data)
        print('self.ozon_article_amount', self.ozon_article_amount)
        return self.ozon_article_amount

    def confirm_delivery_create_document(self):
        """Функция подтверждает отгрузку и запускает создание документов на стороне ОЗОН"""
        url = 'https://api-seller.ozon.ru/v2/posting/fbs/act/create'
        if self.ware_house_amount_dict[self.delivary_method_id]['quantity'] > 0:
            payload = json.dumps({
                "containers_count": self.ware_house_amount_dict[self.delivary_method_id]['containers_count'],
                "delivery_method_id": self.delivary_method_id,
                "departure_date": self.departure_date
            })
            response = requests.request(
                "POST", url, headers=ozon_headers_karavaev, data=payload)
            self.delivery_id = json.loads(response.text)['result']['id']

    def check_delivery_create(self):
        """
        Функция проверяет, что отгрузка создана.
        Формирует список отправлений для дальнейшей работы
        """
        url = 'https://api-seller.ozon.ru/v2/posting/fbs/act/check-status'

        payload = json.dumps(
            {
                "id": self.delivery_id
                # "id": 35178630
            }
        )
        response = requests.request(
            "POST", url, headers=ozon_headers_karavaev, data=payload)

        data = json.loads(response.text)['result']
        if data['status'] == "ready":

            self.delivery_number_list = data["added_to_act"]
            print('Функция check_delivery_create сработала. Спать 5 мин не нужно')
        else:
            print('уснули на 5 мин в функции check_delivery_create')
            time.sleep(300)
            self.check_delivery_create()

    def receive_barcode_delivery(self):
        """Получает и сохраняет штрихкод поставки"""
        url = 'https://api-seller.ozon.ru/v2/posting/fbs/act/get-barcode'
        payload = json.dumps(
            {
                "id": self.delivery_id
                # "id": 35275670
            }
        )
        # self.delivery_id = 35275670
        response = requests.request(
            "POST", url, headers=ozon_headers_karavaev, data=payload)

        image = Image.open(io.BytesIO(response.content))
        save_folder_docs = f"{self.main_save_folder_server}/ozon_delivery_barcode/{self.delivery_id}_баркод {self.date_for_files}.png"
        image.save(save_folder_docs)

        # Сохраняем штрихкод в PDF формате
        now_date = datetime.now().strftime(("%d.%m"))
        im = Image.open(save_folder_docs)
        text_common = "Штрихкод для отгрузки ОЗОН"
        text_company = 'ИП Караваев'
        font = ImageFont.truetype("arial.ttf", size=50)
        A4 = (1033, 1462)
        white_A4 = Image.new('RGB', A4, 'white')
        text_draw = ImageDraw.Draw(white_A4)
        text_common_lenght = text_draw.textlength(text_common, font=font)
        text_company_lenght = text_draw.textlength(text_company, font=font)
        x = (A4[0] - im.width) // 2
        text_common_x = (A4[0] - round(text_common_lenght)) // 2
        text_company_x = (A4[0] - round(text_company_lenght)) // 2
        white_A4.paste(im, (x, 330))
        text_draw.text((text_common_x, 100), text_common,
                       font=font, fill=('#000000'))
        text_draw.text((text_company_x, 200), text_company,
                       font=font, fill=('#000000'))
        save_folder_docs_pdf = f"{self.main_save_folder_server}/ozon_delivery_barcode/{self.delivery_id}_баркод {now_date}.pdf"
        white_A4.save(save_folder_docs_pdf, 'PDF', quality=100)

        folder = (
            f'{self.dropbox_current_assembling_folder}/OZON-ИП акт {now_date}.pdf')
        with open(save_folder_docs_pdf, 'rb') as f:
            dbx_db.files_upload(f.read(), folder)

    def check_status_formed_invoice(self):
        """
        Проверяет статус формирования накладной: /v2/posting/fbs/digital/act/check-status. 
        Когда статус документа перейдёт в FORMED или CONFIRMED, получите файлы:
        Получает и сохраняет этикетки для каждой коробки: /v2/posting/fbs/act/get-container-labels.
        Этикетки для каждой отправки: /v2/posting/fbs/package-label.
        """

        url = 'https://api-seller.ozon.ru/v2/posting/fbs/digital/act/check-status'
        payload = json.dumps(
            {
                "id": self.delivery_id
                # "id": 35178630
            }
        )
        response = requests.request(
            "POST", url, headers=ozon_headers_karavaev, data=payload)

        data = json.loads(response.text)['status']

        if data == "CONFIRMED" or data == "FORMED":
            print('Функция check_status_formed_invoice сработала. Спать 5 мин не нужно')
            self.receive_barcode_delivery()
            self.get_box_tickets()
            self.forming_package_ticket_with_article()
        else:
            print('уснули на 5 мин в функции check_status_formed_invoice')
            time.sleep(300)
            self.check_status_formed_invoice()

    def get_box_tickets(self):
        """
        Функция получает и сохраняет этикетки для каждой коробки в формате PDF.
        Данные получают из эндпоинта /v2/posting/fbs/act/get-container-labels.
        """
        url = 'https://api-seller.ozon.ru/v2/posting/fbs/act/get-container-labels'

        payload = json.dumps(
            {
                "id": self.delivery_id
                # "id": 35275670
            }
        )
        response = requests.request(
            "POST", url, headers=ozon_headers_karavaev, data=payload)

        # получение данных PDF из входящих данных
        pdf_data = response.content  # замените на фактические входные данные
        save_folder_docs = f'{self.main_save_folder_server}/ozon_docs/OZON - ИП ШК на 1 коробку {self.date_for_files}.pdf'
        # сохранение PDF-файла
        with open(save_folder_docs, 'wb') as f:
            f.write(pdf_data)
        folder = (
            f'{self.dropbox_current_assembling_folder}/OZON-ИП ШК на 1 коробку {self.date_for_files}.pdf')
        with open(save_folder_docs, 'rb') as f:
            dbx_db.files_upload(f.read(), folder)

    def forming_package_ticket_with_article(self):
        """
        Функция получает и сохраняет этикетки для каждой отправки.
        Данные получают из эндпоинта /v2/posting/fbs/package-label.
        После получения и сохранения всех этикеток помещает артикул продавца
        на этикетку.
        """
        url = 'https://api-seller.ozon.ru/v2/posting/fbs/package-label'
        for package in self.fbs_ozon_common_data.keys():
            payload = json.dumps(
                {
                    "posting_number": [package]
                }
            )
            response = requests.request(
                "POST", url, headers=ozon_headers_karavaev, data=payload)
            save_folder = f'{self.main_save_folder_server}/package_tickets'
            save_folder_docs = f'{self.main_save_folder_server}/package_tickets/{package}.pdf'
            # сохранение PDF-файла
            with open(save_folder_docs, 'wb') as f:
                f.write(response.content)

        done_files_folder = f'{self.main_save_folder_server}/package_tickets/done'
        if not os.path.isdir(done_files_folder):
            os.mkdir(done_files_folder)

        # Записываем артикул продавциа в файл с этикеткой
        new_data_for_ozon_ticket(save_folder, self.fbs_ozon_common_data)
        # Формируем список всех файлов, которые нужно объединять в один документ
        list_filename = glob.glob(f'{done_files_folder}/*.pdf')
        # Адрес и имя конечного файла
        file_name = f'{self.main_save_folder_server}/package_tickets/done/done_tickets.pdf'
        merge_barcode_for_ozon_two_on_two(list_filename, file_name)
        folder = f'{self.dropbox_current_assembling_folder}/OZON-ИП этикетки.pdf'
        with open(file_name, 'rb') as f:
            dbx_db.files_upload(f.read(), folder)


class CreatePivotFile(WildberriesFbsMode, OzonFbsMode):
    def __init__(self):
        self.amount_articles = WildberriesFbsMode().article_data_for_tickets()
        self.ozon_article_amount = OzonFbsMode().prepare_data_for_confirm_delivery()
        self.dropbox_main_fbs_folder = '/DATABASE/beta'

    def create_pivot_xls(self):
        '''Создает сводный файл excel с количеством каждого артикула.
        Подключается к базе данных на сервере'''
        # Задаем словарь с данными WB, а входящий становится общим для всех маркетплейсов
        wb_article_amount = self.amount_articles.copy()
        hour = datetime.now().hour
        date_folder = datetime.today().strftime('%Y-%m-%d')
        if hour >= 18 or hour <= 6:
            self.delivary_method_id = 22655170176000
            self.dropbox_current_assembling_folder = f'{self.dropbox_main_fbs_folder}/НОЧЬ СБОРКА ФБС/{date_folder}'

        else:
            self.delivary_method_id = 1020000089903000
            self.dropbox_current_assembling_folder = f'{self.dropbox_main_fbs_folder}/ДЕНЬ СБОРКА ФБС/{date_folder}'
        CELL_LIMIT = 16  # ограничение символов в ячейке Excel
        COUNT_HELPER = 2

        for article in self.ozon_article_amount.keys():
            if article in self.amount_articles.keys():
                self.amount_articles[article] = int(
                    self.amount_articles[article]) + int(self.ozon_article_amount[article])
            else:
                self.amount_articles[article] = int(
                    self.ozon_article_amount[article])

        sorted_data_for_pivot_xls = dict(
            sorted(self.amount_articles.items(), key=lambda v: v[0].upper()))
        pivot_xls = openpyxl.Workbook()
        create = pivot_xls.create_sheet(title='pivot_list', index=0)
        sheet = pivot_xls['pivot_list']
        sheet['A1'] = 'Артикул продавца'
        sheet['B1'] = 'Ячейка стеллажа'
        # Можно вернуть столбец для производства
        # sheet['C1'] = 'На производство'
        sheet['D1'] = 'Всего для FBS'
        sheet['E1'] = 'FBS WB'
        sheet['F1'] = 'FBS Ozon'
        # ========== РАСКРЫТЬ КОГДА ПОЯВИТСЯ ЯНДЕКС МАРКЕТ ========= #
        # sheet['G1'] = 'FBY Yandex'

        for key, value in sorted_data_for_pivot_xls.items():
            create.cell(row=COUNT_HELPER, column=1).value = key
            create.cell(row=COUNT_HELPER, column=4).value = value
            COUNT_HELPER += 1
        name_pivot_xls = 'web_barcode/fbs_mode/data_for_barcodes/pivot_excel/На производство.xlsx'
        path_file = os.path.abspath(name_pivot_xls)
        # file_name_dir = path.parent

        pivot_xls.save(name_pivot_xls)
        # ========= Подключение к базе данных ========= #
        engine = create_engine(
            "postgresql://databaseadmin:Up3psv8x@158.160.28.219:5432/innotreid")

        data = pd.read_sql_table(
            "database_shelvingstocks",
            con=engine,
            columns=['task_start_date',
                     'task_finish_date',
                     'seller_article_wb',
                     'seller_article',
                     'shelf_number',
                     'amount']
        )
        connection = psycopg2.connect(user="databaseadmin",
                                      # пароль, который указали при установке PostgreSQL
                                      password="Up3psv8x",
                                      host="158.160.28.219",
                                      port="5432",
                                      database="innotreid")
        cursor = connection.cursor()

        shelf_seller_article_list = data['seller_article_wb'].to_list()
        shelf_number_list = data['shelf_number'].to_list()
        shelf_amount_list = data['amount'].to_list()
        w_b = load_workbook(name_pivot_xls)
        source_page = w_b.active
        name_article = source_page['A']
        amount_all_fbs = source_page['D']

        for i in range(1, len(name_article)):
            for j in self.ozon_article_amount.keys():
                if name_article[i].value == j:
                    source_page.cell(
                        row=i+1, column=6).value = self.ozon_article_amount[j]
            # ========== РАСКРЫТЬ КОГДА ПОЯВИТСЯ ЯНДЕКС МАРКЕТ ========= #
            # for k in range(len(name_article_ya)):
            #     if name_article[i].value == name_article_ya[k]:
            #         source_page.cell(row=i+1,column=7).value = amount_article_ya[k]

            if name_article[i].value in wb_article_amount.keys():
                source_page.cell(
                    row=i+1, column=5).value = self.amount_articles[name_article[i].value]
            # Заполняется столбец "В" - номер ячейки на внутреннем складе
            for s in range(len(shelf_seller_article_list)):
                if name_article[i].value == shelf_seller_article_list[s] and (
                        int(amount_all_fbs[i].value) < int(shelf_amount_list[s])):
                    source_page.cell(
                        row=i+1, column=2).value = shelf_number_list[s]
                    new_amount = int(
                        shelf_amount_list[s]) - int(amount_all_fbs[i].value)
                    select_table_query = f'''UPDATE database_shelvingstocks SET amount={new_amount},
                        task_start_date=current_timestamp, task_finish_date=NULL WHERE seller_article='{shelf_seller_article_list[s]}';'''
                    cursor.execute(select_table_query)
                elif name_article[i].value == shelf_seller_article_list[s] and (
                        int(amount_all_fbs[i].value) >= int(shelf_amount_list[s])):
                    # ========== Сюда вставить отметку, если мало артикулов в полке =========
                    source_page.cell(
                        row=i+1, column=2).value = f'{shelf_number_list[s]}'
        connection.commit()
        w_b.save(name_pivot_xls)
        w_b2 = load_workbook(name_pivot_xls)
        source_page2 = w_b2.active
        amount_all_fbs = source_page2['D']
        amount_for_production = source_page2['C']
        PROD_DETAIL_CONST = 4
        for r in range(1, len(amount_all_fbs)):
            # Заполняет столбец ['C'] = 'Производство'
            if amount_all_fbs[r].value == 1:
                source_page2.cell(
                    row=r+1, column=3).value = int(PROD_DETAIL_CONST)
            elif 2 <= int(amount_all_fbs[r].value) <= PROD_DETAIL_CONST-1:
                source_page2.cell(
                    row=r+1, column=3).value = int(2 * PROD_DETAIL_CONST)
            elif PROD_DETAIL_CONST <= int(amount_all_fbs[r].value) <= 2 * PROD_DETAIL_CONST - 1:
                source_page2.cell(
                    row=r+1, column=3).value = int(3 * PROD_DETAIL_CONST)
            else:
                source_page2.cell(row=r+1, column=3).value = ' '
        w_b2.save(name_pivot_xls)
        w_b2 = load_workbook(name_pivot_xls)
        source_page2 = w_b2.active
        amount_all_fbs = source_page2['D']
        al = Alignment(horizontal="center", vertical="center")
        al_left = Alignment(horizontal="left", vertical="center")
        # Задаем толщину и цвет обводки ячейки
        font_bold = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="medium", color="000000")
        for i in range(len(amount_all_fbs)):
            for c in source_page2[f'A{i+1}:G{i+1}']:
                if i == 0:
                    c[0].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[1].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[2].border = Border(top=thick, left=thick,
                                         bottom=thin, right=thick)
                    c[3].border = Border(top=thick, left=thick,
                                         bottom=thin, right=thick)
                    c[4].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[5].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[6].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                elif i == len(amount_all_fbs)-1:
                    c[0].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[1].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[2].border = Border(top=thin, left=thick,
                                         bottom=thick, right=thick)
                    c[3].border = Border(top=thin, left=thick,
                                         bottom=thick, right=thick)
                    c[4].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[5].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[6].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                else:
                    c[0].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[1].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[2].border = Border(top=thin, left=thick,
                                         bottom=thin, right=thick)
                    c[3].border = Border(top=thin, left=thick,
                                         bottom=thin, right=thick)
                    c[4].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[5].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[6].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                c[0].alignment = al_left
                c[1].alignment = al
                c[2].alignment = al
                c[3].alignment = al
                c[4].alignment = al
                c[5].alignment = al
                c[6].alignment = al
        source_page2.column_dimensions['A'].width = 18
        source_page2.column_dimensions['B'].width = 18
        source_page2.column_dimensions['C'].width = 18
        source_page2.column_dimensions['D'].width = 10
        source_page2.column_dimensions['E'].width = 10
        source_page2.column_dimensions['F'].width = 12
        source_page2.column_dimensions['G'].width = 12

        # Когда понадобится столбец НА ПРОИЗВОДСТВО - удалить следующую строку
        source_page2.delete_cols(3, 1)
        w_b2.save(name_pivot_xls)
        xl = DispatchEx("Excel.Application")
        xl.DisplayAlerts = False
        # print(f'{file_name_dir}/На производство.xlsx')
        folder_path = os.path.dirname(os.path.abspath(path_file))
        name_for_file = f'Общий файл производство ИП {delivery_date}'
        name_xls_dropbox = f'На производство ИП {delivery_date}'
        wb = xl.Workbooks.Open(path_file)
        xl.CalculateFull()
        pythoncom.PumpWaitingMessages()
        try:
            wb.ExportAsFixedFormat(
                0, f'{folder_path}/{name_for_file}.pdf')
        except Exception as e:
            print("Failed to convert in PDF format.Please confirm environment meets all the requirements  and try again")
        finally:
            wb.Close()

        # Сохраняем на DROPBOX
        with open(f'{path_file}', 'rb') as f:
            dbx_db.files_upload(
                f.read(), f'{self.dropbox_current_assembling_folder}/{name_xls_dropbox}.xlsx')
        with open(f'{folder_path}/{name_for_file}.pdf', 'rb') as f:
            dbx_db.files_upload(
                f.read(), f'{self.dropbox_current_assembling_folder}/{name_for_file}.pdf')


# ========== ВЫЗЫВАЕМ ФУНКЦИИ ПООЧЕРЕДИ ========== #

def common_action():
    wb_actions = WildberriesFbsMode()
    ozon_actions = OzonFbsMode()

    clearning_folders()
    # =========== АЛГОРИТМ  ДЕЙСТВИЙ С WILDBERRIES ========== #
    # 1. Обрабатываю новые сборочные задания.
    # wb_actions.article_data_for_tickets()

    # 2. Создаю поставку
    # wb_actions.create_delivery()

    # 3. добавляю сборочные задания по их id в созданную поставку и получаю qr стикер каждого
    # задания и сохраняю его в папку
    # wb_actions.qrcode_order()

    # 4. Создаю лист сборки
    # wb_actions.create_selection_list()

    # 5. Добавляю поставку в доставку, получаю QR код поставки
    # и преобразует этот QR код в необходимый формат.
    # wb_actions.qrcode_supply()

    # 6. Создаю список с полными именами файлов, которые нужно объединить
    # wb_actions.list_for_print_create()

    # =========== АЛГОРИТМ  ДЕЙСТВИЙ С ОЗОН ========== #
    # 1. Собираю информацию о новых заказах с Озон.
    ozon_actions.awaiting_packaging_orders()

    # # 2. Делю заказ на отправления и перевожу его в статус awaiting_deliver.
    ozon_actions.awaiting_deliver_orders()

    # 3. Готовлю данные для подтверждения отгрузки
    ozon_actions.prepare_data_for_confirm_delivery()

    # # 4. Подтверждаю отгрузку и запускаю создание документов на стороне ОЗОН
    ozon_actions.confirm_delivery_create_document()

    # # 5. Проверяю, что отгрузка создана. Формирую список отправлений для дальнейшей работы
    ozon_actions.check_delivery_create()

    # # 6. Проверяю статус формирования накладной.
    # # Получаю файлы с этикетками для коробок и этикетки для каждой отправки
    ozon_actions.check_status_formed_invoice()

    # =========== СОЗДАЮ СВОДНЫЙ ФАЙЛ ========== #
    # 1. Создаю сводный файл для производства
    pivot_file = CreatePivotFile()
    pivot_file.create_pivot_xls()

    # Очищаем все папки на сервере

    clearning_folders()


common_action()
