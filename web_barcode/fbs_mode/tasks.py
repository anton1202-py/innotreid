import base64
import glob
import io
import json
import logging
import math
import os
import shutil
import textwrap
import time
import traceback
from collections import Counter
from contextlib import closing
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

import dropbox
import openpyxl
import pandas as pd
import psycopg2
import requests
import telegram
from celery_tasks.celery import app
from dotenv import load_dotenv
from msoffice2pdf import convert
from openpyxl import Workbook, load_workbook
from openpyxl.drawing import image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont
from sqlalchemy import create_engine

from .helpers import (design_barcodes_dict_spec,
                      merge_barcode_for_ozon_two_on_two,
                      merge_barcode_for_yandex_two_on_two,
                      new_data_for_ozon_ticket, new_data_for_yandex_ticket,
                      print_barcode_to_pdf2, qrcode_print_for_products,
                      supply_qrcode_to_standart_view)

# Загрузка переменных окружения из файла .env
dotenv_path = os.path.join(os.path.dirname(
    __file__), '..', 'web_barcode', '.env')
load_dotenv(dotenv_path)

REFRESH_TOKEN_DB = os.getenv('REFRESH_TOKEN_DB')
APP_KEY_DB = os.getenv('APP_KEY_DB')
APP_SECRET_DB = os.getenv('APP_SECRET_DB')

API_KEY_WB_IP = os.getenv('API_KEY_WB_IP')
YANDEX_IP_KEY = os.getenv('YANDEX_IP_KEY')
API_KEY_OZON_KARAVAEV = os.getenv('API_KEY_OZON_KARAVAEV')
CLIENT_ID_OZON_KARAVAEV = os.getenv('CLIENT_ID_OZON_KARAVAEV')

OZON_OOO_API_TOKEN = os.getenv('OZON_OOO_API_TOKEN')
OZON_OOO_CLIENT_ID = os.getenv('OZON_OOO_CLIENT_ID')
YANDEX_OOO_KEY = os.getenv('YANDEX_OOO_KEY')
WB_OOO_API_KEY = os.getenv('WB_OOO_API_KEY')


TELEGRAM_TOKEN = os.getenv('TELEGRAM_TOKEN')
CHAT_ID_ADMIN = os.getenv('CHAT_ID_ADMIN')
CHAT_ID_MANAGER = os.getenv('CHAT_ID_MANAGER')
CHAT_ID_EU = os.getenv('CHAT_ID_EU')
CHAT_ID_AN = os.getenv('CHAT_ID_AN')

bot = telegram.Bot(token=TELEGRAM_TOKEN)

wb_headers_karavaev = {
    'Content-Type': 'application/json',
    'Authorization': API_KEY_WB_IP
}
wb_headers_ooo = {
    'Content-Type': 'application/json',
    'Authorization': WB_OOO_API_KEY
}

ozon_headers_karavaev = {
    'Api-Key': API_KEY_OZON_KARAVAEV,
    'Content-Type': 'application/json',
    'Client-Id': CLIENT_ID_OZON_KARAVAEV
}
ozon_headers_ooo = {
    'Api-Key': OZON_OOO_API_TOKEN,
    'Content-Type': 'application/json',
    'Client-Id': OZON_OOO_CLIENT_ID
}

yandex_headers_karavaev = {
    'Authorization': YANDEX_IP_KEY,
}
yandex_headers_ooo = {
    'Authorization': YANDEX_OOO_KEY,
}

db_folder = '/!На производство'

file_add_name_ip = 'ИП'
file_add_name_ooo = 'ООО'

dbx_db = dropbox.Dropbox(oauth2_refresh_token=REFRESH_TOKEN_DB,
                         app_key=APP_KEY_DB,
                         app_secret=APP_SECRET_DB)

logging.basicConfig(level=logging.DEBUG,
                    filename="tasks_log.log",
                    format="%(asctime)s %(levelname)s %(message)s")


def stream_dropbox_file(path):
    _, res = dbx_db.files_download(path)
    with closing(res) as result:
        byte_data = result.content
        return io.BytesIO(byte_data)


def clearning_folders():
    dirs = ['fbs_mode/data_for_barcodes/cache_dir',
            'fbs_mode/data_for_barcodes/done_data',
            'fbs_mode/data_for_barcodes/pivot_excel',
            'fbs_mode/data_for_barcodes/qrcode_folder',
            'fbs_mode/data_for_barcodes/qrcode_supply',
            'fbs_mode/data_for_barcodes/package_tickets',
            'fbs_mode/data_for_barcodes/ozon_docs',
            'fbs_mode/data_for_barcodes/ozon_delivery_barcode',
            'fbs_mode/data_for_barcodes/yandex',
            'fbs_mode/data_for_barcodes/new_pivot_excel'
            ]
    for dir in dirs:
        for filename in glob.glob(os.path.join(dir, "*")):
            file_path = os.path.join(dir, filename)
            try:
                if os.path.isfile(filename) or os.path.islink(filename):
                    os.unlink(filename)
                elif os.path.isdir(filename):
                    shutil.rmtree(filename)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (filename, e))


def wb_clearning_folders():
    dirs = ['fbs_mode/data_for_barcodes/cache_dir',
            'fbs_mode/data_for_barcodes/done_data',
            'fbs_mode/data_for_barcodes/pivot_excel',
            'fbs_mode/data_for_barcodes/qrcode_folder',
            'fbs_mode/data_for_barcodes/qrcode_supply'
            ]
    for dir in dirs:
        for filename in glob.glob(os.path.join(dir, "*")):
            file_path = os.path.join(dir, filename)
            try:
                if os.path.isfile(filename) or os.path.islink(filename):
                    os.unlink(filename)
                elif os.path.isdir(filename):
                    shutil.rmtree(filename)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (filename, e))


def error_message(function_name: str, function, error_text: str) -> str:
    """
    Формирует текст ошибки и выводит информацию в модальном окне
    function_name - название функции.
    function - сама функция, как объект, а не результат работы. Без скобок.
    error_text - текст ошибки.
    """
    tb_str = traceback.format_exc()
    message_error = (f'Ошибка в функции: <b>{function_name}</b>\n'
                     f'<b>Функция выполняет</b>: {function.__doc__}\n'
                     f'<b>Ошибка</b>\n: {error_text}\n\n'
                     f'<b>Техническая информация</b>:\n {tb_str}')
    return message_error


class WildberriesFbsMode():
    """Класс отвечает за работу с заказами Wildberries"""

    def __init__(self, headers, db_forder, file_add_name):
        """Основные данные класса"""
        self.amount_articles = {}
        self.dropbox_main_fbs_folder = db_forder
        self.headers = headers
        self.file_add_name = file_add_name

    def check_folder_exists(self, path):
        try:
            dbx_db.files_list_folder(path)
            return True
        except dropbox.exceptions.ApiError as e:
            return False

    def create_dropbox_folder(self):
        """
        WILDBERRIES
        Создает папку для созданных документов на Дропбоксе.
        """
        hour = datetime.now().hour
        date_folder = datetime.today().strftime('%Y-%m-%d')

        if hour >= 6 and hour < 18:
            self.dropbox_current_assembling_folder = f'{self.dropbox_main_fbs_folder}/ДЕНЬ СБОРКА ФБС/{date_folder}'
        else:
            self.dropbox_current_assembling_folder = f'{self.dropbox_main_fbs_folder}/НОЧЬ СБОРКА ФБС/{date_folder}'
        # Создаем папку на dropbox, если ее еще нет
        if self.check_folder_exists(self.dropbox_current_assembling_folder) == False:
            dbx_db.files_create_folder_v2(
                self.dropbox_current_assembling_folder)

    def process_new_orders(self):
        """
        WILDBERRIES
        Функция обрабатывает новые сборочные задания.
        """
        try:
            url = "https://suppliers-api.wildberries.ru/api/v3/orders/new"
            response = requests.request(
                "GET", url, headers=self.headers)
            orders_data = json.loads(response.text)['orders']
            if response.status_code == 200:
                return orders_data
            else:
                time.sleep(10)
                return self.process_new_orders()
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'process_new_orders', self.process_new_orders, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def time_filter_orders(self):
        """
        WILDBERRIES
        Функция фильтрует новые заказы по времени.
        Если заказ создан меньше, чем 1 час назад, в работу он не берется
        """
        try:
            orders_data = self.process_new_orders()
            filters_order_data = []
            now_time = datetime.now()
            for order in orders_data:
                # Время создания заказа в переводе с UTC на московское
                create_order_time = datetime.strptime(
                    order['createdAt'], '%Y-%m-%dT%H:%M:%SZ') + timedelta(hours=3)
                delta_order_time = now_time - create_order_time
                if delta_order_time > timedelta(hours=1):
                    filters_order_data.append(order)
            return filters_order_data
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'time_filter_orders', self.time_filter_orders, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def article_info(self, article):
        """
        WILDBERRIES
        Получает развернутый ответ про каждый артикул. 
        """
        try:
            url_data = "https://suppliers-api.wildberries.ru/content/v2/get/cards/list"
            payload = json.dumps({
                "settings": {
                    "cursor": {
                        "limit": 1
                    },
                    "filter": {
                        "textSearch": article,
                        "withPhoto": -1
                    }
                }
            })
            response_data = requests.request(
                "POST", url_data, headers=self.headers, data=payload)
            if response_data.status_code == 200:
                return response_data.text
            else:
                text = f'Статус код = {response_data.status_code} у артикула {article}'
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=text, parse_mode='HTML')
                time.sleep(5)
                return self.article_info(article)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'article_info', self.article_info, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def article_data_for_tickets(self):
        """
        WILDBERRIES
        Выделяет артикулы продавца светильников, их баркоды и наименования.
        Создает словарь с данными каждого артикулы и словарь с количеством каждого
        артикула. 
        """
        try:
            orders_data = self.time_filter_orders()
            # Словарь с данными артикула: {артикул_продавца: [баркод, наименование]}
            self.data_article_info_dict = {}
            # Список только для артикулов ночников. Остальные отфильтровывает
            self.clear_article_list = []
            # Словарь для данных листа подбора {order_id: [photo_link, brand, name, seller_article]}
            self.selection_dict = {}
            for data in orders_data:
                answer = self.article_info(data['article'])

                if json.loads(answer)['cards'][0]['subjectName'] == "Ночники":
                    self.clear_article_list.append(data['article'])
                    # Достаем баркод артикула (первый из списка, если их несколько)
                    barcode = json.loads(answer)[
                        'cards'][0]['sizes'][0]['skus'][0]
                    # Достаем название артикула
                    title = json.loads(answer)[
                        'cards'][0]['title']
                    self.data_article_info_dict[data['article']] = [
                        title, barcode]
                    photo = json.loads(answer)[
                        'cards'][0]['photos'][0]['big']
                    brand = json.loads(answer)[
                        'cards'][0]['brand']
                    title_article = json.loads(answer)[
                        'cards'][0]['title']
                    seller_article = data['article']
                    # Заполняем словарь данными для Листа подбора
                    self.selection_dict[data['id']] = [
                        photo, brand, title_article, seller_article]
                time.sleep(2)
            # Словарь с данными: {артикул_продавца: количество}
            self.amount_articles = dict(Counter(self.clear_article_list))
            return self.amount_articles
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'article_data_for_tickets', self.article_data_for_tickets, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def create_barcode_tickets(self):
        """WILDBERRIES. Функция создает этикетки со штрихкодами для артикулов"""
        try:
            if self.clear_article_list and self.data_article_info_dict:
                design_barcodes_dict_spec(
                    self.clear_article_list, self.data_article_info_dict)
            else:
                text = 'не сработала create_barcode_tickets так как нет данных'
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=text, parse_mode='HTML')
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'create_barcode_tickets', self.create_barcode_tickets, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def create_delivery(self):
        """WILDBERRIES. Создание поставки"""
        try:
            amount_articles = self.article_data_for_tickets()
            if amount_articles:
                delivery_date = datetime.today().strftime("%d.%m.%Y %H-%M-%S")
                url_data = 'https://suppliers-api.wildberries.ru/api/v3/supplies'
                hour = datetime.now().hour
                delivery_name = f"Поставка {delivery_date}"
                if hour >= 6 and hour < 18:
                    delivery_name = f'День {delivery_date}'
                else:
                    delivery_name = f'Ночь {delivery_date}'

                payload = json.dumps(
                    {"name": delivery_name}
                )
                # Из этой переменной достать ID поставки
                response_data = requests.request(
                    "POST", url_data, headers=self.headers, data=payload)
                # print(response_data)
                self.supply_id = json.loads(response_data.text)['id']
            else:
                text = f'Нет артикулов на WB для сборки {self.file_add_name}'
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=text, parse_mode='HTML')
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'create_delivery', self.create_delivery, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def qrcode_order(self):
        """
        WILDBERRIES.
        Функция добавляет сборочные задания по их id
        в созданную поставку и получает qr стикер каждого
        задания и сохраняет его в папку
        """
        try:
            if self.supply_id:
                # Добавляем заказы в поставку
                for order in self.selection_dict.keys():
                    add_url = f'https://suppliers-api.wildberries.ru/api/v3/supplies/{self.supply_id}/orders/{order}'
                    response_add_orders = requests.request(
                        "PATCH", add_url, headers=self.headers)

                # Создаем qr коды добавленных ордеров.
                for order in self.selection_dict.keys():
                    ticket_url = 'https://suppliers-api.wildberries.ru/api/v3/orders/stickers?type=png&width=58&height=40'
                    payload_ticket = json.dumps({"orders": [order]})
                    response_ticket = requests.request(
                        "POST", ticket_url, headers=self.headers, data=payload_ticket)

                    # Расшифровываю ответ, чтобы сохранить файл этикетки задания
                    ticket_data = json.loads(response_ticket.text)[
                        "stickers"][0]["file"]

                    # Узнаю стикер сборочного задания и помещаю его в словарь с данными для
                    # листа подбора
                    sticker_code_first_part = json.loads(response_ticket.text)[
                        "stickers"][0]["partA"]
                    sticker_code_second_part = json.loads(response_ticket.text)[
                        "stickers"][0]["partB"]
                    sticker_code = f'{sticker_code_first_part} {sticker_code_second_part}'
                    self.selection_dict[order].append(sticker_code)
                    # декодируем строку из base64 в бинарные данные
                    binary_data = base64.b64decode(ticket_data)
                    # создаем объект изображения из бинарных данных
                    img = Image.open(BytesIO(binary_data))
                    # сохраняем изображение в файл
                    folder_path = os.path.join(
                        os.getcwd(), "fbs_mode/data_for_barcodes/qrcode_folder")
                    if not os.path.exists(folder_path):
                        os.makedirs(folder_path)
                    img.save(
                        f"{folder_path}/{order} {self.selection_dict[order][-1]}.png")
            else:
                text = 'не сработала qrcode_order из за отсутвия self.supply_id'
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=text, parse_mode='HTML')
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message('qrcode_order', self.qrcode_order, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def create_selection_list(self):
        """WILDBERRIES. Создает лист сборки"""
        try:
            if self.selection_dict:
                delivery_date = datetime.today().strftime("%d.%m.%Y %H-%M-%S")
                # создаем новую книгу Excel
                selection_file = Workbook()
                COUNT_HELPER = 2
                # выбираем лист Sheet1
                create = selection_file.create_sheet(
                    title='pivot_list', index=0)
                sheet = selection_file['pivot_list']

                # Установка параметров печати
                create.page_setup.paperSize = create.PAPERSIZE_A4
                create.page_setup.orientation = create.ORIENTATION_PORTRAIT
                create.page_margins.left = 0.25
                create.page_margins.right = 0.25
                create.page_margins.top = 0.25
                create.page_margins.bottom = 0.25
                create.page_margins.header = 0.3
                create.page_margins.footer = 0.3

                sheet['A1'] = '№ Задания'
                sheet['B1'] = 'Фото'
                sheet['C1'] = 'Бренд'
                sheet['D1'] = 'Наименование'
                sheet['E1'] = 'Артикул продавца'
                sheet['F1'] = 'Стикер'

                for key, value in self.selection_dict.items():
                    # # загружаем изображение
                    response = requests.get(value[0])
                    img = image.Image(io.BytesIO(response.content))
                    # задаем размеры изображения
                    img.width = 30
                    img.height = 50

                    create.cell(row=COUNT_HELPER, column=1).value = key
                    # вставляем изображение в Столбец В
                    sheet.add_image(img, f'B{COUNT_HELPER}')
                    # create.cell(row=COUNT_HELPER, column=2).value = value[0]
                    create.cell(row=COUNT_HELPER, column=3).value = value[1]
                    create.cell(row=COUNT_HELPER, column=4).value = value[2]
                    create.cell(row=COUNT_HELPER, column=5).value = value[3]
                    create.cell(row=COUNT_HELPER, column=6).value = value[4]
                    COUNT_HELPER += 1
                folder_path = os.path.join(
                    os.getcwd(), 'fbs_mode/data_for_barcodes/pivot_excel')
                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)
                name_selection_file = f'{folder_path}/selection_list.xlsx'
                path_file = os.path.abspath(name_selection_file)

                selection_file.save(name_selection_file)

                w_b2 = load_workbook(name_selection_file)
                source_page2 = w_b2.active

                al = Alignment(horizontal="center", vertical="center")
                al_left = Alignment(horizontal="left",
                                    vertical="center", wrapText=True)

                source_page2.column_dimensions['A'].width = 10  # Номер задания
                source_page2.column_dimensions['B'].width = 5  # Картинка
                source_page2.column_dimensions['C'].width = 15  # Бренд
                source_page2.column_dimensions['D'].width = 25  # Наименование
                # Артикул продавца
                source_page2.column_dimensions['E'].width = 16
                source_page2.column_dimensions['F'].width = 16  # Стикер

                thin = Side(border_style="thin", color="000000")
                for i in range(len(self.selection_dict)+1):
                    for c in source_page2[f'A{i+1}:F{i+1}']:
                        c[0].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[0].font = Font(size=9)
                        c[0].alignment = al

                        c[1].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[1].font = Font(size=9)
                        c[1].alignment = al

                        c[2].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[2].font = Font(size=9)
                        c[2].alignment = al_left

                        c[3].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[3].font = Font(size=9)
                        c[3].alignment = al_left

                        c[4].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[4].font = Font(size=9)
                        c[4].alignment = al_left

                        c[5].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[5].font = Font(size=9)
                        c[5].alignment = al_left

                # Увеличиваем высоту строки
                for i in range(2, len(self.selection_dict) + 2):
                    source_page2.row_dimensions[i].height = 40
                w_b2.save(name_selection_file)
                folder_path = os.path.dirname(os.path.abspath(path_file))
                name_for_file = f'WB - {self.file_add_name} лист подбора {delivery_date}'
                name_xls_dropbox = f'WB - {self.file_add_name} Лист подбора {delivery_date}'

                output = convert(source=path_file,
                                 output_dir=folder_path, soft=1)
                # Сохраняем на DROPBOX
                with open(output, 'rb') as f:
                    dbx_db.files_upload(
                        f.read(), f'{self.dropbox_current_assembling_folder}/{name_for_file}.pdf')
            else:
                text = 'Не сработала create_selection_list потому что нет self.selection_dict'
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=text, parse_mode='HTML')
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'create_selection_list', self.create_selection_list, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def supply_to_delivery(self, numb=0):
        """
        WILDBERRIES.
        Функция добавляет поставку в доставку.
        """
        try:
            if self.supply_id:
                time.sleep(30)
                # Переводим поставку в доставку
                url_to_supply = f'https://suppliers-api.wildberries.ru/api/v3/supplies/{self.supply_id}/deliver'
                response_to_supply = requests.request(
                    "PATCH", url_to_supply, headers=self.headers)

                if response_to_supply.status_code != 204 and numb < 10:
                    numb += 1
                    self.supply_to_delivery(numb)
                elif response_to_supply.status_code != 204 and numb >= 10:
                    text = 'Не получилось перевести поставку в доставку, поэтому не будет QR-кода'
                    bot.send_message(chat_id=CHAT_ID_ADMIN,
                                     text=text, parse_mode='HTML')
                elif response_to_supply.status_code == 204:
                    text = 'Поставку перевел в доставку'
                    bot.send_message(chat_id=CHAT_ID_ADMIN,
                                     text=text, parse_mode='HTML')
            else:
                text = 'Поставка не добавлена в доставку (supply_to_delivery), так как нет артикулов'
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=text, parse_mode='HTML')
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'supply_to_delivery', self.supply_to_delivery, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def qrcode_supply(self):
        """
        WILDBERRIES.
        Функция получает QR код поставки
        и преобразует этот QR код в необходимый формат.
        """
        try:
            if self.supply_id:
                time.sleep(60)
                # Получаем QR код поставки:
                url_supply_qrcode = f"https://suppliers-api.wildberries.ru/api/v3/supplies/{self.supply_id}/barcode?type=png"
                response_supply_qrcode = requests.request(
                    "GET", url_supply_qrcode, headers=self.headers)

                # Создаем QR код поставки
                qrcode_base64_data = json.loads(
                    response_supply_qrcode.text)["file"]

                # декодируем строку из base64 в бинарные данные
                binary_data = base64.b64decode(qrcode_base64_data)
                # создаем объект изображения из бинарных данных
                img = Image.open(BytesIO(binary_data))
                # сохраняем изображение в файл
                folder_path = os.path.join(
                    os.getcwd(), 'fbs_mode/data_for_barcodes/qrcode_supply')
                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)
                img.save(
                    f"{folder_path}/{self.supply_id}.png")
            else:
                text = 'Поставка не сформирована (qrcode_supply), так как нет артикулов'
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=text, parse_mode='HTML')
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'qrcode_supply', self.qrcode_supply, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def list_for_print_create(self):
        """
        WILDBERRIES.
        Функция создает список с полными именами файлов, которые нужно объединить
        amount_articles - словарь с данными {артикул_продавца: количество}.
        Объединяет эти файлы и сохраняет конечный файл на дропбоксе.
        """
        try:
            if self.amount_articles:
                qrcode_list = qrcode_print_for_products()
                pdf_filenames = glob.glob(
                    'fbs_mode/data_for_barcodes/cache_dir/*.pdf')
                logging.info(f"len(pdf_filenames): {len(pdf_filenames)}")
                mes_text = f'длина списка из папки cache_dir/*.pdf {len(pdf_filenames)}'
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=mes_text, parse_mode='HTML')
                text = str(self.amount_articles)
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=text, parse_mode='HTML')
                list_pdf_file_ticket_for_complect = []
                for j in pdf_filenames:
                    while self.amount_articles[str(Path(j).stem)] > 0:
                        list_pdf_file_ticket_for_complect.append(j)
                        self.amount_articles[str(Path(j).stem)] -= 1
                logging.info(
                    f"list_pdf_file_ticket_for_complect после добавления количества: {list_pdf_file_ticket_for_complect}")
                # Определяем число qr кодов для поставки.
                amount_of_supply_qrcode = math.ceil(
                    len(list_pdf_file_ticket_for_complect)/20)
                for file in qrcode_list:
                    list_pdf_file_ticket_for_complect.append(file)
                logging.info(
                    f"list_pdf_file_ticket_for_complect после добавления qr кодов: {list_pdf_file_ticket_for_complect}")
                outer_list = []  # Внешний список для процесса сортировки
                for i in list_pdf_file_ticket_for_complect:
                    # Разделяю полное название файла на путь к файлу и имя файла
                    # Оказывается в python знаком \ отделяется последняя папка перед файлом
                    # А все внешние отделяются знаком /
                    last_slash_index = i.rfind("/")
                    result = [i[:last_slash_index], i[last_slash_index+1:]]
                    new_name = result
                    full_new_name = []  # Список с полным именени файла после разделения
                    # Имена QR кодов у меня составные. Состоят из нескольких слов с пробелами
                    # В этом цикле разделяю имена из предыдущих списков по пробелу.
                    for j in new_name:
                        split_name = j.split(' ')
                        full_new_name.append(split_name)
                    outer_list.append(full_new_name)
                # Сортирую самый внешний список по последнему элемену самого внутреннего списка
                sorted_list = sorted(outer_list, key=lambda x: x[-1][-1])
                # Далее идет обратный процесс - процесс объединения элементов списка
                # в первоначальные имена файлов, но уже отсортированные
                new_sort = []
                for i in sorted_list:
                    inner_new_sort = []
                    for j in i:
                        j = ' '.join(j)
                        inner_new_sort.append(j)
                    new_sort.append(inner_new_sort)
                last_sorted_list = []
                for i in new_sort:
                    i = '/'.join(i)
                    last_sorted_list.append(i)
                list_pdf_file_ticket_for_complect = last_sorted_list
                logging.info(
                    f"list_pdf_file_ticket_for_complect перед группировкой файлов: {list_pdf_file_ticket_for_complect}")
                mes_text = f'длина списка list_pdf_file_ticket_for_complect для печати этикеток после сортировки {len(list_pdf_file_ticket_for_complect)}'
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=mes_text, parse_mode='HTML')
                qrcode_supply_amount = supply_qrcode_to_standart_view()
                if len(qrcode_supply_amount) != 0:
                    while amount_of_supply_qrcode > 0:
                        list_pdf_file_ticket_for_complect.append(
                            qrcode_supply_amount[0])
                        amount_of_supply_qrcode -= 1
                folder_path = os.path.join(
                    os.getcwd(), 'fbs_mode/data_for_barcodes/done_data')
                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)
                file_name = (f'{folder_path}/Наклейки для комплектовщиков '
                             f'{time.strftime("%Y-%m-%d %H-%M")}.pdf')
                saved_on_dropbox_filename = f'{self.dropbox_current_assembling_folder}/WB - {self.file_add_name} этикетки FBS {time.strftime("%Y-%m-%d %H-%M-%S")}.pdf'
                print_barcode_to_pdf2(list_pdf_file_ticket_for_complect,
                                      file_name,
                                      saved_on_dropbox_filename)
            else:
                text = 'не сработала list_for_print_create потому что нет данных'
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=text, parse_mode='HTML')
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'list_for_print_create', self.list_for_print_create, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')


class OzonFbsMode():
    """Класс отвечает за работу с заказами ОЗОН"""

    def __init__(self, ozon_headers, db_folder, file_add_name):

        # Данные отправлений для FBS OZON
        self.ozon_headers = ozon_headers
        self.dropbox_main_fbs_folder = db_folder
        self.file_add_name = file_add_name

        self.posting_data = []
        self.warehouse_dict = {}
        if self.file_add_name == 'ООО':
            self.warehouse_dict = {
                'day_stock': 1020001030027000, 'night_stock': 22676408482000}
        elif self.file_add_name == 'ИП':
            self.warehouse_dict = {
                'day_stock': 1020000089903000, 'night_stock': 22655170176000}
        self.date_for_files = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
        self.main_save_folder_server = 'fbs_mode/data_for_barcodes'
        self.date_before = datetime.now() - timedelta(days=20)
        self.tomorrow_date = datetime.now() + timedelta(days=1)
        self.future_date = datetime.now() + timedelta(days=20)

        self.since_date_for_filter = self.date_before.strftime(
            '%Y-%m-%d') + 'T08:00:00Z'

        # Словарь с данными {'артикул продавца': 'количество'}. Для сводной таблицы по FBS
        self.ozon_article_amount = {}

    def awaiting_packaging_orders(self):
        """
        OZON.
        Функция собирает новые заказы и преобразует данные
        от заказа в словарь для дальнейшей работы с ним
        """
        try:
            # Метод для просмотра новых заказов
            url = "https://api-seller.ozon.ru/v3/posting/fbs/unfulfilled/list"

            future_date_for_filter = self.future_date.strftime(
                '%Y-%m-%d') + 'T08:00:00Z'

            payload = json.dumps({
                "dir": "ASC",
                "filter": {
                    "cutoff_from": self.since_date_for_filter,
                    "cutoff_to": future_date_for_filter,
                    "delivery_method_id": [],
                    "provider_id": [],
                    "status": "awaiting_packaging",
                    "warehouse_id": []
                },
                "limit": 100,
                "offset": 0,
                "with": {
                    "analytics_data": True,
                    "barcodes": True,
                    "financial_data": True,
                    "translit": True
                }
            })

            response = requests.request(
                "POST", url, headers=self.ozon_headers, data=payload)
            # Заполняю словарь данными
            for i in json.loads(response.text)['result']['postings']:
                inner_dict_data = {}
                inner_dict_data['posting_number'] = i['posting_number']
                inner_dict_data['order_id'] = i['order_id']
                inner_dict_data['delivery_method'] = i['delivery_method']
                inner_dict_data['in_process_at'] = i['in_process_at']
                inner_dict_data['shipment_date'] = i['shipment_date']
                inner_dict_data['products'] = i['products']
                self.posting_data.append(inner_dict_data)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'awaiting_packaging_orders', self.awaiting_packaging_orders, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def awaiting_deliver_orders(self):
        """
        OZON.
        Делит заказ на отправления и переводит его в статус awaiting_deliver.
        Каждый элемент в packages может содержать несколько элементов products или отправлений.
        Каждый элемент в products — это товар, включённый в данное отправление.
        """
        try:
            url = 'https://api-seller.ozon.ru/v4/posting/fbs/ship'
            for data_dict in self.posting_data:
                for sku_data in data_dict['products']:
                    payload = json.dumps({
                        "packages": [
                            {
                                "products": [
                                    {
                                        "product_id": sku_data['sku'],
                                        "quantity": sku_data['quantity']
                                    }
                                ]
                            }
                        ],
                        "posting_number": data_dict['posting_number'],
                        "with": {
                            "additional_data": True
                        }
                    })
                    response = requests.request(
                        "POST", url, headers=self.ozon_headers, data=payload)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'awaiting_deliver_orders', self.awaiting_deliver_orders, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def check_folder_exists(self, path):
        try:
            dbx_db.files_list_folder(path)
            return True
        except dropbox.exceptions.ApiError as e:
            return False

    def prepare_data_for_confirm_delivery(self):
        """OZON. Подготовка данных для подтверждения отгрузки"""
        try:
            hour = datetime.now().hour
            date_folder = datetime.today().strftime('%Y-%m-%d')

            to_date_for_filter = self.tomorrow_date.strftime(
                '%Y-%m-%d') + 'T23:59:00Z'
            date_confirm_delivery = self.tomorrow_date.strftime('%Y-%m-%d')

            # Словарь с данными {'Номер склада': {quantity: 'Количество артикулов', containers_count: 'количество коробок'}}
            self.ware_house_amount_dict = {}
            # Словарь с данными {'номер отправления': {'Артикул продавца': 'количество'}}
            self.fbs_ozon_common_data = {}

            # Словарь с данными {'номер отправления': [{'Артикул продавца': 'seller_article',
            # 'Наименование': 'article_name', 'Количество': 'amount'}]}
            self.fbs_ozon_common_data_buils_dict = {}

            if hour >= 6 and hour < 18:
                self.delivary_method_id = self.warehouse_dict['day_stock']
                self.dropbox_current_assembling_folder = f'{self.dropbox_main_fbs_folder}/!ДЕНЬ СБОРКА ФБС/{date_folder}'
            else:
                self.delivary_method_id = self.warehouse_dict['night_stock']
                self.dropbox_current_assembling_folder = f'{self.dropbox_main_fbs_folder}/!НОЧЬ СБОРКА ФБС/{date_folder}'
            self.departure_date = date_confirm_delivery + 'T10:00:00Z'

            # Создаем папку на dropbox, если ее еще нет
            if self.check_folder_exists(self.dropbox_current_assembling_folder) == False:
                dbx_db.files_create_folder_v2(
                    self.dropbox_current_assembling_folder)

            # Проверяем все отгрузки, которые буду завтра
            url = 'https://api-seller.ozon.ru/v3/posting/fbs/list'
            amount_products = 0
            payload = json.dumps(
                {
                    "dir": "asc",
                    "filter": {
                        "delivery_method_id": [self.delivary_method_id],
                        "provider_id": [24],
                        "since": self.since_date_for_filter,
                        "status": "awaiting_deliver",
                        "to": to_date_for_filter,
                        "warehouse_id": [self.delivary_method_id]
                    },
                    "limit": 999,
                    "offset": 0,
                    "with": {
                        "analytics_data": True,
                        "barcodes": True,
                        "financial_data": True,
                        "translit": True
                    }
                }
            )
            response = requests.request(
                "POST", url, headers=self.ozon_headers, data=payload)
            for data in json.loads(response.text)['result']['postings']:
                inner_article_amount_dict = {}
                inner_bilding_list = []
                if self.tomorrow_date.strftime('%Y-%m-%d') in data["shipment_date"]:
                    # if '29' in data["shipment_date"]:
                    for product in data['products']:
                        inner_bilding_dict = {}
                        inner_bilding_dict['Артикул продавца'] = product['offer_id']
                        inner_bilding_dict['Наименование'] = product['name']
                        inner_bilding_dict['Количество'] = product['quantity']

                        amount_products += product['quantity']
                        inner_article_amount_dict[product['offer_id']
                                                  ] = product['quantity']
                        inner_bilding_list.append(inner_bilding_dict)
                        if product['offer_id'] not in self.ozon_article_amount.keys():
                            self.ozon_article_amount[product['offer_id']] = int(
                                product['quantity'])
                        else:
                            self.ozon_article_amount[product['offer_id']
                                                     ] = self.ozon_article_amount[product['offer_id']] + int(product['quantity'])

                    self.fbs_ozon_common_data[data['posting_number']
                                              ] = inner_article_amount_dict

                    # Словарь для файла сборки
                    self.fbs_ozon_common_data_buils_dict[data['posting_number']
                                                         ] = inner_bilding_list
            containers_count = math.ceil(amount_products/20)
            self.ware_house_amount_dict[self.delivary_method_id] = {
                'quantity': amount_products, 'containers_count': containers_count}
            return self.ozon_article_amount
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'prepare_data_for_confirm_delivery', self.prepare_data_for_confirm_delivery, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def confirm_delivery_create_document(self):
        """
        OZON.
        Функция подтверждает отгрузку и запускает создание документов на стороне ОЗОН"""
        try:
            url = 'https://api-seller.ozon.ru/v2/posting/fbs/act/create'
            if self.ware_house_amount_dict[self.delivary_method_id]['quantity'] > 0:
                payload = json.dumps({
                    "containers_count": self.ware_house_amount_dict[self.delivary_method_id]['containers_count'],
                    "delivery_method_id": self.delivary_method_id,
                    "departure_date": self.departure_date
                })
                response = requests.request(
                    "POST", url, headers=self.ozon_headers, data=payload)

                self.delivery_id = json.loads(response.text)['result']['id']
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'confirm_delivery_create_document', self.confirm_delivery_create_document, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def check_delivery_create(self):
        """
        OZON.
        Функция проверяет, что отгрузка создана.
        Формирует список отправлений для дальнейшей работы
        """
        try:
            url = 'https://api-seller.ozon.ru/v2/posting/fbs/act/check-status'

            payload = json.dumps(
                {
                    "id": self.delivery_id
                    # "id": 35178630
                }
            )
            response = requests.request(
                "POST", url, headers=self.ozon_headers, data=payload)

            data = json.loads(response.text)['result']
            if data['status'] == "ready":

                self.delivery_number_list = data["added_to_act"]
                print('Функция check_delivery_create сработала. Спать 5 мин не нужно')
            else:
                print('уснули на 5 мин в функции check_delivery_create')
                time.sleep(300)
                self.check_delivery_create()
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'check_delivery_create', self.check_delivery_create, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def receive_barcode_delivery(self):
        """OZON. Получает и сохраняет штрихкод поставки"""
        try:
            url = 'https://api-seller.ozon.ru/v2/posting/fbs/act/get-barcode'
            payload = json.dumps(
                {
                    "id": self.delivery_id
                }
            )
            response = requests.request(
                "POST", url, headers=self.ozon_headers, data=payload)

            image = Image.open(io.BytesIO(response.content))
            folder_path = os.path.join(
                os.getcwd(), f'{self.main_save_folder_server}/ozon_delivery_barcode')
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            save_folder_docs = f"{folder_path}/{self.delivery_id}_баркод {self.date_for_files}.png"
            image.save(save_folder_docs)

            # Сохраняем штрихкод в PDF формате
            now_date = datetime.now().strftime(("%d.%m"))
            im = Image.open(save_folder_docs)
            text_common = "Штрихкод для отгрузки ОЗОН"
            text_company = ''
            if self.file_add_name == 'ООО':
                text_company = 'ООО Иннотрейд'
            elif self.file_add_name == 'ИП':
                text_company = 'ИП Караваев'
            font = ImageFont.truetype("arial.ttf", size=50)
            A4 = (1033, 1462)
            white_A4 = Image.new('RGB', A4, 'white')
            text_draw = ImageDraw.Draw(white_A4)
            text_common_lenght = text_draw.textlength(text_common, font=font)
            text_company_lenght = text_draw.textlength(text_company, font=font)
            x = (A4[0] - im.width) // 2
            text_common_x = (A4[0] - round(text_common_lenght)) // 2
            text_company_x = (A4[0] - round(text_company_lenght)) // 2
            white_A4.paste(im, (x, 330))
            text_draw.text((text_common_x, 100), text_common,
                           font=font, fill=('#000000'))
            text_draw.text((text_company_x, 200), text_company,
                           font=font, fill=('#000000'))
            save_folder_docs_pdf = f"{self.main_save_folder_server}/ozon_delivery_barcode/{self.delivery_id}_баркод {now_date}.pdf"
            white_A4.save(save_folder_docs_pdf, 'PDF', quality=100)

            folder = (
                f'{self.dropbox_current_assembling_folder}/OZON - {self.file_add_name} акт {now_date}.pdf')
            with open(save_folder_docs_pdf, 'rb') as f:
                dbx_db.files_upload(f.read(), folder)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'receive_barcode_delivery', self.receive_barcode_delivery, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def create_ozone_selection_sheet_pdf(self):
        """OZON. Создает лист подбора для OZON"""
        try:
            sorted_data_buils_dict = dict(
                sorted(self.fbs_ozon_common_data_buils_dict.items(), key=lambda x: x[0][-6:]))

            number_of_departure_oz = sorted_data_buils_dict.keys()

            ozone_selection_sheet_xls = openpyxl.Workbook()
            create = ozone_selection_sheet_xls.create_sheet(
                title='pivot_list', index=0)
            sheet = ozone_selection_sheet_xls['pivot_list']
            sheet['A1'] = 'Номер отправления'
            sheet['B1'] = 'Наименование товара'
            sheet['C1'] = 'Артикул'
            sheet['D1'] = 'Количество'

            al = Alignment(horizontal="center",
                           vertical="center", wrap_text=True)
            al2 = Alignment(vertical="center", wrap_text=True)
            thin = Side(border_style="thin", color="000000")
            thick = Side(border_style="medium", color="000000")
            pattern = PatternFill('solid', fgColor="fcff52")

            upd_number_of_departure_oz = []
            upd_product_name_oz = []
            upd_name_for_print_oz = []
            upd_amount_for_print_oz = []

            for key, value in sorted_data_buils_dict.items():
                for data in value:
                    upd_number_of_departure_oz.append(key)
                    upd_product_name_oz.append(data['Наименование'])
                    upd_name_for_print_oz.append(data['Артикул продавца'])
                    upd_amount_for_print_oz.append(data['Количество'])

            create.column_dimensions['A'].width = 18
            create.column_dimensions['B'].width = 38
            create.column_dimensions['C'].width = 18
            create.column_dimensions['D'].width = 10

            for i in range(len(upd_number_of_departure_oz)):
                create.cell(
                    row=i+2, column=1).value = upd_number_of_departure_oz[i]
                create.cell(row=i+2, column=2).value = upd_product_name_oz[i]
                create.cell(row=i+2, column=3).value = upd_name_for_print_oz[i]
                create.cell(
                    row=i+2, column=4).value = upd_amount_for_print_oz[i]
            for i in range(1, len(upd_number_of_departure_oz)+2):
                wrapped_lines = textwrap.wrap(create.cell(
                    row=i, column=2).value, width=38)
                num_lines = len(wrapped_lines)
                row_height = 18 * num_lines
                create.row_dimensions[i].height = row_height
                for c in create[f'A{i}:D{i}']:
                    c[0].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[1].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[2].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[3].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[3].alignment = al
                    for j in range(3):
                        c[j].alignment = al2

            folder_path_docs = os.path.join(
                os.getcwd(), f'{self.main_save_folder_server}/ozon_docs')
            if not os.path.exists(folder_path_docs):
                os.makedirs(folder_path_docs)
            name_for_file = f'{folder_path_docs}/selection_sheet'

            ozone_selection_sheet_xls.save(f'{name_for_file}.xlsx')
            w_b2 = load_workbook(f'{name_for_file}.xlsx')
            source_page2 = w_b2.active
            number_of_departure_oz = source_page2['A']
            amount_for_print_oz = source_page2['D']

            for i in range(1, len(number_of_departure_oz)+2):
                if i < len(number_of_departure_oz)-1:
                    for c in source_page2[f'A{i+1}:D{i+1}']:
                        if (number_of_departure_oz[i].value == number_of_departure_oz[i+1].value
                                and number_of_departure_oz[i].value != number_of_departure_oz[i-1].value):
                            c[0].border = Border(
                                top=thick, left=thick, bottom=thin, right=thin)
                            c[1].border = Border(
                                top=thick, left=thin, bottom=thin, right=thin)
                            c[2].border = Border(
                                top=thick, left=thin, bottom=thin, right=thin)
                            c[3].border = Border(
                                top=thick, left=thin, bottom=thin, right=thick)
                            for j in range(4):
                                c[j].fill = pattern
                        if (number_of_departure_oz[i].value == number_of_departure_oz[i+1].value
                                and number_of_departure_oz[i].value == number_of_departure_oz[i-1].value):
                            c[0].border = Border(
                                top=thin, left=thick, bottom=thin, right=thin)
                            c[1].border = Border(
                                top=thin, left=thin, bottom=thin, right=thin)
                            c[2].border = Border(
                                top=thin, left=thin, bottom=thin, right=thin)
                            c[3].border = Border(
                                top=thin, left=thin, bottom=thin, right=thick)
                            for j in range(4):
                                c[j].fill = pattern
                        elif (number_of_departure_oz[i].value != number_of_departure_oz[i+1].value
                                and number_of_departure_oz[i].value == number_of_departure_oz[i-1].value):
                            c[0].border = Border(
                                top=thin, left=thick, bottom=thick, right=thin)
                            c[1].border = Border(
                                top=thin, left=thin, bottom=thick, right=thin)
                            c[2].border = Border(
                                top=thin, left=thin, bottom=thick, right=thin)
                            c[3].border = Border(
                                top=thin, left=thin, bottom=thick, right=thick)
                            for j in range(4):
                                c[j].fill = pattern
                        if amount_for_print_oz[i].value > 1:
                            c[0].border = Border(
                                top=thick, left=thick, bottom=thick, right=thin)
                            c[1].border = Border(
                                top=thick, left=thin, bottom=thick, right=thin)
                            c[2].border = Border(
                                top=thick, left=thin, bottom=thick, right=thin)
                            c[3].border = Border(
                                top=thick, left=thin, bottom=thick, right=thick)
                            for j in range(4):
                                c[j].fill = pattern

            w_b2.save(f'{name_for_file}.xlsx')

            path_file = os.path.abspath(f'{name_for_file}.xlsx')
            only_file_name = os.path.splitext(os.path.basename(path_file))[0]

            folder_path = os.path.dirname(os.path.abspath(path_file))
            output = convert(source=path_file, output_dir=folder_path, soft=1)

            folder = (
                f'{self.dropbox_current_assembling_folder}/OZON - {self.file_add_name} лист подбора {self.date_for_files}.pdf')
            with open(output, 'rb') as f:
                dbx_db.files_upload(f.read(), folder)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'create_ozone_selection_sheet_pdf', self.create_ozone_selection_sheet_pdf, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def check_status_formed_invoice(self):
        """
        OZON.
        Проверяет статус формирования накладной: /v2/posting/fbs/digital/act/check-status. 
        Когда статус документа перейдёт в FORMED или CONFIRMED, получите файлы:
        Получает и сохраняет этикетки для каждой коробки: /v2/posting/fbs/act/get-container-labels.
        Этикетки для каждой отправки: /v2/posting/fbs/package-label.
        """
        try:
            url = 'https://api-seller.ozon.ru/v2/posting/fbs/digital/act/check-status'
            payload = json.dumps(
                {
                    "id": self.delivery_id
                    # "id": 35178630
                }
            )
            response = requests.request(
                "POST", url, headers=self.ozon_headers, data=payload)

            data = json.loads(response.text)['status']

            if data == "CONFIRMED" or data == "FORMED":
                print(
                    'Функция check_status_formed_invoice сработала. Спать 5 мин не нужно')
                self.receive_barcode_delivery()
                self.get_box_tickets()
            else:
                print('уснули на 5 мин в функции check_status_formed_invoice')
                time.sleep(300)
                self.check_status_formed_invoice()
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'check_status_formed_invoice', self.check_status_formed_invoice, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def get_box_tickets(self):
        """
        OZON.
        Функция получает и сохраняет этикетки для каждой коробки в формате PDF.
        Данные получают из эндпоинта /v2/posting/fbs/act/get-container-labels.
        """
        try:
            url = 'https://api-seller.ozon.ru/v2/posting/fbs/act/get-container-labels'

            payload = json.dumps(
                {
                    "id": self.delivery_id
                }
            )
            response = requests.request(
                "POST", url, headers=self.ozon_headers, data=payload)

            # получение данных PDF из входящих данных
            pdf_data = response.content  # замените на фактические входные данные
            folder_path = os.path.join(
                os.getcwd(), f'{self.main_save_folder_server}/ozon_docs')
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            save_folder_docs = f'{folder_path}/OZON - {self.file_add_name} ШК на 1 коробку {self.date_for_files}.pdf'
            # сохранение PDF-файла
            with open(save_folder_docs, 'wb') as f:
                f.write(pdf_data)
            folder = (
                f'{self.dropbox_current_assembling_folder}/OZON - {self.file_add_name} ШК на 1 коробку {self.date_for_files}.pdf')
            with open(save_folder_docs, 'rb') as f:
                dbx_db.files_upload(f.read(), folder)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'get_box_tickets', self.get_box_tickets, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def forming_package_ticket_with_article(self):
        """
        OZON.
        Функция получает и сохраняет этикетки для каждой отправки.
        Данные получают из эндпоинта /v2/posting/fbs/package-label.
        После получения и сохранения всех этикеток помещает артикул продавца
        на этикетку.
        """
        try:
            url = 'https://api-seller.ozon.ru/v2/posting/fbs/package-label'
            save_folder = os.path.join(
                os.getcwd(), f'{self.main_save_folder_server}/package_tickets')
            if not os.path.exists(save_folder):
                os.makedirs(save_folder)
            for package in self.fbs_ozon_common_data.keys():
                payload = json.dumps(
                    {
                        "posting_number": [package]
                    }
                )
                response = requests.request(
                    "POST", url, headers=self.ozon_headers, data=payload)

                save_folder_docs = f'{save_folder}/{package}.pdf'
                # сохранение PDF-файла
                with open(save_folder_docs, 'wb') as f:
                    f.write(response.content)

            done_files_folder = os.path.join(
                os.getcwd(), f'{self.main_save_folder_server}/package_tickets/done')
            if not os.path.isdir(done_files_folder):
                os.mkdir(done_files_folder)

            # Записываем артикул продавциа в файл с этикеткой
            new_data_for_ozon_ticket(save_folder, self.fbs_ozon_common_data)
            # Формируем список всех файлов, которые нужно объединять в один документ
            list_filename = glob.glob(f'{done_files_folder}/*.pdf')
            # Адрес и имя конечного файла
            file_name = f'{done_files_folder}/done_tickets.pdf'
            merge_barcode_for_ozon_two_on_two(list_filename, file_name)
            folder = f'{self.dropbox_current_assembling_folder}/OZON - {self.file_add_name} этикетки.pdf'
            with open(file_name, 'rb') as f:
                dbx_db.files_upload(f.read(), folder)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'forming_package_ticket_with_article', self.forming_package_ticket_with_article, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')


class YandexMarketFbsMode():
    """Класс отвечает за работу с заказами Yandex"""

    def __init__(self, yandex_headers, db_folder, file_add_name):
        """Основные данные класса"""
        self.yandex_headers = yandex_headers
        self.dropbox_main_fbs_folder = db_folder
        self.file_add_name = file_add_name

        if self.file_add_name == 'ООО':
            self.compaign_id = 23746359
        elif self.file_add_name == 'ИП':
            self.compaign_id = 74448338

        self.amount_articles = {}
        self.shipment_id = ''
        self.orders_list = []
        self.main_save_folder_server = 'fbs_mode/data_for_barcodes'
        self.date_for_files = datetime.now().strftime('%Y-%m-%d %H-%M')
        hour = datetime.now().hour
        date_folder = datetime.today().strftime('%Y-%m-%d')

        if hour >= 6 and hour < 18:
            self.dropbox_current_assembling_folder = f'{self.dropbox_main_fbs_folder}/!ДЕНЬ СБОРКА ФБС/{date_folder}'
        else:
            self.dropbox_current_assembling_folder = f'{self.dropbox_main_fbs_folder}/!НОЧЬ СБОРКА ФБС/{date_folder}'

    def check_folder_exists(self, path):
        try:
            dbx_db.files_list_folder(path)
            return True
        except dropbox.exceptions.ApiError as e:
            return False

    def receive_orders_data(self):
        """
        YANDEXMARKET
        Функция обрабатывает заказы.
        Выделяет номера заказов, артикулы продавца в них, кол-во и название артикула.
        Создает список с данными каждого заказа: [{Заказ: [{артикул_продавца: артикул, название_артикула: название,
        количество_в_заказе: количество}]}]
        """
        try:
            url = f"https://api.partner.market.yandex.ru/campaigns/{self.compaign_id}/orders?status=PROCESSING&substatus=STARTED"
            response = requests.request(
                "GET", url, headers=self.yandex_headers)
            orders_list = json.loads(response.text)['orders']

            main_orders_list = []
            for order in orders_list:
                order_dict = {}
                article_list_in_order = order['items']
                inner_article_list = []
                inner_article_dict = {}
                for article in article_list_in_order:

                    inner_article_dict = {}

                    inner_article_dict['seller_article'] = article['offerId']
                    inner_article_dict['article_name'] = article['offerName']
                    inner_article_dict['amount'] = article['count']

                    inner_article_list.append(inner_article_dict)
                inner_article_dict['articles_info'] = inner_article_list
                inner_article_dict['delivery_date'] = order['delivery']['dates']['toDate']
                order_dict['order_id'] = order['id']
                order_dict['order_data'] = inner_article_dict

                main_orders_list.append(order_dict)
            return main_orders_list
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'receive_orders_data', self.receive_orders_data, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def change_orders_status(self):
        """
        YANDEXMARKET
        Функция изменяет статус новых заказов.
        """
        try:
            orders_list = self.receive_orders_data()
            for order in orders_list:
                status_url = f"https://api.partner.market.yandex.ru/campaigns/{self.compaign_id}/orders/{order['order_id']}/status"

                payload = json.dumps(
                    {
                        "order": {
                            "status": "PROCESSING",
                            "substatus": "READY_TO_SHIP",
                        }
                    }
                )
                response_data = requests.request(
                    "PUT", status_url, headers=self.yandex_headers, data=payload)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'change_orders_status', self.change_orders_status, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def receive_delivery_number(self):
        """
        YANDEXMARKET
        Определяет id поставки для подтверждения отгрузки.
        """
        try:
            url_delivery = f'https://api.partner.market.yandex.ru/campaigns/{self.compaign_id}/first-mile/shipments'
            date_for_delivery = datetime.now() + timedelta(days=1)
            date_for_delivery = date_for_delivery.strftime('%d-%m-%Y')
            payload = json.dumps({
                "dateFrom": date_for_delivery,
                "dateTo": date_for_delivery,
                "statuses": [
                    "OUTBOUND_CREATED", "OUTBOUND_CONFIRMED",
                    "OUTBOUND_READY_FOR_CONFIRMATION", "FINISHED"
                ]
            })
            response_delivery = requests.request(
                "PUT", url_delivery, headers=self.yandex_headers, data=payload)

            shipments_list = json.loads(response_delivery.text)[
                'result']['shipments']
            if len(shipments_list) == 0:
                message_text = 'На завтра нет поставки Яндекс Маркета'
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=message_text, parse_mode='HTML')
            else:
                shipment_dict = {}
                shipment_id = shipments_list[0]['id']
                self.shipment_id = shipments_list[0]['id']
                shipment_dict['shipment_id'] = shipment_id
                return shipment_dict
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'receive_delivery_number', self.receive_delivery_number, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def received_info_about_delivery(self):
        """
        YANDEXMARKET.
        Получает информацию о предстоящей отгрузке.
        """
        try:
            shipment_data = self.receive_delivery_number()
            if shipment_data:
                shipment_id = shipment_data['shipment_id']
                # shipment_id = 45793720
                url_info = f'https://api.partner.market.yandex.ru/campaigns/{self.compaign_id}/first-mile/shipments/{shipment_id}'
                response = requests.request(
                    "GET", url_info, headers=self.yandex_headers)
                info_main_data = json.loads(response.text)['result']
                orders_list = info_main_data['orderIds']
                shipment_data = {}
                shipment_data['shipment_id'] = shipment_id
                shipment_data['orders_list'] = orders_list
                return shipment_data
            else:
                return shipment_data

        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'received_info_about_delivery', self.received_info_about_delivery, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def check_actual_orders(self):
        """
        YANDEXMARKET
        Проверяет все ордеры в поставке и исключает отмененнные
        """
        try:
            shipment_data = self.received_info_about_delivery()
            if shipment_data:
                orders_info_list = []
                inner_info_dict = {}
                raw_orders_list = shipment_data['orders_list']
                for order in raw_orders_list:
                    inner_info_list = []
                    url_check = f'https://api.partner.market.yandex.ru/campaigns/{self.compaign_id}/orders/{order}'
                    response = requests.request(
                        "GET", url_check, headers=self.yandex_headers)
                    check_main_data = json.loads(response.text)['order']
                    if check_main_data['status'] != 'CANCELLED':
                        self.orders_list.append(order)
                        for article in check_main_data['items']:
                            inner_article_list = {}
                            inner_article_list['seller_article'] = article['offerId']
                            inner_article_list['name'] = article['offerName']
                            inner_article_list['amount'] = article['count']
                            inner_info_list.append(inner_article_list)
                        inner_info_dict[order] = inner_info_list
                        orders_info_list.append(inner_info_dict)
                return inner_info_dict
            else:
                return None
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'check_actual_orders', self.check_actual_orders, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def yandex_prepare_data(self):
        """
        YANDEX.
        Готовит данные для общего файла для производства"""
        try:
            raw_data = self.check_actual_orders()
            if raw_data:
                article_count_dict = {}
                for order_id, article_list in raw_data.items():
                    for article_dict in article_list:
                        if article_dict['seller_article'] in article_count_dict.keys():
                            article_count_dict[article_dict['seller_article']] = int(
                                article_dict['amount']) + int(article_count_dict[article_dict['seller_article']])
                        else:
                            article_count_dict[article_dict['seller_article']] = int(
                                article_dict['amount'])
                return article_count_dict
            else:
                return None
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'yandex_prepare_data', self.yandex_prepare_data, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def approve_shipment(self):
        """
        YANDEXMARKET
        Подтверждение отгрузки
        """
        try:
            shipment_data = self.received_info_about_delivery()
            shipment_id = shipment_data['shipment_id']

            approve_url = f'https://api.partner.market.yandex.ru/campaigns/{self.compaign_id}/first-mile/shipments/{shipment_id}/confirm'
            payload_approve = json.dumps({
                "externalShipmentId": f'{shipment_id}',
                "orderIds": self.orders_list
            })
            response = requests.request(
                "POST", approve_url, headers=self.yandex_headers, data=payload_approve)

        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'approve_shipment', self.approve_shipment, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def check_dropbox_folder_exist(self):
        """Проверяет, что на Dropbox существует необходимая папка"""
        if self.check_folder_exists(self.dropbox_current_assembling_folder) == False:
            dbx_db.files_create_folder_v2(
                self.dropbox_current_assembling_folder)

    def saving_act(self):
        """
        YANDEXMARKET
        Сохраняет акт приема - передачи в PDF формате.
        """
        try:
            self.check_dropbox_folder_exist()
            url_act = f'https://api.partner.market.yandex.ru/campaigns/{self.compaign_id}/first-mile/shipments/{self.shipment_id}/act'
            response_act = requests.request(
                "GET", url_act, headers=self.yandex_headers)
            pdf_data = response_act.content
            folder_path = os.path.join(
                os.getcwd(), f'{self.main_save_folder_server}/yandex')
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            save_folder_docs = f'{folder_path}/YANDEX - {self.file_add_name} акт {self.date_for_files}.pdf'
            # сохранение PDF-файла
            with open(save_folder_docs, 'wb') as f:
                f.write(pdf_data)
            folder = (
                f'{self.dropbox_current_assembling_folder}/YANDEX - {self.file_add_name} акт {self.date_for_files}.pdf')
            with open(save_folder_docs, 'rb') as f:
                dbx_db.files_upload(f.read(), folder)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'saving_act', self.saving_act, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def saving_tickets(self):
        """
        YANDEXMARKET
        Сохраняет этикетки в PDF формате.
        """
        try:
            orders_info_list = self.check_actual_orders()
            self.check_dropbox_folder_exist()
            # Проверка существования папок
            raw_ticket_folder = os.path.join(
                os.getcwd(), f'{self.main_save_folder_server}/yandex')
            if not os.path.exists(raw_ticket_folder):
                os.makedirs(raw_ticket_folder)
            folder_path = os.path.join(
                os.getcwd(), f'{self.main_save_folder_server}/yandex/tickets')
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            done_tickets_folder = os.path.join(
                os.getcwd(), f'{self.main_save_folder_server}/yandex/tickets/done')
            if not os.path.exists(done_tickets_folder):
                os.makedirs(done_tickets_folder)

            for order in self.orders_list:
                url_tickets = f'https://api.partner.market.yandex.ru/campaigns/{self.compaign_id}/orders/{order}/delivery/labels'
                response_tickets = requests.request(
                    "GET", url_tickets, headers=self.yandex_headers)
                pdf_data = response_tickets.content  # замените на фактические входные данные
                save_folder_docs = f'{folder_path}/{order}.pdf'
                # сохранение PDF-файла
                with open(save_folder_docs, 'wb') as f:
                    f.write(pdf_data)

            new_data_for_yandex_ticket(folder_path, orders_info_list)
            list_filenames = glob.glob(f'{done_tickets_folder}/*.pdf')
            folder_summary_file_name = f'{self.main_save_folder_server}/yandex/YANDEX - {self.file_add_name} этикетки {self.date_for_files}.pdf'
            merge_barcode_for_yandex_two_on_two(
                list_filenames, folder_summary_file_name)
            folder = (
                f'{self.dropbox_current_assembling_folder}/YANDEX - {self.file_add_name} этикетки {self.date_for_files}.pdf')
            with open(folder_summary_file_name, 'rb') as f:
                dbx_db.files_upload(f.read(), folder)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'saving_tickets', self.saving_tickets, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def create_yandex_selection_sheet_pdf(self):
        """
        YANDEXMARKET
        Формирует файл подбора
        """
        try:
            date_for_files = datetime.now().strftime('%Y-%m-%d')
            main_shipment_data = self.check_actual_orders()
            number_of_departure_ya = main_shipment_data.keys()
            yandex_selection_sheet_xls = openpyxl.Workbook()
            create = yandex_selection_sheet_xls.create_sheet(
                title='pivot_list', index=0)
            create.page_setup.paperSize = create.PAPERSIZE_A4
            create.page_setup.orientation = create.ORIENTATION_PORTRAIT
            create.page_margins.left = 0.25
            create.page_margins.right = 0.25
            create.page_margins.top = 0.25
            create.page_margins.bottom = 0.25
            create.page_margins.header = 0.3
            create.page_margins.footer = 0.3

            sheet = yandex_selection_sheet_xls['pivot_list']
            sheet['A1'] = f'Лист подбора Яндекс Маркета {date_for_files}'
            font = Font(size=16)
            # Применяем стиль шрифта к ячейке с надписью
            sheet['A1'].font = font
            sheet.merge_cells('A1:D1')

            sheet['A3'] = 'Номер отправления'
            sheet['B3'] = 'Наименование товара'
            sheet['C3'] = 'Артикул'
            sheet['D3'] = 'Количество'

            al = Alignment(horizontal="center",
                           vertical="center", wrap_text=True)
            al2 = Alignment(vertical="center", wrap_text=True)
            thin = Side(border_style="thin", color="000000")
            thick = Side(border_style="medium", color="000000")
            pattern = PatternFill('solid', fgColor="fcff52")

            upd_number_of_departure_ya = []
            upd_product_name_ya = []
            upd_name_for_print_ya = []
            upd_amount_for_print_ya = []

            for key, value in main_shipment_data.items():
                for data in value:
                    upd_number_of_departure_ya.append(key)
                    upd_product_name_ya.append(data['name'])
                    upd_name_for_print_ya.append(data['seller_article'])
                    upd_amount_for_print_ya.append(data['amount'])

            create.column_dimensions['A'].width = 18
            create.column_dimensions['B'].width = 38
            create.column_dimensions['C'].width = 18
            create.column_dimensions['D'].width = 12

            for i in range(len(upd_number_of_departure_ya)):
                create.cell(
                    row=i+4, column=1).value = upd_number_of_departure_ya[i]
                create.cell(row=i+4, column=2).value = upd_product_name_ya[i]
                create.cell(row=i+4, column=3).value = upd_name_for_print_ya[i]
                create.cell(
                    row=i+4, column=4).value = upd_amount_for_print_ya[i]
            for i in range(3, len(upd_number_of_departure_ya)+4):
                wrapped_lines = textwrap.wrap(create.cell(
                    row=i, column=2).value, width=38)
                num_lines = len(wrapped_lines)
                row_height = 18 * num_lines
                create.row_dimensions[i].height = row_height
                for c in create[f'A{i}:D{i}']:
                    c[0].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[1].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[2].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[3].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[3].alignment = al
                    for j in range(4):
                        c[j].alignment = al

            select_file_folder = os.path.join(
                os.getcwd(), f'{self.main_save_folder_server}/yandex')
            if not os.path.exists(select_file_folder):
                os.makedirs(select_file_folder)
            name_for_file = f'{select_file_folder}/YANDEX - selection_sheet {self.date_for_files}.xlsx'
            yandex_selection_sheet_xls.save(name_for_file)

            folder_path = os.path.dirname(
                os.path.abspath(name_for_file))
            path_file = os.path.abspath(name_for_file)
            output = convert(source=path_file, output_dir=folder_path, soft=1)

            folder = (
                f'{self.dropbox_current_assembling_folder}/YANDEX - {self.file_add_name} лист подбора {self.date_for_files}.pdf')
            # Сохраняем на DROPBOX
            with open(output, 'rb') as f:
                dbx_db.files_upload(f.read(), folder)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'create_yandex_selection_sheet_pdf', self.create_yandex_selection_sheet_pdf, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')


class CreatePivotFile(WildberriesFbsMode, OzonFbsMode, YandexMarketFbsMode):
    """Создает общий файл для производства"""

    def __init__(self, db_folder, file_add_name, headers_wb, headers_ozon, headers_yandex):
        self.dropbox_main_fbs_folder = db_folder
        self.file_add_name = file_add_name
        self.headers_wb = headers_wb
        self.headers_ozon = headers_ozon
        self.headers_yandex = headers_yandex
        self.main_save_folder_server = 'fbs_mode/data_for_barcodes'
        # Получаем текущую дату
        today = datetime.today()
        hour = today.hour
        wb_ip_days = ['Friday']
        # Используем метод strftime() для форматирования даты и вывода дня недели
        day_of_week = today.strftime('%A')

        if self.file_add_name == 'ИП' and day_of_week in wb_ip_days and hour >= 20:
            self.ozon_article_amount = None
            self.yandex_article_amount = None
        else:
            self.ozon_article_amount = OzonFbsMode(
                self.headers_ozon, self.dropbox_main_fbs_folder, self.file_add_name).prepare_data_for_confirm_delivery()
            self.yandex_article_amount = YandexMarketFbsMode(
                self.headers_yandex, self.dropbox_main_fbs_folder, self.file_add_name).yandex_prepare_data()
        if self.file_add_name == 'ООО':
            self.warehouse_dict = {
                'day_stock': 1020001030027000, 'night_stock': 22676408482000}
        elif self.file_add_name == 'ИП':
            self.warehouse_dict = {
                'day_stock': 1020000089903000, 'night_stock': 22655170176000}

    def delivery_data(self, next_number=0, limit_number=1000):
        """
        СВОДНЫЙ ФАЙЛ
        Собирает данные по поставке, которая уже создалась
        """
        try:
            url_data = f'https://suppliers-api.wildberries.ru/api/v3/supplies?limit={limit_number}&next={next_number}'
            response_data = requests.request(
                "GET", url_data, headers=wb_headers_karavaev)
            delivery_date = datetime.today().strftime("%d.%m.%Y")
            hour = datetime.now().hour
            delivery_name = ''
            if hour >= 6 and hour < 18:
                delivery_name = f'День {delivery_date}'
            else:
                delivery_name = f'Ночь {delivery_date}'
            if response_data.status_code == 200:
                if len(json.loads(response_data.text)['supplies']) >= limit_number:
                    next_number_new = json.loads(response_data.text)['next']
                    return self.delivery_data(next_number_new)
                else:
                    last_sup = json.loads(response_data.text)['supplies'][-1]
                    if delivery_name in last_sup['name']:
                        supply_id = last_sup['id']
                        return supply_id
                    else:
                        return ''
            else:
                time.sleep(5)
                return self.delivery_data()
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'delivery_data', self.delivery_data, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def data_for_production_list(self):
        """
        СВОДНЫЙ ФАЙЛ
        Сводит данные из данных о поставке к словарю: {артикул: количество}
        """
        try:
            supply_id = self.delivery_data()
            article_amount = {}
            if supply_id:
                url = f'https://suppliers-api.wildberries.ru/api/v3/supplies/{supply_id}/orders'
                response_data = requests.request(
                    "GET", url, headers=wb_headers_karavaev)
                if response_data.status_code == 200:

                    orders_data = json.loads(response_data.text)['orders']
                    for data in orders_data:
                        if data['article'] in article_amount:
                            article_amount[data['article']] += 1
                        else:
                            article_amount[data['article']] = 1
                else:
                    time.sleep(5)
                    return self.data_for_production_list()
            return article_amount
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'data_for_production_list', self.data_for_production_list, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')
            return article_amount

    def create_pivot_xls(self):
        '''
        СВОДНЫЙ ФАЙЛ.
        Создает сводный файл excel с количеством каждого артикула.
        Подключается к базе данных на сервере'''
        try:
            self.amount_articles = self.data_for_production_list()
            delivery_date = datetime.today().strftime("%d.%m.%Y %H-%M-%S")
            # Задаем словарь с данными WB, а входящий становится общим для всех маркетплейсов
            if self.amount_articles:
                wb_article_amount = self.amount_articles.copy()
            else:
                wb_article_amount = {}
            hour = datetime.now().hour
            date_folder = datetime.today().strftime('%Y-%m-%d')

            if hour >= 6 and hour < 18:
                self.delivary_method_id = self.warehouse_dict['day_stock']
                self.dropbox_current_assembling_folder = f'{self.dropbox_main_fbs_folder}/!ДЕНЬ СБОРКА ФБС/{date_folder}'
            else:
                self.delivary_method_id = self.warehouse_dict['night_stock']
                self.dropbox_current_assembling_folder = f'{self.dropbox_main_fbs_folder}/!НОЧЬ СБОРКА ФБС/{date_folder}'

            CELL_LIMIT = 16
            COUNT_HELPER = 2
            if self.ozon_article_amount:
                for article in self.ozon_article_amount.keys():
                    if article in self.amount_articles.keys():
                        self.amount_articles[article] = int(
                            self.amount_articles[article]) + int(self.ozon_article_amount[article])
                    else:
                        self.amount_articles[article] = int(
                            self.ozon_article_amount[article])
            if self.yandex_article_amount:
                for article in self.yandex_article_amount.keys():
                    if article in self.amount_articles.keys():
                        self.amount_articles[article] = int(
                            self.amount_articles[article]) + int(self.yandex_article_amount[article])
                    else:
                        self.amount_articles[article] = int(
                            self.yandex_article_amount[article])

            sorted_data_for_pivot_xls = dict(
                sorted(self.amount_articles.items(), key=lambda v: v[0].upper()))
            pivot_xls = openpyxl.Workbook()
            create = pivot_xls.create_sheet(title='pivot_list', index=0)
            sheet = pivot_xls['pivot_list']
            sheet['A1'] = 'Артикул продавца'
            sheet['B1'] = 'Ячейка стеллажа'
            # Можно вернуть столбец для производства
            # sheet['C1'] = 'На производство'
            sheet['D1'] = 'Всего для FBS'
            sheet['E1'] = 'FBS WB'
            sheet['F1'] = 'FBS Ozon'
            sheet['G1'] = 'FBY Yandex'

            for key, value in sorted_data_for_pivot_xls.items():
                create.cell(row=COUNT_HELPER, column=1).value = key
                create.cell(row=COUNT_HELPER, column=4).value = value
                COUNT_HELPER += 1
            select_file_folder = os.path.join(
                os.getcwd(), f'{self.main_save_folder_server}/pivot_excel')
            if not os.path.exists(select_file_folder):
                os.makedirs(select_file_folder)
            name_pivot_xls = f'{select_file_folder}/for_production.xlsx'
            path_file = os.path.abspath(name_pivot_xls)
            # file_name_dir = path.parent

            pivot_xls.save(name_pivot_xls)
            # ========= Подключение к базе данных ========= #
            engine = create_engine(
                "postgresql://databaseadmin:Up3psv8x@158.160.28.219:5432/innotreid")

            data = pd.read_sql_table(
                "database_shelvingstocks",
                con=engine,
                columns=['task_start_date',
                         'task_finish_date',
                         'seller_article_wb',
                         'seller_article',
                         'shelf_number',
                         'amount']
            )
            connection = psycopg2.connect(user="databaseadmin",
                                          # пароль, который указали при установке PostgreSQL
                                          password="Up3psv8x",
                                          host="158.160.28.219",
                                          port="5432",
                                          database="innotreid")
            cursor = connection.cursor()

            shelf_seller_article_list = data['seller_article_wb'].to_list()
            shelf_number_list = data['shelf_number'].to_list()
            shelf_amount_list = data['amount'].to_list()
            w_b = load_workbook(name_pivot_xls)
            source_page = w_b.active
            name_article = source_page['A']
            amount_all_fbs = source_page['D']

            for i in range(1, len(name_article)):
                if self.ozon_article_amount:
                    for j in self.ozon_article_amount.keys():
                        if name_article[i].value == j:
                            source_page.cell(
                                row=i+1, column=6).value = self.ozon_article_amount[j]
                if self.yandex_article_amount:
                    for k in self.yandex_article_amount.keys():
                        if name_article[i].value == k:
                            source_page.cell(
                                row=i+1, column=7).value = self.yandex_article_amount[k]

                if name_article[i].value in wb_article_amount.keys():
                    source_page.cell(
                        row=i+1, column=5).value = wb_article_amount[name_article[i].value]
                # Заполняется столбец "В" - номер ячейки на внутреннем складе
                for s in range(len(shelf_seller_article_list)):
                    if name_article[i].value == shelf_seller_article_list[s] and (
                            int(amount_all_fbs[i].value) < int(shelf_amount_list[s])):
                        source_page.cell(
                            row=i+1, column=2).value = shelf_number_list[s]
                        new_amount = int(
                            shelf_amount_list[s]) - int(amount_all_fbs[i].value)
                        select_table_query = f'''UPDATE database_shelvingstocks SET amount={new_amount},
                            task_start_date=current_timestamp, task_finish_date=NULL WHERE seller_article='{shelf_seller_article_list[s]}';'''
                        cursor.execute(select_table_query)
                    elif name_article[i].value == shelf_seller_article_list[s] and (
                            int(amount_all_fbs[i].value) >= int(shelf_amount_list[s])):
                        # ========== Сюда вставить отметку, если мало артикулов в полке =========
                        source_page.cell(
                            row=i+1, column=2).value = f'{shelf_number_list[s]}'
            connection.commit()
            w_b.save(name_pivot_xls)
            w_b2 = load_workbook(name_pivot_xls)
            source_page2 = w_b2.active
            amount_all_fbs = source_page2['D']
            amount_for_production = source_page2['C']
            PROD_DETAIL_CONST = 4
            for r in range(1, len(amount_all_fbs)):
                # Заполняет столбец ['C'] = 'Производство'
                if amount_all_fbs[r].value == 1:
                    source_page2.cell(
                        row=r+1, column=3).value = int(PROD_DETAIL_CONST)
                elif 2 <= int(amount_all_fbs[r].value) <= PROD_DETAIL_CONST-1:
                    source_page2.cell(
                        row=r+1, column=3).value = int(2 * PROD_DETAIL_CONST)
                elif PROD_DETAIL_CONST <= int(amount_all_fbs[r].value) <= 2 * PROD_DETAIL_CONST - 1:
                    source_page2.cell(
                        row=r+1, column=3).value = int(3 * PROD_DETAIL_CONST)
                else:
                    source_page2.cell(row=r+1, column=3).value = ' '
            w_b2.save(name_pivot_xls)
            w_b2 = load_workbook(name_pivot_xls)
            source_page2 = w_b2.active
            amount_all_fbs = source_page2['D']
            al = Alignment(horizontal="center", vertical="center")
            al_left = Alignment(horizontal="left", vertical="center")
            # Задаем толщину и цвет обводки ячейки
            font_bold = Font(bold=True)
            thin = Side(border_style="thin", color="000000")
            thick = Side(border_style="medium", color="000000")
            for i in range(len(amount_all_fbs)):
                for c in source_page2[f'A{i+1}:G{i+1}']:
                    if i == 0:
                        c[0].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[1].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[2].border = Border(top=thick, left=thick,
                                             bottom=thin, right=thick)
                        c[3].border = Border(top=thick, left=thick,
                                             bottom=thin, right=thick)
                        c[4].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[5].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[6].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                    elif i == len(amount_all_fbs)-1:
                        c[0].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[1].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[2].border = Border(top=thin, left=thick,
                                             bottom=thick, right=thick)
                        c[3].border = Border(top=thin, left=thick,
                                             bottom=thick, right=thick)
                        c[4].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[5].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[6].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                    else:
                        c[0].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[1].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[2].border = Border(top=thin, left=thick,
                                             bottom=thin, right=thick)
                        c[3].border = Border(top=thin, left=thick,
                                             bottom=thin, right=thick)
                        c[4].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[5].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                        c[6].border = Border(top=thin, left=thin,
                                             bottom=thin, right=thin)
                    c[0].alignment = al_left
                    c[1].alignment = al
                    c[2].alignment = al
                    c[3].alignment = al
                    c[4].alignment = al
                    c[5].alignment = al
                    c[6].alignment = al
            source_page2.column_dimensions['A'].width = 18
            source_page2.column_dimensions['B'].width = 18
            source_page2.column_dimensions['C'].width = 18
            source_page2.column_dimensions['D'].width = 10
            source_page2.column_dimensions['E'].width = 10
            source_page2.column_dimensions['F'].width = 12
            source_page2.column_dimensions['G'].width = 12

            # Когда понадобится столбец НА ПРОИЗВОДСТВО - удалить следующую строку
            source_page2.delete_cols(3, 1)
            w_b2.save(name_pivot_xls)

            folder_path = os.path.dirname(os.path.abspath(path_file))
            name_for_file = f'Общий файл производство {self.file_add_name} {delivery_date}'
            name_xls_dropbox = f'На производство {self.file_add_name} {delivery_date}'

            output = convert(source=path_file, output_dir=folder_path, soft=1)

            # Сохраняем на DROPBOX
            with open(f'{path_file}', 'rb') as f:
                dbx_db.files_upload(
                    f.read(), f'{self.dropbox_current_assembling_folder}/{name_xls_dropbox}.xlsx')
            with open(output, 'rb') as f:
                dbx_db.files_upload(
                    f.read(), f'{self.dropbox_current_assembling_folder}/{name_for_file}.pdf')
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'create_pivot_xls', self.create_pivot_xls, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def analyze_fbs_amount(self):
        """
        СВОДНЫЙ ФАЙЛ.
        Анализирует количество артикулов в текущей сборке
        """
        try:
            wb_article_amount = self.amount_articles.copy()
            if self.ozon_article_amount:
                for article in self.ozon_article_amount.keys():
                    if article in self.amount_articles.keys():
                        self.amount_articles[article] = int(
                            self.amount_articles[article]) + int(self.ozon_article_amount[article])
                    else:
                        self.amount_articles[article] = int(
                            self.ozon_article_amount[article])
            if self.yandex_article_amount:
                for article in self.yandex_article_amount.keys():
                    if article in self.amount_articles.keys():
                        self.amount_articles[article] = int(
                            self.amount_articles[article]) + int(self.yandex_article_amount[article])
                    else:
                        self.amount_articles[article] = int(
                            self.yandex_article_amount[article])
            sum_all_fbs = sum(self.amount_articles.values())
            sum_fbs_wb = 0
            if wb_article_amount:
                for i in wb_article_amount.values():
                    sum_fbs_wb += int(i)
            sum_fbs_ozon = 0
            if self.ozon_article_amount:
                for i in self.ozon_article_amount.values():
                    sum_fbs_ozon += int(i)
            if len(self.amount_articles) == 0:
                max_amount_all_fbs = 0
                articles_for_fbs = []
                max_article_amount_all_fbs = 0
            else:
                max_amount_all_fbs = max(list(self.amount_articles.values()))
                amount_for_fbs = list(self.amount_articles.values())
                articles_for_fbs = list(self.amount_articles.keys())
                max_article_amount_all_fbs = articles_for_fbs[amount_for_fbs.index(
                    max_amount_all_fbs)]
            return (sum_fbs_wb,
                    sum_fbs_ozon,
                    sum_all_fbs,
                    articles_for_fbs,
                    max_article_amount_all_fbs,
                    max_amount_all_fbs)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'analyze_fbs_amount', self.analyze_fbs_amount, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    def sender_message_to_telegram(self):
        """Отправляет количество артикулов в телеграм бот"""

        try:
            list_chat_id_tg = [CHAT_ID_EU, CHAT_ID_AN]
            sum_fbs_wb, sum_fbs_ozon, sum_all_fbs, articles_for_fbs, max_article_amount_all_fbs, max_amount_all_fbs = self.analyze_fbs_amount()
            ur_lico_for_message = ''
            if self.file_add_name == 'ООО':
                ur_lico_for_message = 'Amstek'
            elif self.file_add_name == 'ИП':
                ur_lico_for_message = '3Д Ночник'
            message = f''' Отправлено на сборку Фбс {ur_lico_for_message}
                ВБ: {sum_fbs_wb}, Озон: {sum_fbs_ozon}
                Итого по ФБС {ur_lico_for_message}: {sum_all_fbs} штук
                В сборке {len(articles_for_fbs)} артикулов
                Артикул с максимальным количеством {max_article_amount_all_fbs}. В сборке {max_amount_all_fbs} штук'''
            message = message.replace('            ', '')
            for chat_id in list_chat_id_tg:
                bot.send_message(chat_id=chat_id, text=message)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'sender_message_to_telegram', self.sender_message_to_telegram, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

# ========== ВЫЗЫВАЕМ ФУНКЦИИ ПО ОЧЕРЕДИ ========== #


# ==================== Сборка WILDBERRIES =================== #


def action_wb(db_folder, file_add_name, headers_wb,
              headers_ozon, headers_yandex):
    wb_actions = WildberriesFbsMode(
        headers_wb, db_folder, file_add_name)

    clearning_folders()
    # =========== АЛГОРИТМ  ДЕЙСТВИЙ С WILDBERRIES ========== #
    # 1. Создаю папку на dropbox для документов
    wb_actions.create_dropbox_folder()
    # 2. Создаю поставку
    wb_actions.create_delivery()
    # 3. добавляю сборочные задания по их id в созданную поставку и получаю qr стикер каждого
    # задания и сохраняю его в папку
    wb_actions.qrcode_order()
    # 4. Создаю лист сборки
    wb_actions.create_selection_list()
    # 5. Создаю шрихкоды для артикулов
    wb_actions.create_barcode_tickets()
    # 6. Добавляю поставку в доставку.
    wb_actions.supply_to_delivery()
    # 7. Получаю QR код поставки
    # и преобразует этот QR код в необходимый формат.
    wb_actions.qrcode_supply()
    # 8. Создаю список с полными именами файлов, которые нужно объединить
    wb_actions.list_for_print_create()


# =========== Сборка ОЗОН ========== #
def action_ozon_ooo(ozon_headers, db_folder, file_add_name):
    ozon_actions = OzonFbsMode(ozon_headers, db_folder, file_add_name)
    # 1. Собираю информацию о новых заказах с Озон.
    ozon_actions.awaiting_packaging_orders()
    # 2. Делю заказ на отправления и перевожу его в статус awaiting_deliver.
    ozon_actions.awaiting_deliver_orders()
    # 3. Готовлю данные для подтверждения отгрузки
    ozon_actions.prepare_data_for_confirm_delivery()
    # 4. Создает лист подбора для отправки
    ozon_actions.create_ozone_selection_sheet_pdf()
    # 5. Сохраняет этикетки для каждой отправки
    ozon_actions.forming_package_ticket_with_article()
    # 6. Подтверждаю отгрузку и запускаю создание документов на стороне ОЗОН
    ozon_actions.confirm_delivery_create_document()
    # 7. Проверяю, что отгрузка создана. Формирую список отправлений для дальнейшей работы
    ozon_actions.check_delivery_create()
    # 8. Проверяю статус формирования накладной.
    # Получаю файлы с этикетками для коробок и этикетки для каждой отправки
    ozon_actions.check_status_formed_invoice()


def action_ozon_ip_morning(ozon_headers, db_folder, file_add_name):
    ozon_actions = OzonFbsMode(ozon_headers, db_folder, file_add_name)
    # 1. Собираю информацию о новых заказах с Озон.
    ozon_actions.awaiting_packaging_orders()
    # 2. Делю заказ на отправления и перевожу его в статус awaiting_deliver.
    ozon_actions.awaiting_deliver_orders()
    # 3. Готовлю данные для подтверждения отгрузки
    ozon_actions.prepare_data_for_confirm_delivery()
    # 4. Создает лист подбора для отправки
    ozon_actions.create_ozone_selection_sheet_pdf()
    # 5. Сохраняет этикетки для каждой отправки
    ozon_actions.forming_package_ticket_with_article()


def action_ozon_ip_day(ozon_headers, db_folder, file_add_name):
    ozon_actions = OzonFbsMode(ozon_headers, db_folder, file_add_name)
    # 1. Готовлю данные для подтверждения отгрузки
    ozon_actions.prepare_data_for_confirm_delivery()
    # 2. Подтверждаю отгрузку и запускаю создание документов на стороне ОЗОН
    ozon_actions.confirm_delivery_create_document()
    # 3. Проверяю, что отгрузка создана. Формирую список отправлений для дальнейшей работы
    ozon_actions.check_delivery_create()
    # 4. Проверяю статус формирования накладной.
    # Получаю файлы с этикетками для коробок и этикетки для каждой отправки
    ozon_actions.check_status_formed_invoice()
    # Очищаем все папки на сервере


# =========== АЛГОРИТМ ДЕЙСТВИЙ С ЯНДЕКС ========== #
def action_yandex(yandex_headers, db_folder, file_add_name):
    yandex_actions = YandexMarketFbsMode(
        yandex_headers, db_folder, file_add_name)
    # 1. Меняет статус ордеров
    yandex_actions.change_orders_status()
    # 2. Формирует файл подбора
    yandex_actions.create_yandex_selection_sheet_pdf()
    # 3. Подтверждение отгрузки
    yandex_actions.approve_shipment()
    # 4. Сохраняем акт
    yandex_actions.saving_act()
    # 5. Сохраняем этикетки
    yandex_actions.saving_tickets()


def production_file(db_folder, file_add_name, headers_wb,
                    headers_ozon, headers_yandex):

    # =========== СОЗДАЮ СВОДНЫЙ ФАЙЛ ========== #
    pivot_file = CreatePivotFile(db_folder, file_add_name,
                                 headers_wb, headers_ozon,
                                 headers_yandex)
    # 1. Создаю сводный файл для производства
    pivot_file.create_pivot_xls()
    # 2. Отправляю данные по сборке FBS
    pivot_file.sender_message_to_telegram()

    message_text = f'Сборка {file_add_name} сформирована'
    bot.send_message(chat_id=CHAT_ID_MANAGER,
                     text=message_text, parse_mode='HTML')
    bot.send_message(chat_id=CHAT_ID_ADMIN,
                     text=message_text, parse_mode='HTML')


@app.task
def ooo_production_file():
    production_file(
        db_folder, file_add_name_ooo, wb_headers_ooo,
        ozon_headers_ooo, yandex_headers_ooo)


@app.task
def ooo_wb_action():
    action_wb(
        db_folder, file_add_name_ooo, wb_headers_ooo,
        ozon_headers_ooo, yandex_headers_ooo)


# ooo_wb_action()

@app.task
def ooo_ozon_action():
    action_ozon_ooo(ozon_headers_ooo, db_folder, file_add_name_ooo)


# ooo_ozon_action()

@app.task
def ooo_yandex_action():
    action_yandex(yandex_headers_ooo, db_folder, file_add_name_ooo)


# ooo_yandex_action()

@app.task
def ip_wb_action():
    action_wb(
        db_folder, file_add_name_ip, wb_headers_karavaev,
        ozon_headers_karavaev, yandex_headers_karavaev)


@app.task
def ip_production_file():
    production_file(
        db_folder, file_add_name_ip, wb_headers_karavaev,
        ozon_headers_karavaev, yandex_headers_karavaev)

# ip_wb_action()


@app.task
def ip_ozon_action_morning():
    action_ozon_ip_morning(ozon_headers_karavaev,
                           db_folder, file_add_name_ip)


# ip_ozon_action_morning()

@app.task
def ip_ozon_action_day():
    action_ozon_ip_day(ozon_headers_karavaev, db_folder, file_add_name_ip)


@app.task
def ip_yandex_action():
    action_yandex(yandex_headers_karavaev, db_folder, file_add_name_ip)


# ip_yandex_action()
