import base64
import glob
import io
import json
import logging
import math
import os
import shutil
import smtplib
import textwrap
import time
import traceback
from collections import Counter
from contextlib import closing
from datetime import datetime, timedelta
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO
from pathlib import Path

import dropbox
import openpyxl
import pandas as pd
import psycopg2
import requests
import telegram
from celery_tasks.celery import app
from dotenv import load_dotenv
from msoffice2pdf import convert
from openpyxl import Workbook, load_workbook
from openpyxl.drawing import image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont
from price_system.supplyment import sender_error_to_tg
from sqlalchemy import create_engine

from web_barcode.constants_file import (CHAT_ID_ADMIN, CHAT_ID_AN, CHAT_ID_EU,
                                        CHAT_ID_MANAGER, EMAIL_ADDRESS_FROM,
                                        EMAIL_ADDRESS_FROM_PASSWORD,
                                        EMAIL_ADDRESS_TO,
                                        MANAGER_EMAIL_ADDRESS,
                                        NKS_EMAIL_ADDRESS, POST_PORT,
                                        POST_SERVER, bot, dbx_db,
                                        header_wb_dict, wb_headers_karavaev,
                                        wb_headers_ooo)

from .helpers import (design_barcodes_dict_spec,
                      merge_barcode_for_ozon_two_on_two,
                      merge_barcode_for_yandex_two_on_two,
                      new_data_for_ozon_ticket, new_data_for_yandex_ticket,
                      print_barcode_to_pdf2,
                      print_barcode_to_pdf_without_dropbox,
                      qrcode_print_for_products,
                      supply_qrcode_to_standart_view)

logging.basicConfig(level=logging.DEBUG,
                    filename="tasks_log.log",
                    format="%(asctime)s %(levelname)s %(message)s")

db_folder = '/!На производство'

file_add_name_ip = 'IP'
file_add_name_ooo = 'OOO'


def stream_dropbox_file(path):
    _, res = dbx_db.files_download(path)
    with closing(res) as result:
        byte_data = result.content
        return io.BytesIO(byte_data)


def clearning_folders():
    dirs = ['fbs_mode/data_for_barcodes/cache_dir',
            'fbs_mode/data_for_barcodes/done_data',
            'fbs_mode/data_for_barcodes/pivot_excel',
            'fbs_mode/data_for_barcodes/qrcode_folder',
            'fbs_mode/data_for_barcodes/qrcode_supply',
            'fbs_mode/data_for_barcodes/package_tickets',
            'fbs_mode/data_for_barcodes/ozon_docs',
            'fbs_mode/data_for_barcodes/ozon_delivery_barcode',
            'fbs_mode/data_for_barcodes/yandex',
            'fbs_mode/data_for_barcodes/new_pivot_excel',
            'fbs_mode/data_for_barcodes/photo',
            'fbs_mode/data_for_barcodes/raw_photo'
            ]
    for dir in dirs:
        for filename in glob.glob(os.path.join(dir, "*")):
            file_path = os.path.join(dir, filename)
            try:
                if os.path.isfile(filename) or os.path.islink(filename):
                    os.unlink(filename)
                elif os.path.isdir(filename):
                    shutil.rmtree(filename)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (filename, e))


def wb_clearning_folders():
    dirs = ['fbs_mode/data_for_barcodes/cache_dir',
            'fbs_mode/data_for_barcodes/done_data',
            'fbs_mode/data_for_barcodes/pivot_excel',
            'fbs_mode/data_for_barcodes/qrcode_folder',
            'fbs_mode/data_for_barcodes/qrcode_supply'
            ]
    for dir in dirs:
        for filename in glob.glob(os.path.join(dir, "*")):
            file_path = os.path.join(dir, filename)
            try:
                if os.path.isfile(filename) or os.path.islink(filename):
                    os.unlink(filename)
                elif os.path.isdir(filename):
                    shutil.rmtree(filename)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (filename, e))


def error_message(function_name: str, function, error_text: str) -> str:
    """
    Формирует текст ошибки и выводит информацию в модальном окне
    function_name - название функции.
    function - сама функция, как объект, а не результат работы. Без скобок.
    error_text - текст ошибки.
    """
    tb_str = traceback.format_exc()
    message_error = (f'Ошибка в функции: <b>{function_name}</b>\n'
                     f'<b>Функция выполняет</b>: {function.__doc__}\n'
                     f'<b>Ошибка</b>\n: {error_text}\n\n'
                     f'<b>Техническая информация</b>:\n {tb_str}')
    return message_error


class WildberriesFbsMode():
    """Класс отвечает за работу с заказами Wildberries"""

    def __init__(self, headers, db_forder, file_add_name):
        """Основные данные класса"""
        self.amount_articles = {}
        self.dropbox_main_fbs_folder = db_forder
        self.headers = headers
        self.file_add_name = file_add_name
        self.files_for_send = []

    def check_folder_exists(self, path):
        try:
            dbx_db.files_list_folder(path)
            return True
        except dropbox.exceptions.ApiError as e:
            return False

    @sender_error_to_tg
    def check_folder_availability(self, folder):
        """Проверяет наличие папки, если ее нет, то создает"""
        folder_path = os.path.join(
            os.getcwd(), folder)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

    @sender_error_to_tg
    def convert_photo_from_webp_to_jpg(self, photo_address, article_barcode, name_raw_photo_folder, name_photo_folder):
        """Конвертирует webp изображение в jpg и сохраняет в нужных папках."""
        response = requests.get(photo_address)
        image = Image.open(BytesIO(response.content))
        name_raw_photo = f"{name_raw_photo_folder}/{article_barcode}.jpg"
        image.save(name_raw_photo)
        # Открываем изображение в формате We
        webp_image = Image.open(name_raw_photo)
        name_photo = f"{name_photo_folder}/{article_barcode}.jpg"
        # Сохраняем изображение в формате JPEG
        webp_image.save(name_photo, 'JPEG')
        return name_photo

    @sender_error_to_tg
    def process_new_orders(self):
        """
        WILDBERRIES
        Функция обрабатывает новые сборочные задания.
        """
        url = "https://suppliers-api.wildberries.ru/api/v3/orders/new"
        response = requests.request(
            "GET", url, headers=self.headers)
        if response.status_code == 200:
            returned_data_list = []
            orders_data = json.loads(response.text)['orders']
            # Сделал обход склада в НСК (его id = 1003917). Если склад = НСК, то товары не участвуют в сборке.
            for data in orders_data:
                if data['warehouseId'] == 1003917 or data['warehouseId'] == 1057680:
                    returned_data_list.append(data)
            return returned_data_list
        else:
            time.sleep(10)
            return self.process_new_orders()

    @sender_error_to_tg
    def time_filter_orders(self):
        """
        WILDBERRIES
        Функция фильтрует новые заказы по времени.
        Если заказ создан меньше, чем 1 час назад, в работу он не берется
        """
        orders_data = self.process_new_orders()
        filters_order_data = []
        now_time = datetime.now()
        for order in orders_data:
            # Время создания заказа в переводе с UTC на московское
            create_order_time = datetime.strptime(
                order['createdAt'], '%Y-%m-%dT%H:%M:%SZ') + timedelta(hours=3)
            delta_order_time = now_time - create_order_time
            if delta_order_time > timedelta(hours=1):
                filters_order_data.append(order)
        return filters_order_data

    @sender_error_to_tg
    def article_info(self, article):
        """
        WILDBERRIES
        Получает развернутый ответ про каждый артикул. 
        """
        url_data = "https://suppliers-api.wildberries.ru/content/v2/get/cards/list"
        payload = json.dumps({
            "settings": {
                "cursor": {
                    "limit": 1
                },
                "filter": {
                    "textSearch": article,
                    "withPhoto": -1
                }
            }
        })
        response_data = requests.request(
            "POST", url_data, headers=self.headers, data=payload)
        if response_data.status_code == 200:
            return response_data.text
        else:
            text = f'Статус код = {response_data.status_code} у артикула {article}'
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=text, parse_mode='HTML')
            time.sleep(5)
            return self.article_info(article)

    @sender_error_to_tg
    def article_data_for_tickets(self):
        """
        WILDBERRIES
        Выделяет артикулы продавца светильников, их баркоды и наименования.
        Создает словарь с данными каждого артикулы и словарь с количеством каждого
        артикула. 
        """
        orders_data = self.time_filter_orders()
        # Словарь с данными артикула: {артикул_продавца: [баркод, наименование]}
        self.data_article_info_dict = {}
        # Список только для артикулов ночников. Остальные отфильтровывает
        self.clear_article_list = []
        # Словарь для данных листа подбора {order_id: [photo_link, brand, name, seller_article]}
        self.selection_dict = {}
        # Папка для сохранения фото в webp формате
        name_raw_photo_folder = f"fbs_mode/data_for_barcodes/raw_photo"
        self.check_folder_availability(name_raw_photo_folder)
        # Папка для сохранения фото в jpg формате
        name_photo_folder = f"fbs_mode/data_for_barcodes/photo"
        self.check_folder_availability(name_photo_folder)
        CATEGORY_LIST = ['Ночники', 'Светильники']
        for data in orders_data:
            answer = self.article_info(data['article'])
            if json.loads(answer)['cards']:
                if json.loads(answer)['cards'][0]['subjectName'] in CATEGORY_LIST:
                    self.clear_article_list.append(data['article'])
                    # Достаем баркод артикула (первый из списка, если их несколько)
                    barcode = json.loads(answer)[
                        'cards'][0]['sizes'][0]['skus'][0]
                    # Достаем название артикула
                    title = json.loads(answer)[
                        'cards'][0]['title']
                    self.data_article_info_dict[data['article']] = [
                        title, barcode]
                    photo = json.loads(answer)[
                        'cards'][0]['photos'][1]['big']
                    photo_name = self.convert_photo_from_webp_to_jpg(
                        photo, barcode, name_raw_photo_folder, name_photo_folder)
                    title_article = json.loads(answer)[
                        'cards'][0]['title']
                    seller_article = data['article']
                    # Заполняем словарь данными для Листа подбора
                    self.selection_dict[data['id']] = [
                        title_article, seller_article]
            time.sleep(2)
        # Словарь с данными: {артикул_продавца: количество}
        self.amount_articles = dict(Counter(self.clear_article_list))
        return self.amount_articles

    @sender_error_to_tg
    def create_barcode_tickets(self):
        """WILDBERRIES. Функция создает этикетки со штрихкодами для артикулов"""
        if self.clear_article_list and self.data_article_info_dict:
            design_barcodes_dict_spec(
                self.clear_article_list, self.data_article_info_dict)
        else:
            text = 'не сработала create_barcode_tickets так как нет данных'
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=text, parse_mode='HTML')

    @sender_error_to_tg
    def create_delivery(self):
        """WILDBERRIES. Создание поставки"""
        amount_articles = self.article_data_for_tickets()
        if amount_articles:
            delivery_date = datetime.today().strftime("%d.%m.%Y %H-%M-%S")
            url_data = 'https://suppliers-api.wildberries.ru/api/v3/supplies'
            hour = datetime.now().hour
            delivery_name = f"Поставка {delivery_date}"
            if hour >= 6 and hour < 18:
                delivery_name = f'НСК День {delivery_date}'
            else:
                delivery_name = f'НСК Ночь {delivery_date}'
            payload = json.dumps(
                {"name": delivery_name}
            )
            # Из этой переменной достать ID поставки
            response_data = requests.request(
                "POST", url_data, headers=self.headers, data=payload)
            # print(response_data)
            self.supply_id = json.loads(response_data.text)['id']
        else:
            text = f'Нет артикулов на WB для сборки {self.file_add_name}'
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=text, parse_mode='HTML')

    def add_shelf_number_to_selection_dict(self, selection_dict: dict):
        """Добавляет полку хранения в словарь с данными отправления"""
        SHELVES_DATA = '/DATABASE/Стеллажи Новосибирск.xlsx'
        db_file_data = stream_dropbox_file(SHELVES_DATA)
        pd_file_data = pd.read_excel(db_file_data)
        special_tickets_data_file = pd.DataFrame(
            pd_file_data, columns=['Артикул', 'Ячейка'])
        articles_list = special_tickets_data_file['Артикул'].to_list()
        shelves_list = special_tickets_data_file['Ячейка'].to_list()
        article_address = {}
        for i in range(len(articles_list)):
            article = str(articles_list[i])
            address = str(shelves_list[i])
            if str(articles_list[i]) != 'nan':
                if article.capitalize() in article_address:
                    article_address[article.capitalize()] += f'\n{address}'
                else:
                    article_address[article.capitalize()] = address
        for _, order_data in selection_dict.items():
            if order_data[3].capitalize() in article_address:
                order_data.append(article_address[order_data[3].capitalize()])
            else:
                order_data.append('')
        return selection_dict

    @sender_error_to_tg
    def qrcode_order(self):
        """
        WILDBERRIES.
        Функция добавляет сборочные задания по их id
        в созданную поставку и получает qr стикер каждого
        задания и сохраняет его в папку
        """
        if self.supply_id:
            # Добавляем заказы в поставку
            for order in self.selection_dict.keys():
                add_url = f'https://suppliers-api.wildberries.ru/api/v3/supplies/{self.supply_id}/orders/{order}'
                response_add_orders = requests.request(
                    "PATCH", add_url, headers=self.headers)
            # Создаем qr коды добавленных ордеров.
            for order in self.selection_dict.keys():
                ticket_url = 'https://suppliers-api.wildberries.ru/api/v3/orders/stickers?type=png&width=40&height=30'
                payload_ticket = json.dumps({"orders": [order]})
                response_ticket = requests.request(
                    "POST", ticket_url, headers=self.headers, data=payload_ticket)
                # Расшифровываю ответ, чтобы сохранить файл этикетки задания
                ticket_data = json.loads(response_ticket.text)[
                    "stickers"][0]["file"]
                # Узнаю стикер сборочного задания и помещаю его в словарь с данными для
                # листа подбора
                sticker_code_first_part = json.loads(response_ticket.text)[
                    "stickers"][0]["partA"]
                sticker_code_second_part = json.loads(response_ticket.text)[
                    "stickers"][0]["partB"]
                sticker_code = f'{sticker_code_first_part} {sticker_code_second_part}'
                self.selection_dict[order].append(sticker_code)
                # декодируем строку из base64 в бинарные данные
                binary_data = base64.b64decode(ticket_data)
                # создаем объект изображения из бинарных данных
                img = Image.open(BytesIO(binary_data))
                # сохраняем изображение в файл
                folder_path = os.path.join(
                    os.getcwd(), "fbs_mode/data_for_barcodes/qrcode_folder")
                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)
                img.save(
                    f"{folder_path}/{order} {self.selection_dict[order][-2]}.png")
        else:
            text = 'не сработала qrcode_order из за отсутвия self.supply_id'
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=text, parse_mode='HTML')

    @sender_error_to_tg
    def create_selection_list(self):
        """WILDBERRIES. Создает лист сборки"""
        if self.selection_dict:
            self.selection_dict = self.add_shelf_number_to_selection_dict(
                self.selection_dict)
            sorted_dict = dict(
                sorted(self.selection_dict.items(), key=lambda item: item[1][1]))
            delivery_date = datetime.today().strftime("%d-%m-%Y_%H-%M-%S")
            # создаем новую книгу Excel
            selection_file = Workbook()
            COUNT_HELPER = 2
            # выбираем лист Sheet1
            create = selection_file.create_sheet(
                title='pivot_list', index=0)
            sheet = selection_file['pivot_list']
            # Установка параметров печати
            create.page_setup.paperSize = create.PAPERSIZE_A4
            create.page_setup.orientation = create.ORIENTATION_PORTRAIT
            create.page_margins.left = 0.25
            create.page_margins.right = 0.25
            create.page_margins.top = 0.25
            create.page_margins.bottom = 0.25
            create.page_margins.header = 0.3
            create.page_margins.footer = 0.3
            sheet['A1'] = '№ Задания'
            sheet['B1'] = 'Наименование'
            sheet['C1'] = 'Артикул продавца'
            sheet['D1'] = 'Стикер'
            sheet['E1'] = 'Ячейка'
            for key, value in sorted_dict.items():
                create.cell(row=COUNT_HELPER, column=1).value = key
                create.cell(row=COUNT_HELPER, column=2).value = value[1]
                create.cell(row=COUNT_HELPER, column=3).value = value[2]
                create.cell(row=COUNT_HELPER, column=4).value = value[3]
                create.cell(row=COUNT_HELPER, column=6).value = value[4]
                COUNT_HELPER += 1
            folder_path = os.path.join(
                os.getcwd(), 'fbs_mode/data_for_barcodes/pivot_excel')
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            name_selection_file = f'{folder_path}/NSK_WB_{self.file_add_name}_Selection_list_{delivery_date}.xlsx'
            path_file = os.path.abspath(name_selection_file)
            selection_file.save(name_selection_file)
            w_b2 = load_workbook(name_selection_file)
            source_page2 = w_b2.active
            al = Alignment(horizontal="center", vertical="center")
            al_left = Alignment(horizontal="left",
                                vertical="center", wrapText=True)
            source_page2.column_dimensions['A'].width = 16  # Номер задания
            source_page2.column_dimensions['B'].width = 7  # Картинка
            source_page2.column_dimensions['C'].width = 16  # Бренд
            source_page2.column_dimensions['D'].width = 25  # Наименование
            # Артикул продавца
            source_page2.column_dimensions['E'].width = 16
            thin = Side(border_style="thin", color="000000")
            for i in range(len(sorted_dict)+1):
                for c in source_page2[f'A{i+1}:E{i+1}']:
                    c[0].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[0].font = Font(size=12)
                    c[0].alignment = al

                    c[1].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[1].font = Font(size=12)
                    c[1].alignment = al

                    c[2].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[2].font = Font(size=12)
                    c[2].alignment = al_left

                    c[3].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[3].font = Font(size=12)
                    c[3].alignment = al_left

                    c[4].border = Border(top=thin, left=thin,
                                         bottom=thin, right=thin)
                    c[4].font = Font(size=12, bold=True)
                    c[4].alignment = al_left

            w_b2.save(name_selection_file)
            folder_path = os.path.dirname(os.path.abspath(path_file))
            output = convert(source=path_file,
                             output_dir=folder_path, soft=1)
            self.files_for_send.append(name_selection_file)
        else:
            text = 'Не сработала create_selection_list потому что нет self.selection_dict'
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=text, parse_mode='HTML')

    @sender_error_to_tg
    def supply_to_delivery(self, numb=0):
        """
        WILDBERRIES.
        Функция добавляет поставку в доставку.
        """
        if self.supply_id:
            time.sleep(30)
            # Переводим поставку в доставку
            url_to_supply = f'https://suppliers-api.wildberries.ru/api/v3/supplies/{self.supply_id}/deliver'
            response_to_supply = requests.request(
                "PATCH", url_to_supply, headers=self.headers)
            if response_to_supply.status_code != 204 and numb < 10:
                numb += 1
                self.supply_to_delivery(numb)
            elif response_to_supply.status_code != 204 and numb >= 10:
                text = f'Не получилось перевести поставку в доставку, поэтому не будет QR-кода. Статус код {response_to_supply.status_code}. {response_to_supply.text}'
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                                 text=text, parse_mode='HTML')
        else:
            text = 'Поставка не добавлена в доставку (supply_to_delivery), так как нет артикулов'
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=text, parse_mode='HTML')

    @sender_error_to_tg
    def qrcode_supply(self):
        """
        WILDBERRIES.
        Функция получает QR код поставки
        и преобразует этот QR код в необходимый формат.
        """
        if self.supply_id:
            time.sleep(60)
            # Получаем QR код поставки:
            url_supply_qrcode = f"https://suppliers-api.wildberries.ru/api/v3/supplies/{self.supply_id}/barcode?type=png"
            response_supply_qrcode = requests.request(
                "GET", url_supply_qrcode, headers=self.headers)
            # Создаем QR код поставки
            qrcode_base64_data = json.loads(
                response_supply_qrcode.text)["file"]
            # декодируем строку из base64 в бинарные данные
            binary_data = base64.b64decode(qrcode_base64_data)
            # создаем объект изображения из бинарных данных
            img = Image.open(BytesIO(binary_data))
            # сохраняем изображение в файл
            folder_path = os.path.join(
                os.getcwd(), 'fbs_mode/data_for_barcodes/qrcode_supply')
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            img.save(
                f"{folder_path}/{self.supply_id}.png")
        else:
            text = 'Поставка не сформирована (qrcode_supply), так как нет артикулов'
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=text, parse_mode='HTML')

    @sender_error_to_tg
    def list_for_print_create(self):
        """
        WILDBERRIES.
        Функция создает список с полными именами файлов, которые нужно объединить
        amount_articles - словарь с данными {артикул_продавца: количество}.
        Объединяет эти файлы и сохраняет конечный файл на дропбоксе.
        """
        if self.amount_articles:
            qrcode_list = qrcode_print_for_products()
            pdf_filenames = glob.glob(
                'fbs_mode/data_for_barcodes/cache_dir/*.pdf')
            logging.info(f"len(pdf_filenames): {len(pdf_filenames)}")

            list_pdf_file_ticket_for_complect = []
            for j in pdf_filenames:
                while self.amount_articles[str(Path(j).stem)] > 0:
                    list_pdf_file_ticket_for_complect.append(j)
                    self.amount_articles[str(Path(j).stem)] -= 1
            logging.info(
                f"list_pdf_file_ticket_for_complect после добавления количества: {list_pdf_file_ticket_for_complect}")
            # Определяем число qr кодов для поставки.
            amount_of_supply_qrcode = math.ceil(
                len(list_pdf_file_ticket_for_complect)/20)
            for file in qrcode_list:
                list_pdf_file_ticket_for_complect.append(file)
            logging.info(
                f"list_pdf_file_ticket_for_complect после добавления qr кодов: {list_pdf_file_ticket_for_complect}")
            outer_list = []  # Внешний список для процесса сортировки
            for i in list_pdf_file_ticket_for_complect:
                # Разделяю полное название файла на путь к файлу и имя файла
                # Оказывается в python знаком \ отделяется последняя папка перед файлом
                # А все внешние отделяются знаком /
                last_slash_index = i.rfind("/")
                result = [i[:last_slash_index], i[last_slash_index+1:]]
                new_name = result
                full_new_name = []  # Список с полным именени файла после разделения
                # Имена QR кодов у меня составные. Состоят из нескольких слов с пробелами
                # В этом цикле разделяю имена из предыдущих списков по пробелу.
                for j in new_name:
                    split_name = j.split(' ')
                    full_new_name.append(split_name)
                outer_list.append(full_new_name)
            # Сортирую самый внешний список по последнему элемену самого внутреннего списка
            sorted_list = sorted(outer_list, key=lambda x: x[-1][-1])
            # Далее идет обратный процесс - процесс объединения элементов списка
            # в первоначальные имена файлов, но уже отсортированные
            new_sort = []
            for i in sorted_list:
                inner_new_sort = []
                for j in i:
                    j = ' '.join(j)
                    inner_new_sort.append(j)
                new_sort.append(inner_new_sort)
            last_sorted_list = []
            for i in new_sort:
                i = '/'.join(i)
                last_sorted_list.append(i)
            list_pdf_file_ticket_for_complect = last_sorted_list
            logging.info(
                f"list_pdf_file_ticket_for_complect перед группировкой файлов: {list_pdf_file_ticket_for_complect}")

            qrcode_supply_amount = supply_qrcode_to_standart_view()
            if len(qrcode_supply_amount) != 0:
                while amount_of_supply_qrcode > 0:
                    list_pdf_file_ticket_for_complect.append(
                        qrcode_supply_amount[0])
                    amount_of_supply_qrcode -= 1
            folder_path = os.path.join(
                os.getcwd(), 'fbs_mode/data_for_barcodes/done_data')
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            file_name = (
                f'{folder_path}/NSK_WB_{self.file_add_name}_tickets_FBS_{time.strftime("%Y-%m-%d_%H-%M-%S")}.pdf')
            print_barcode_to_pdf_without_dropbox(list_pdf_file_ticket_for_complect,
                                                 file_name)
            self.files_for_send.append(file_name)
        else:
            text = 'не сработала list_for_print_create потому что нет данных'
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=text, parse_mode='HTML')

    @sender_error_to_tg
    def send_email(self):
        # Настройки
        delivery_date = datetime.today().strftime("%d-%m-%Y %H-%M-%S")
        smtp_server = POST_SERVER  # Замените на SMTP-сервер вашего провайдера
        smtp_port = POST_PORT  # Обычно 587 для TLS
        username = EMAIL_ADDRESS_FROM
        password = EMAIL_ADDRESS_FROM_PASSWORD
        from_email = EMAIL_ADDRESS_FROM
        addresses_for_email = [EMAIL_ADDRESS_TO,
                               NKS_EMAIL_ADDRESS, MANAGER_EMAIL_ADDRESS]
        subject = f'Сборка {self.file_add_name} {delivery_date}'

        # Создаем сообщение
        for to_email in addresses_for_email:
            msg = MIMEMultipart()
            msg['From'] = from_email
            msg['To'] = to_email
            msg['Subject'] = subject
            if self.files_for_send:
                body = f'Сборка {self.file_add_name} {delivery_date}'
                msg.attach(MIMEText(body, 'plain'))
                for filename in self.files_for_send:
                    try:
                        # Открываем файл в бинарном режиме
                        with open(filename, 'rb') as attachment:
                            # Создаем вложение
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload(attachment.read())
                            encoders.encode_base64(part)
                            part.add_header('Content-Disposition',
                                            f'attachment; filename*=UTF-8\'\'{filename}')

                            # Добавляем вложение в сообщение
                            msg.attach(part)
                    except Exception as e:
                        print(f'Не удалось прикрепить файл {filename}: {e}')
            else:
                body = f'В сборке {self.file_add_name} {delivery_date} нет товаров'
                msg.attach(MIMEText(body, 'plain'))
            # Отправляем письмо
            try:
                server = smtplib.SMTP(smtp_server, smtp_port)
                server.starttls()  # Начинаем TLS шифрование
                server.login(username, password)
                server.send_message(msg)
            except Exception as e:
                print(f'Ошибка: {e}')
            finally:
                server.quit()

    @sender_error_to_tg
    def analyze_fbs_amount(self):
        """
        СВОДНЫЙ ФАЙЛ.
        Анализирует количество артикулов в текущей сборке
        """

        try:
            sum_all_fbs = sum(self.amount_articles.values())
            if not self.amount_articles:
                max_amount_all_fbs = 0
                articles_for_fbs = []
                max_article_amount_all_fbs = 0
            else:
                max_amount_all_fbs = max(list(self.amount_articles.values()))
                amount_for_fbs = list(self.amount_articles.values())
                articles_for_fbs = list(self.amount_articles.keys())
                max_article_amount_all_fbs = articles_for_fbs[amount_for_fbs.index(
                    max_amount_all_fbs)]
            return (sum_all_fbs,
                    articles_for_fbs,
                    max_article_amount_all_fbs,
                    max_amount_all_fbs)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'analyze_fbs_amount', self.analyze_fbs_amount, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

    @sender_error_to_tg
    def sender_message_to_telegram(self):
        """Отправляет количество артикулов в телеграм бот"""

        try:
            list_chat_id_tg = [CHAT_ID_EU, CHAT_ID_AN]
            sum_all_fbs, articles_for_fbs, max_article_amount_all_fbs, max_amount_all_fbs = self.analyze_fbs_amount()
            ur_lico_for_message_dict = {
                'OOO': 'Amstek',
                'IP': '3Д Ночник'
            }
            message = f'''Отправлено на сборку Фбс в Новосибирске {ur_lico_for_message_dict[self.file_add_name]}
                Итого по ФБС WB {ur_lico_for_message_dict[self.file_add_name]}: {sum_all_fbs} штук
                В сборке {len(articles_for_fbs)} артикулов
                Артикул с максимальным количеством {max_article_amount_all_fbs}. В сборке {max_amount_all_fbs} штук'''
            message = message.replace('            ', '')
            for chat_id in list_chat_id_tg:
                bot.send_message(chat_id=chat_id, text=message)
        except Exception as e:
            # обработка ошибки и отправка сообщения через бота
            message_text = error_message(
                'sender_message_to_telegram', self.sender_message_to_telegram, e)
            bot.send_message(chat_id=CHAT_ID_ADMIN,
                             text=message_text, parse_mode='HTML')

# ========== ВЫЗЫВАЕМ ФУНКЦИИ ПО ОЧЕРЕДИ ========== #


# ==================== Сборка WILDBERRIES =================== #

@sender_error_to_tg
def action_wb(db_folder, file_add_name, headers_wb):
    wb_actions = WildberriesFbsMode(
        headers_wb, db_folder, file_add_name)

    clearning_folders()
    # =========== АЛГОРИТМ  ДЕЙСТВИЙ С WILDBERRIES ========== #
    # 1. Создаю поставку
    wb_actions.create_delivery()
    # 2. добавляю сборочные задания по их id в созданную поставку и получаю qr стикер каждого
    # задания и сохраняю его в папку
    wb_actions.qrcode_order()
    # 3. Создаю лист сборки
    wb_actions.create_selection_list()
    # 4. Создаю шрихкоды для артикулов
    wb_actions.create_barcode_tickets()
    # 5. Добавляю поставку в доставку.
    wb_actions.supply_to_delivery()
    # 6. Получаю QR код поставки
    # и преобразует этот QR код в необходимый формат.
    wb_actions.qrcode_supply()
    # 7. Создаю список с полными именами файлов, которые нужно объединить
    wb_actions.list_for_print_create()
    # 8. Отрпавляю документы на элктронный адрес
    wb_actions.send_email()
    # 9. Отрпавляю данные о сборке в ТГ
    wb_actions.sender_message_to_telegram()


@app.task
def nsk_fbs_task():
    """Запускает утреннюю FBS сборку ИП (Без документов ОЗОН)"""
    try:
        action_wb(
            db_folder, file_add_name_ip, wb_headers_karavaev)
        message_text = f'Сборка Новосибирск {file_add_name_ip} сформирована'
        bot.send_message(chat_id=CHAT_ID_MANAGER,
                         text=message_text, parse_mode='HTML')
        bot.send_message(chat_id=CHAT_ID_ADMIN,
                         text=message_text, parse_mode='HTML')
    except:
        text = f'Приложение fbs_mode. Не сработала функция action_wb для ИП'
        bot.send_message(chat_id=CHAT_ID_ADMIN,
                         text=text, parse_mode='HTML')

    try:
        time.sleep(60)
        action_wb(
            db_folder, file_add_name_ooo, wb_headers_ooo)
        message_text = f'Сборка Новосибирск {file_add_name_ooo} сформирована'
        bot.send_message(chat_id=CHAT_ID_MANAGER,
                         text=message_text, parse_mode='HTML')
        bot.send_message(chat_id=CHAT_ID_ADMIN,
                         text=message_text, parse_mode='HTML')
    except:
        text = f'Приложение fbs_mode. Не сработала функция action_wb для ООО'
        bot.send_message(chat_id=CHAT_ID_ADMIN,
                         text=text, parse_mode='HTML')
