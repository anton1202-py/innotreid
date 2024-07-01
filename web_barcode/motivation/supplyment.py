from datetime import datetime, timedelta

import pandas as pd
from database.models import CodingMarketplaces, OzonSales
from django.http import HttpResponse
# from celery_tasks.celery import app
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from price_system.models import Articles, DesignUser
from users.models import InnotreidUser

from .models import Selling


def get_main_sales_data(year):
    """Отдает данные по продажам артикулов"""

    sale_data = Selling.objects.filter(year=year).values(
        'lighter', 'month', 'quantity', 'summ')
    # Словарь с данными артикула по продажам по месяцам
    main_sales_dict = {}
    # Словарь с продажами артикула за текущий год
    year_sales_dict = {}
    for data in sale_data:
        if data['lighter'] in main_sales_dict:
            if data['month'] in main_sales_dict[data['lighter']]:
                main_sales_dict[data['lighter']
                                ][data['month']] += int(data['quantity'])
            else:
                main_sales_dict[data['lighter']
                                ][data['month']] = int(data['quantity'])
        else:
            main_sales_dict[data['lighter']] = {
                data['month']: int(data['quantity'])}
        if data['lighter'] in year_sales_dict:
            year_sales_dict[data['lighter']] += int(data['quantity'])
        else:
            year_sales_dict[data['lighter']] = int(data['quantity'])
    return year_sales_dict, main_sales_dict


def get_amount_summ_sales_data(year):
    """Отдает данные по продажам артикулов"""
    sale_data = Selling.objects.filter(year=year).values(
        'lighter', 'month', 'quantity', 'summ')
    # Словарь с данными артикула по продажам по месяцам
    main_sales_dict = {}
    # Словарь с продажами артикула за текущий год
    year_sales_dict = {}
    for data in sale_data:
        if data['lighter'] in main_sales_dict:
            if data['month'] in main_sales_dict[data['lighter']]:
                main_sales_dict[data['lighter']
                                ][data['month']]['quantity'] += int(data['quantity'])
                main_sales_dict[data['lighter']
                                ][data['month']]['summ'] += int(data['summ'])
            else:
                main_sales_dict[data['lighter']
                                ][data['month']] = {'quantity': int(data['quantity']),
                                                    'summ': int(data['summ'])}
        else:
            main_sales_dict[data['lighter']] = {
                data['month']: {
                    'quantity': int(data['quantity']),
                    'summ': int(data['summ'])}}
        if data['lighter'] in year_sales_dict:
            year_sales_dict[data['lighter']
                            ]['quantity'] += int(data['quantity'])
            year_sales_dict[data['lighter']]['summ'] += int(data['summ'])
        else:
            year_sales_dict[data['lighter']] = {'quantity': int(data['quantity']),
                                                'summ': int(data['summ'])}
    # print(main_sales_dict)
    return year_sales_dict, main_sales_dict


def get_designers_amount_summ_sales_data(innotreiduser, year):
    """Отдает данные по продажам артикулов"""
    sale_data = Selling.objects.filter(year=year).values(
        'lighter', 'month', 'quantity', 'summ', 'lighter__designer_article', 'lighter__copy_right', 'lighter__designer')
    designer_obj = DesignUser.objects.filter(designer=innotreiduser)[0]
    main_percent = designer_obj.main_reward_persent/100
    copyright_percent = designer_obj.copyright_reward_persent/100

    # Словарь с данными артикула по продажам по месяцам
    main_sales_dict = {}
    # Словарь с продажами артикула за текущий год
    year_sales_dict = {}
    for data in sale_data:
        percent = 0
        if data['lighter__copy_right'] == True:
            percent = copyright_percent
        else:
            percent = main_percent

        if data['lighter__designer']:
            if data['lighter__designer'] in main_sales_dict:
                if data['lighter'] in main_sales_dict[data['lighter__designer']]:
                    if data['month'] in main_sales_dict[data['lighter__designer']][data['lighter']]:
                        main_sales_dict[data['lighter__designer']][data['lighter']
                                                                   ][data['month']]['quantity'] += int(data['quantity'])
                        main_sales_dict[data['lighter__designer']][data['lighter']
                                                                   ][data['month']]['summ'] += int(data['summ'])*percent

                    else:
                        main_sales_dict[data['lighter__designer']][data['lighter']
                                                                   ][data['month']] = {'quantity': int(data['quantity']),
                                                                                       'summ': int(data['summ'])*percent}
                else:
                    main_sales_dict[data['lighter__designer']][data['lighter']] = {
                        data['month']: {
                            'quantity': int(data['quantity']),
                            'summ': int(data['summ'])*percent}}
            else:
                main_sales_dict[data['lighter__designer']] = {data['lighter']: {data['month']: {
                    'quantity': int(data['quantity']), 'summ': int(data['summ'])*percent}}}

            if data['lighter__designer'] in year_sales_dict:
                if data['lighter'] in year_sales_dict[data['lighter__designer']]:
                    year_sales_dict[data['lighter__designer']][data['lighter']
                                                               ]['quantity'] += int(data['quantity'])
                    year_sales_dict[data['lighter__designer']][data['lighter']
                                                               ]['summ'] += int(data['summ'])*percent
                else:
                    year_sales_dict[data['lighter__designer']][data['lighter']] = {'quantity': int(data['quantity']),
                                                                                   'summ': int(data['summ'])*percent}
            else:
                year_sales_dict[data['lighter__designer']] = {data['lighter']: {'quantity': int(data['quantity']),
                                                                                'summ': int(data['summ'])*percent}}
    return year_sales_dict, main_sales_dict


def get_team_amount_summ_sales(year):
    """Отдает данные по продажам и количеству артикулов за все дизайнерские ночники"""
    sale_data = Selling.objects.filter(year=year, lighter__designer_article=True).values(
        'lighter', 'month', 'quantity', 'summ', 'lighter__designer_article', 'lighter__copy_right', 'lighter__designer')

    team_obj = InnotreidUser.objects.filter(
        username='team')[0]
    team_designer_obj = DesignUser.objects.get(designer=team_obj)
    team_main_percent = team_designer_obj.main_reward_persent/100
    team_copyright_percent = team_designer_obj.copyright_reward_persent/100

    # Словарь с данными артикула по продажам по месяцам
    main_sales_dict = {}
    # Словарь с продажами артикула за текущий год
    year_sales_dict = {}
    for data in sale_data:
        team_percent = 0
        if data['lighter__copy_right'] == True:
            team_percent = team_copyright_percent
        else:
            team_percent = team_main_percent
        # ========== % ДЛЯ ВСЕЙ КОМАНДЫ ЕЖЕМЕСЯЧНО ========== #
        if team_obj.pk in main_sales_dict:
            if data['lighter'] in main_sales_dict[team_obj.pk]:
                if data['month'] in main_sales_dict[team_obj.pk][data['lighter']]:
                    main_sales_dict[team_obj.pk][data['lighter']
                                                 ][data['month']]['quantity'] += int(data['quantity'])
                    main_sales_dict[team_obj.pk][data['lighter']
                                                 ][data['month']]['summ'] += int(data['summ'])*team_percent
                else:
                    main_sales_dict[team_obj.pk][data['lighter']
                                                 ][data['month']] = {'quantity': int(data['quantity']),
                                                                     'summ': int(data['summ'])*team_percent}
            else:
                main_sales_dict[team_obj.pk][data['lighter']] = {
                    data['month']: {
                        'quantity': int(data['quantity']),
                        'summ': int(data['summ'])*team_percent}}
        else:
            main_sales_dict[team_obj.pk] = {data['lighter']: {data['month']: {
                'quantity': int(data['quantity']), 'summ': int(data['summ'])*team_percent}}}
        # ========== КОНЕЦ % ДЛЯ ВСЕЙ КОМАНДЫ ЕЖЕМЕСЯЧНО ========== #

        # ========== % ДЛЯ ВСЕЙ КОМАНДЫ В ГОД ========== #
        if team_obj.pk in year_sales_dict:
            if data['lighter'] in year_sales_dict[team_obj.pk]:
                year_sales_dict[team_obj.pk][data['lighter']
                                             ]['quantity'] += int(data['quantity'])
                year_sales_dict[team_obj.pk][data['lighter']
                                             ]['summ'] += int(data['summ'])*team_percent
            else:
                year_sales_dict[team_obj.pk][data['lighter']] = {'quantity': int(data['quantity']),
                                                                 'summ': int(data['summ'])*team_percent}
        else:
            year_sales_dict[team_obj.pk] = {data['lighter']: {'quantity': int(data['quantity']),
                                                              'summ': int(data['summ'])*team_percent}}
        # ========== КОНЕЦ % ДЛЯ ВСЕЙ КОМАНДЫ В ГОД ========== #
    return year_sales_dict, main_sales_dict


def get_designers_sales_data(year):
    """Отдает данные по продажам дизайнеров"""
    sale_data = Selling.objects.filter(year=year).values(
        'lighter', 'month', 'summ', 'lighter__designer', 'lighter__designer_article', 'lighter__copy_right')
    designer_rew_dict = {}
    designer_persent = DesignUser.objects.all().values(
        'designer', 'main_reward_persent', 'copyright_reward_persent')
    for i in designer_persent:
        if i['copyright_reward_persent']:
            designer_rew_dict[i['designer']
                              ] = i['copyright_reward_persent']/100
        elif not i['copyright_reward_persent'] and i['main_reward_persent']:
            designer_rew_dict[i['designer']] = i['main_reward_persent']/100
        else:
            designer_rew_dict[i['designer']] = 0
    # Словарь с данными артикула по продажам по месяцам
    monthly_sales_dict = {}
    # Словарь с продажами артикула за текущий год
    year_sales_dict = {}
    team_obj = InnotreidUser.objects.filter(
        username='team')[0]
    team_designer_obj = DesignUser.objects.get(designer=team_obj)
    team_main_percent = team_designer_obj.main_reward_persent/100
    team_copyright_percent = team_designer_obj.copyright_reward_persent/100
    for data in sale_data:
        if data['lighter__designer']:

            # ========== % ДЛЯ ВСЕЙ КОМАНДЫ ========== #
            team_percent = 0
            if data['lighter__copy_right'] == True:
                team_percent = team_copyright_percent
            else:
                team_percent = team_main_percent
            # ========== КОНЕЦ % ДЛЯ ВСЕЙ КОМАНДЫ ========== #
            designer_obj = DesignUser.objects.filter(
                designer__id=data['lighter__designer'])[0]
            main_percent = designer_obj.main_reward_persent/100
            if designer_obj.copyright_reward_persent:
                copyright_percent = designer_obj.copyright_reward_persent/100
            percent = 0
            if data['lighter__copy_right'] == True:
                percent = copyright_percent
            else:
                percent = main_percent

            if data['lighter__designer'] in monthly_sales_dict:
                if data['month'] in monthly_sales_dict[data['lighter__designer']]:
                    monthly_sales_dict[data['lighter__designer']
                                       ][data['month']] += int(data['summ'])*percent
                else:
                    monthly_sales_dict[data['lighter__designer']
                                       ][data['month']] = int(data['summ'])*percent
            else:
                monthly_sales_dict[data['lighter__designer']] = {
                    data['month']: int(data['summ'])*percent}
            # ========== % ДЛЯ ВСЕЙ КОМАНДЫ В МЕСЯЦ ========== #
            if team_obj.pk in monthly_sales_dict:
                if data['month'] in monthly_sales_dict[team_obj.pk]:
                    monthly_sales_dict[team_obj.pk
                                       ][data['month']] += int(data['summ'])*team_percent
                else:
                    monthly_sales_dict[team_obj.pk
                                       ][data['month']] = int(data['summ'])*team_percent
            else:
                monthly_sales_dict[team_obj.pk] = {
                    data['month']: int(data['summ'])*team_percent}
            # ========== КОНЕЦ % ДЛЯ ВСЕЙ КОМАНДЫ В МЕСЯЦ ========== #

            if data['lighter__designer'] in year_sales_dict:
                year_sales_dict[data['lighter__designer']
                                ] += int(data['summ'])*percent
            else:
                year_sales_dict[data['lighter__designer']] = int(
                    data['summ'])*percent

            # ========== % ДЛЯ ВСЕЙ КОМАНДЫ В ГОД ========== #
            if team_obj.pk in year_sales_dict:
                year_sales_dict[team_obj.pk
                                ] += int(data['summ'])*team_percent
            else:
                year_sales_dict[team_obj.pk] = int(
                    data['summ'])*team_percent
            # ========== КОНЕЦ % ДЛЯ ВСЕЙ КОМАНДЫ В ГОД ========== #
    return year_sales_dict, monthly_sales_dict


def get_team_sales_data(year):
    """Отдает данные по продажам дизайнеров"""
    sale_data = Selling.objects.filter(year=year, lighter__designer_article=True).values(
        'lighter', 'month', 'summ', 'lighter__designer', 'lighter__designer_article', 'lighter__copy_right')

    # Словарь с данными артикула по продажам по месяцам
    monthly_sales_dict = {}
    # Словарь с продажами артикула за текущий год
    year_sales_dict = {}
    team_obj = InnotreidUser.objects.filter(
        username='team')[0]
    team_designer_obj = DesignUser.objects.get(designer=team_obj)
    team_main_percent = team_designer_obj.main_reward_persent/100
    team_copyright_percent = team_designer_obj.copyright_reward_persent/100
    for data in sale_data:
        # ========== % ДЛЯ ВСЕЙ КОМАНДЫ ========== #
        team_percent = 0
        if data['lighter__copy_right'] == True:
            team_percent = team_copyright_percent
        else:
            team_percent = team_main_percent
        # ========== КОНЕЦ % ДЛЯ ВСЕЙ КОМАНДЫ ========== #

        # ========== % ДЛЯ ВСЕЙ КОМАНДЫ В МЕСЯЦ ========== #
        if team_obj.pk in monthly_sales_dict:
            if data['month'] in monthly_sales_dict[team_obj.pk]:
                monthly_sales_dict[team_obj.pk
                                   ][data['month']] += int(data['summ'])*team_percent
            else:
                monthly_sales_dict[team_obj.pk
                                   ][data['month']] = int(data['summ'])*team_percent
        else:
            monthly_sales_dict[team_obj.pk] = {
                data['month']: int(data['summ'])*team_percent}
        # ========== КОНЕЦ % ДЛЯ ВСЕЙ КОМАНДЫ В МЕСЯЦ ========== #
        # ========== % ДЛЯ ВСЕЙ КОМАНДЫ В ГОД ========== #
        if team_obj.pk in year_sales_dict:
            year_sales_dict[team_obj.pk
                            ] += int(data['summ'])*team_percent
        else:
            year_sales_dict[team_obj.pk] = int(
                data['summ'])*team_percent
        # ========== КОНЕЦ % ДЛЯ ВСЕЙ КОМАНДЫ В ГОД ========== #
    return year_sales_dict, monthly_sales_dict


def get_article_sales_data(year):
    """Отдает данные по продажам артикулов"""
    sale_data = Selling.objects.filter(year=year).values(
        'lighter', 'month', 'summ', 'lighter__designer')
    # Словарь с продажами артикула за текущий год
    year_sales_dict = {}
    for data in sale_data:
        if data['lighter__designer']:

            if data['lighter__designer'] in year_sales_dict:
                year_sales_dict[data['lighter__designer']
                                ] += int(data['summ'])
            else:
                year_sales_dict[data['lighter__designer']] = int(
                    data['summ'])
            # ========== ОБЪЕКТ ДЛЯ ВСЕЙ КОМАНДЫ И ЕГО % ========== #
            team_obj = InnotreidUser.objects.filter(
                username='team')[0]
            if team_obj.pk in year_sales_dict:
                year_sales_dict[team_obj.pk
                                ] += int(data['summ'])
            else:
                year_sales_dict[team_obj.pk] = int(
                    data['summ'])
    return year_sales_dict


def get_team_article_sales_data(year):
    """Отдает данные по продажам артикулов для всей команды дизайнеров"""
    sale_data = Selling.objects.filter(year=year, lighter__designer_article=True).values(
        'lighter', 'month', 'summ', 'lighter__designer')
    # Словарь с продажами артикула за текущий год
    year_sales_dict = {}
    for data in sale_data:
        team_obj = InnotreidUser.objects.filter(
            username='team')[0]
        if team_obj.pk in year_sales_dict:
            year_sales_dict[team_obj.pk
                            ] += int(data['summ'])
        else:
            year_sales_dict[team_obj.pk] = int(
                data['summ'])
    return year_sales_dict


def get_current_selling():
    """Получаем текущие продажи артикулов"""
    # ozon_sale_data = OzonSales.objects.all()
    article_data = Articles.objects.all()
    date = datetime.now()
    january = date - timedelta(days=20)
    january_month = int(january.strftime('%m'))
    current_year = january.strftime('%Y')
    ozon_marketplace = CodingMarketplaces.objects.get(marketpalce='Ozon')
    for article in article_data:
        article_data = OzonSales.objects.filter(
            offer_id=article.ozon_seller_article, month=january_month, year=current_year).values('total', 'quantity')
        summ_money = 0
        quantity = 0
        for i in article_data:
            quantity += i['quantity']
            summ_money += i['total']
        Selling(
            lighter=article,
            ur_lico=article.company,
            year=current_year,
            month=january_month,
            summ=summ_money,
            quantity=quantity,
            data=date,
            marketplace=ozon_marketplace).save()


def motivation_article_type_excel_file_export(data):
    """Создает и скачивает excel файл"""
    # Создаем DataFrame из данных
    wb = Workbook()
    # Получаем активный лист
    ws = wb.active

    # Заполняем лист данными
    for row, item in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=str(item['common_article']))
        ws.cell(row=row, column=2, value=str(item['company']))
        if item['designer_article'] == False:
            ws.cell(row=row, column=3, value='')
        else:
            ws.cell(row=row, column=3, value=str(item['designer_article']))
        if item['copy_right'] == False:
            ws.cell(row=row, column=4, value='')
        else:
            ws.cell(row=row, column=4, value=str(item['copy_right']))
    # Устанавливаем заголовки столбцов
    ws.cell(row=1, column=1, value='Артикул')
    ws.cell(row=1, column=2, value='Компания')
    ws.cell(row=1, column=3, value='Дизайнерский ночник')
    ws.cell(row=1, column=4, value='С правами')

    al = Alignment(horizontal="center",
                   vertical="center")
    al_left = Alignment(horizontal="left",
                        vertical="center")
    thin = Side(border_style="thin", color="000000")

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10

    for i in range(len(data)+1):
        for c in ws[f'A{i+1}:D{i+1}']:
            for i in range(4):
                c[i].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[i].alignment = al_left

    # Сохраняем книгу Excel в память
    response = HttpResponse(content_type='application/xlsx')
    name = f'Article_type_{datetime.now().strftime("%Y.%m.%d")}.xlsx'
    file_data = 'attachment; filename=' + name
    response['Content-Disposition'] = file_data
    wb.save(response)

    return response


def motivation_article_type_excel_import(xlsx_file, ur_lico):
    """Импортирует данные о группе артикула из Excel"""
    excel_data_common = pd.read_excel(xlsx_file)
    column_list = excel_data_common.columns.tolist()
    if 'Артикул' in column_list and 'Дизайнерский ночник' in column_list and 'С правами' in column_list:
        excel_data = pd.DataFrame(excel_data_common, columns=[
                                  'Артикул', 'Дизайнерский ночник', 'С правами'])
        article_list = excel_data['Артикул'].to_list()
        designer_type_list = excel_data['Дизайнерский ночник'].to_list()
        copyright_type_list = excel_data['С правами'].to_list()

        # Словарь {артикул: название_группы}
        article_value_dict = {}
        # Список для обновления строк в БД
        new_objects = []

        for i in range(len(article_list)):
            article_value_dict[article_list[i]] = [
                designer_type_list[i], copyright_type_list[i]]

        for row in range(len(article_list)):
            if Articles.objects.filter(
                    company=ur_lico, common_article=article_list[row]).exists():
                article_obj = Articles.objects.get(
                    company=ur_lico, common_article=article_list[row])
                if str(article_value_dict[article_list[row]][0]).capitalize() == 'True':
                    article_obj.designer_article = True
                    if str(article_value_dict[article_list[row]][1]) == 'True':
                        article_obj.copy_right = True
                    else:
                        article_obj.copy_right = False
                else:
                    article_obj.designer_article = False
                    article_obj.copy_right = False

                new_objects.append(article_obj)
        Articles.objects.bulk_update(
            new_objects, ['designer_article', 'copy_right'])
    else:
        return f'Вы пытались загрузить ошибочный файл {xlsx_file}.'


def import_sales_2023(xlsx_file, ur_lico):
    """Импортирует продажи за 2023"""
    excel_data_common = pd.read_excel(xlsx_file)
    column_list = excel_data_common.columns.tolist()
    excel_data = pd.DataFrame(excel_data_common, columns=[
                              'День', 'Артикул продавца', 'Заказано, шт.', 'Сумма заказов минус комиссия WB, руб.'])
    article_list = excel_data['Артикул продавца'].to_list()
    date_list = excel_data['День'].to_list()
    orders_list = excel_data['Заказано, шт.'].to_list()
    summ_list = excel_data['Сумма заказов минус комиссия WB, руб.'].to_list()

    # Словарь {артикул: название_группы}
    article_value_dict = {}
    # Список для обновления строк в БД
    new_objects = []
    marketplace_obj = CodingMarketplaces.objects.get(id=1)
    for row in range(len(article_list)):
        if Articles.objects.filter(
                company='ООО Иннотрейд', common_article=article_list[row].capitalize()).exists():
            article_obj = Articles.objects.get(
                company=ur_lico, common_article=article_list[row].capitalize())
            month = date_list[row].month
            year = date_list[row].year
            data = date_list[row]
            marketplace = marketplace_obj
            quantity = orders_list[row]
            summ = summ_list[row]
            ur_lico = 'ООО Иннотрейд'
            if Selling.objects.filter(ur_lico='ООО Иннотрейд',
                                      marketplace=marketplace_obj,
                                      lighter=article_obj,
                                      year=date_list[row].year,
                                      month=date_list[row].month
                                      ).exists():
                sale_obj = Selling.objects.get(ur_lico='ООО Иннотрейд',
                                               marketplace=marketplace_obj,
                                               lighter=article_obj,
                                               year=date_list[row].year,
                                               month=date_list[row].month
                                               )
                sale_obj.quantity += quantity
                sale_obj.summ += summ
                sale_obj.save()
                print('Прибавил к существующией')
            else:
                Selling(ur_lico='ООО Иннотрейд',
                        marketplace=marketplace_obj,
                        lighter=article_obj,
                        year=date_list[row].year,
                        month=date_list[row].month,
                        quantity=quantity,
                        summ=summ
                        ).save()
                print('сохранил новую продажу')
    #         new_objects.append(article_obj)
    # Articles.objects.bulk_update(
    #     new_objects, ['designer_article', 'copy_right'])


def get_article_draw_authors_sales_data(year):
    """
    Возвращает информацию по продажам авторских,
    отрисованных светильников и общие продажи каждого дизайнера
    """
    sale_data = Selling.objects.filter(year=year,
                                       lighter__designer_article=True,
                                       lighter__designer__isnull=False).values(
        'lighter', 'month', 'summ',
        'lighter__designer', 'lighter__designer_article',
        'lighter__copy_right')

    # Словарь с продажами артикула за текущий год
    year_draw_author_data = {}
    for data in sale_data:
        # ========== ОБЪЕКТ ДЛЯ ВСЕЙ КОМАНДЫ И ЕГО % ========== #
        team_obj = InnotreidUser.objects.filter(
            username='team')[0]
        team_designer_obj = DesignUser.objects.get(designer=team_obj)
        team_main_percent = team_designer_obj.main_reward_persent/100
        # ========== ПРОДАЖИ ДЛЯ ВСЕЙ КОМАНДЫ В ГОД ========== #
        if team_obj.pk in year_draw_author_data:
            if data['lighter__copy_right'] == True:
                if 'author' in year_draw_author_data[team_obj.pk]:
                    year_draw_author_data[team_obj.pk
                                          ]['author'] += int(data['summ'])
                else:
                    year_draw_author_data[team_obj.pk
                                          ]['author'] = int(data['summ'])
            else:
                if 'draw' in year_draw_author_data[team_obj.pk]:
                    year_draw_author_data[team_obj.pk
                                          ]['draw'] += int(data['summ'])
                else:
                    year_draw_author_data[team_obj.pk]['draw'] = int(
                        data['summ'])
            year_draw_author_data[team_obj.pk]['summ'] += int(
                data['summ'])
        else:

            if data['lighter__copy_right'] == True:
                year_draw_author_data[team_obj.pk] = {'author': int(
                    data['summ'])}
            else:
                year_draw_author_data[team_obj.pk] = {'draw': int(
                    data['summ'])}
            year_draw_author_data[team_obj.pk]['summ'] = int(
                data['summ'])
        # ========== КОНЕЦ ПРОДАЖИ ДЛЯ ВСЕЙ КОМАНДЫ В ГОД ========== #

        if data['lighter__designer'] in year_draw_author_data:
            if data['lighter__copy_right'] == True:
                if 'author' in year_draw_author_data[data['lighter__designer']]:
                    year_draw_author_data[data['lighter__designer']
                                          ]['author'] += int(data['summ'])
                else:
                    year_draw_author_data[data['lighter__designer']
                                          ]['author'] = int(data['summ'])
            else:
                if 'draw' in year_draw_author_data[data['lighter__designer']]:
                    year_draw_author_data[data['lighter__designer']
                                          ]['draw'] += int(data['summ'])
                else:
                    year_draw_author_data[data['lighter__designer']]['draw'] = int(
                        data['summ'])
            year_draw_author_data[data['lighter__designer']]['summ'] += int(
                data['summ'])
        else:

            if data['lighter__copy_right'] == True:
                year_draw_author_data[data['lighter__designer']] = {'author': int(
                    data['summ'])}
            else:
                year_draw_author_data[data['lighter__designer']] = {'draw': int(
                    data['summ'])}
            year_draw_author_data[data['lighter__designer']]['summ'] = int(
                data['summ'])

    return year_draw_author_data


def get_draw_authors_year_monthly_reward(year):
    """
    Возвращает информацию по вознаграждению для авторских и
    отрисованных светильников за год и за каждый месяц
    """
    sale_data = Selling.objects.filter(year=year,
                                       lighter__designer_article=True,
                                       lighter__designer__isnull=False).values(
        'lighter', 'month', 'summ',
        'lighter__designer', 'lighter__designer_article',
        'lighter__copy_right')
    designer_raw_dict = {}
    designer_persent = DesignUser.objects.all().values(
        'designer', 'main_reward_persent', 'copyright_reward_persent')

    for i in designer_persent:
        if i['copyright_reward_persent']:
            designer_raw_dict[i['designer']
                              ] = i['copyright_reward_persent']/100
        elif not i['copyright_reward_persent'] and i['main_reward_persent']:
            designer_raw_dict[i['designer']] = i['main_reward_persent']/100
        else:
            designer_raw_dict[i['designer']] = 0
    # print(sale_data)
    # Словарь с данными артикула по продажам по месяцам
    monthly_sales_dict = {}
    # Словарь с продажами артикула за текущий год
    year_sales_dict = {}
    for data in sale_data:
        designer_obj = DesignUser.objects.filter(
            designer__id=data['lighter__designer'])[0]
        main_percent = designer_obj.main_reward_persent/100
        if designer_obj.copyright_reward_persent:
            copyright_percent = designer_obj.copyright_reward_persent/100
        percent = 0
        if data['lighter__copy_right'] == True:
            percent = copyright_percent
        else:
            percent = main_percent
        if data['lighter__designer'] in monthly_sales_dict:
            if data['month'] in monthly_sales_dict[data['lighter__designer']]:
                monthly_sales_dict[data['lighter__designer']
                                   ][data['month']] += int(data['summ'])*percent
            else:
                monthly_sales_dict[data['lighter__designer']
                                   ][data['month']] = int(data['summ'])*percent
        else:
            monthly_sales_dict[data['lighter__designer']] = {
                data['month']: int(data['summ'])*percent}

        # ========== ОБЪЕКТ ДЛЯ ВСЕЙ КОМАНДЫ И ЕГО % ========== #
        team_obj = InnotreidUser.objects.filter(
            username='team')[0]
        team_designer_obj = DesignUser.objects.get(designer=team_obj)
        team_main_percent = team_designer_obj.main_reward_persent/100

        if team_designer_obj.copyright_reward_persent:
            team_copyright_percent = team_designer_obj.copyright_reward_persent/100
        team_percent = 0
        if data['lighter__copy_right'] == True:
            team_percent = team_copyright_percent
        else:
            team_percent = team_main_percent
        # ========== % ДЛЯ ВСЕЙ КОМАНДЫ В МЕСЯЦ ========== #
        if team_obj.pk in monthly_sales_dict:
            if data['month'] in monthly_sales_dict[team_obj.pk]:
                monthly_sales_dict[team_obj.pk
                                   ][data['month']] += int(data['summ'])*team_percent
            else:
                monthly_sales_dict[team_obj.pk
                                   ][data['month']] = int(data['summ'])*team_percent
        else:
            monthly_sales_dict[team_obj.pk] = {
                data['month']: int(data['summ'])*team_percent}
        # ========== КОНЕЦ % ДЛЯ ВСЕЙ КОМАНДЫ В МЕСЯЦ ========== #

        if data['lighter__designer'] in year_sales_dict:
            if data['lighter__copy_right'] == True:
                if 'author' in year_sales_dict[data['lighter__designer']]:
                    year_sales_dict[data['lighter__designer']
                                    ]['author'] += int(data['summ'])*percent
                else:
                    year_sales_dict[data['lighter__designer']
                                    ]['author'] = int(data['summ'])*percent
            else:
                if 'draw' in year_sales_dict[data['lighter__designer']]:
                    year_sales_dict[data['lighter__designer']
                                    ]['draw'] += int(data['summ'])*percent
                else:
                    year_sales_dict[data['lighter__designer']
                                    ]['draw'] = int(data['summ'])*percent
            year_sales_dict[data['lighter__designer']
                            ]['summ'] += int(data['summ'])*percent
        else:
            if data['lighter__copy_right'] == True:
                year_sales_dict[data['lighter__designer']] = {'author': int(
                    data['summ'])*percent}
            else:
                year_sales_dict[data['lighter__designer']] = {'draw': int(
                    data['summ'])*percent}
            year_sales_dict[data['lighter__designer']]['summ'] = int(
                data['summ'])*percent

        # ========== % ДЛЯ ВСЕЙ КОМАНДЫ В ГОД ========== #
        if team_obj.pk in year_sales_dict:
            if data['lighter__copy_right'] == True:
                if 'author' in year_sales_dict[team_obj.pk]:
                    year_sales_dict[team_obj.pk
                                    ]['author'] += int(data['summ'])*team_percent
                else:
                    year_sales_dict[team_obj.pk
                                    ]['author'] = int(data['summ'])*team_percent
            else:
                if 'draw' in year_sales_dict[team_obj.pk]:
                    year_sales_dict[team_obj.pk
                                    ]['draw'] += int(data['summ'])*team_percent
                else:
                    year_sales_dict[team_obj.pk
                                    ]['draw'] = int(data['summ'])*team_percent
            year_sales_dict[team_obj.pk
                            ]['summ'] += int(data['summ'])*team_percent
        else:
            if data['lighter__copy_right'] == True:
                year_sales_dict[team_obj.pk] = {'author': int(
                    data['summ'])*team_percent}
            else:
                year_sales_dict[team_obj.pk] = {'draw': int(
                    data['summ'])*team_percent}
            year_sales_dict[team_obj.pk]['summ'] = int(
                data['summ'])*team_percent
        # ========== КОНЕЦ % ДЛЯ ВСЕЙ КОМАНДЫ В ГОД ========== #

    return year_sales_dict, monthly_sales_dict


def motivation_designer_rewards_excel_file_export(article_list, year_sales_dict, main_sales_dict, designer, year, month_list, response):
    """Создает и скачивает excel файл с вознаграждением дизайнера"""
    # Создаем DataFrame из данных
    wb = Workbook()
    # Получаем активный лист
    ws = wb.active
    # Заполняем лист данными
    month_amount = len(month_list)
    ws.cell(row=1, column=1, value=f'Вознаграждение дизайнера {designer}')
    ws.merge_cells('A2:A4')
    ws.cell(row=2, column=1, value='Ночник')
    ws.merge_cells('B2:B4')
    ws.cell(row=2, column=2, value='Название')
    ws.merge_cells('C2:C4')
    ws.cell(row=2, column=3, value='Авторский')
    ws.merge_cells('D2:D3')
    ws.cell(row=2, column=4, value=f'Продажи за {year}')
    ws.merge_cells('E2:E3')
    ws.cell(row=2, column=5, value=f'Вознаграждение за {year}')
    ws.cell(row=4, column=4, value=f'шт')
    ws.cell(row=4, column=5, value=f'руб')

    start_cell = ws.cell(row=2, column=6)
    end_cell = ws.cell(row=2, column=6+2*month_amount - 1)
    # merge_range = f'F2:{chr(70 + 2*month_amount - 1)}2'

    # print(f'F2:chr(70 + 2*month_amount - 1)2')
    ws.merge_cells(start_cell.coordinate + ':' + end_cell.coordinate)
    # print('merge_range', merge_range)

    ws.cell(row=2, column=6, value=f'{year} год')
    start_cell_numb = 70
    start_name_numb = 6
    for i in range(month_amount):
        start_cell = ws.cell(row=3, column=start_name_numb)
        end_cell = ws.cell(row=3, column=start_name_numb+2 - 1)

        ws.merge_cells(start_cell.coordinate + ':' + end_cell.coordinate)
        ws.cell(row=3, column=start_name_numb, value=f'{i+1}')
        ws.cell(row=4, column=start_name_numb, value=f'шт')
        start_cell_numb += 2
        start_name_numb += 2
        ws.cell(row=4, column=start_name_numb-1, value=f'руб')
    row = 5
    for item in article_list:
        copy_right_file = ' '
        if item['copy_right'] == True:
            copy_right_file = 'Да'
        ws.cell(row=row, column=1, value=item['common_article'])
        ws.cell(row=row, column=2, value=item['name'])
        ws.cell(row=row, column=3, value=copy_right_file)
        if item['id'] in year_sales_dict:
            ws.cell(row=row, column=4,
                    value=year_sales_dict[item['id']]['quantity'])
            ws.cell(row=row, column=5,
                    value=round(year_sales_dict[item['id']]['summ']))
            start_name_numb = 6
            for month in month_list:
                if month in main_sales_dict[item['id']]:
                    ws.cell(row=row, column=start_name_numb,
                            value=main_sales_dict[item['id']][month]['quantity'])
                    start_name_numb += 2
                    ws.cell(row=row, column=start_name_numb-1,
                            value=round(main_sales_dict[item['id']][month]['summ']))
        row += 1

    al = Alignment(horizontal="center",
                   vertical="center")
    al_left = Alignment(horizontal="left",
                        vertical="center")
    thin = Side(border_style="thin", color="000000")

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 23

    for i in range(1, len(article_list)+4):
        for c in ws[f'A{i+1}:AZ{i+1}']:
            for i in range(len(month_list)*2+5):
                c[i].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[i].alignment = al

    # Сохраняем книгу Excel в память

    wb.save(response)

    return response


def motivation_designer_sales_excel_file_export(article_list, year_sales_dict, main_sales_dict, designer, year, month_list, response):
    """Создает и скачивает excel файл с вознаграждением дизайнера"""
    # Создаем DataFrame из данных
    wb = Workbook()
    # Получаем активный лист
    ws = wb.active
    # Заполняем лист данными
    month_amount = len(month_list)
    ws.cell(row=1, column=1, value=f'Вознаграждение дизайнера {designer}')
    ws.merge_cells('A2:A4')
    ws.cell(row=2, column=1, value='Ночник')
    ws.merge_cells('B2:B4')
    ws.cell(row=2, column=2, value='Название')
    ws.merge_cells('C2:C4')
    ws.cell(row=2, column=3, value='Авторский')
    ws.merge_cells('D2:E3')
    ws.cell(row=2, column=4, value=f'Продажи за {year}')
    ws.cell(row=4, column=4, value=f'шт')
    ws.cell(row=4, column=5, value=f'руб')

    start_cell = ws.cell(row=2, column=6)
    end_cell = ws.cell(row=2, column=6+2*month_amount - 1)
    # merge_range = f'F2:{chr(70 + 2*month_amount - 1)}2'

    # print(f'F2:chr(70 + 2*month_amount - 1)2')
    ws.merge_cells(start_cell.coordinate + ':' + end_cell.coordinate)
    # print('merge_range', merge_range)

    ws.cell(row=2, column=6, value=f'{year} год')
    start_cell_numb = 70
    start_name_numb = 6
    for i in range(month_amount):
        start_cell = ws.cell(row=3, column=start_name_numb)
        end_cell = ws.cell(row=3, column=start_name_numb+2 - 1)

        ws.merge_cells(start_cell.coordinate + ':' + end_cell.coordinate)
        ws.cell(row=3, column=start_name_numb, value=f'{i+1}')
        ws.cell(row=4, column=start_name_numb, value=f'шт')
        start_cell_numb += 2
        start_name_numb += 2
        ws.cell(row=4, column=start_name_numb-1, value=f'руб')
    row = 5
    for item in article_list:
        copy_right_file = ' '
        if item['copy_right'] == True:
            copy_right_file = 'Да'
        ws.cell(row=row, column=1, value=item['common_article'])
        ws.cell(row=row, column=2, value=item['name'])
        ws.cell(row=row, column=3, value=copy_right_file)
        if item['id'] in year_sales_dict:
            ws.cell(row=row, column=4,
                    value=year_sales_dict[item['id']]['quantity'])
            ws.cell(row=row, column=5,
                    value=round(year_sales_dict[item['id']]['summ']))
            start_name_numb = 6
            for month in month_list:
                if month in main_sales_dict[item['id']]:
                    ws.cell(row=row, column=start_name_numb,
                            value=main_sales_dict[item['id']][month]['quantity'])
                    start_name_numb += 2
                    ws.cell(row=row, column=start_name_numb-1,
                            value=round(main_sales_dict[item['id']][month]['summ']))
        row += 1

    al = Alignment(horizontal="center",
                   vertical="center")
    al_left = Alignment(horizontal="left",
                        vertical="center")
    thin = Side(border_style="thin", color="000000")

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35

    for i in range(1, len(article_list)+4):
        for c in ws[f'A{i+1}:AZ{i+1}']:
            for i in range(len(month_list)*2+5):
                c[i].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[i].alignment = al

    # Сохраняем книгу Excel в память

    wb.save(response)

    return response
