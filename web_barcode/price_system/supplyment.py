import datetime
import json
import math
import os
import time
import traceback
from datetime import datetime

import pandas as pd
import requests
import telegram
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

from web_barcode.constants_file import (CHAT_ID_ADMIN, TELEGRAM_TOKEN,
                                        header_ozon_dict, header_wb_dict,
                                        header_yandex_dict,
                                        wb_headers_karavaev, wb_headers_ooo,
                                        yandex_business_id_dict)

from .models import ArticleGroup, Articles, Groups

bot = telegram.Bot(token=TELEGRAM_TOKEN)

COLUMN_EXCEL_CREATING_GROUP_LIST = ['Название', 'Юр. лицо', 'Старая цена',
                                    'WB скидка продавца', 'Цена WB', 'Цена OZON', 'Цена YANDEX', 'Мин. цена']


def sender_error_to_tg(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            tb_str = traceback.format_exc()
            message_error = (f'Ошибка в функции: {func.__name__}\n'
                             f'Функция выполняет: {func.__doc__}\n'
                             f'Ошибка\n: {e}\n\n'
                             f'Техническая информация:\n {tb_str}')
            try:
                bot.send_message(chat_id=CHAT_ID_ADMIN,
                             # text=message_error[:4000], parse_mode='HTML')
                             text=message_error[:4000])
            except:
                print('Не сработала отправка')
    return wrapper


@sender_error_to_tg
def wb_ip_article_compare():
    """Достаем данные всех артиуклов ВБ, необходимые для сверки"""
    from api_request.wb_requests import wb_article_data_from_api
    all_data = wb_article_data_from_api(wb_headers_karavaev)
    article_dict = {}
    for data in all_data:
        if data["subjectName"] == "Ночники":
            # article = data["vendorCode"].split('-')[0]
            article = data["vendorCode"]
            article_dict[article.capitalize()] = [data["vendorCode"],
                                                  data["sizes"][0]["skus"][0], data["nmID"], data["title"], data['photos'][0]['c246x328']]
        else:
            article_dict[data["vendorCode"]] = [data["vendorCode"],
                                                data["sizes"][0]["skus"][0], data["nmID"], data["title"], data['photos'][0]['c246x328']]

    sorted_article_dict = dict(sorted(article_dict.items()))
    return sorted_article_dict


@sender_error_to_tg
def wb_ooo_article_compare(ur_lico):
    """Достаем данные всех артиуклов ВБ, необходимые для сверки"""
    from api_request.wb_requests import wb_article_data_from_api
    all_data = wb_article_data_from_api(header_wb_dict[ur_lico])
    article_dict = {}
    for data in all_data:
        article = data["vendorCode"]

        article_dict[article.capitalize()] = [data["vendorCode"],
                                              data["sizes"][0]["skus"][0], data["nmID"], data["title"], data['photos'][0]['c246x328']]
    sorted_article_dict = dict(sorted(article_dict.items()))
    return sorted_article_dict


@sender_error_to_tg
def ozon_raw_articles(ur_lico, last_id='', main_stock_data=None):
    """Получает список product_id товаров с OZON"""
    header = header_ozon_dict[ur_lico]
    if main_stock_data is None:
        main_stock_data = []
    url = "https://api-seller.ozon.ru/v2/product/list"

    payload = json.dumps({
        "last_id": last_id,
        "limit": 1000
    })
    response = requests.request(
        "POST", url, headers=header, data=payload)
    ozon_data = json.loads(response.text)["result"]["items"]
    for dat in ozon_data:
        main_stock_data.append(dat["product_id"])
    if len(ozon_data) == 1000:
        return ozon_raw_articles(ur_lico, json.loads(response.text)[
            "result"]['last_id'], main_stock_data)
    # print(raw_article_ozon_list)
    return main_stock_data


@sender_error_to_tg
def ozon_cleaning_articles(ur_lico):
    urls = 'https://api-seller.ozon.ru/v2/product/info/list'
    header = header_ozon_dict[ur_lico]
    raw_article_ozon_list = ozon_raw_articles(ur_lico)
    koef_product = math.ceil(len(raw_article_ozon_list)/900)

    article_ozon_dict = {}
    for i in range(koef_product):
        start_point = i*900
        finish_point = (i+1)*900
        big_info_list = raw_article_ozon_list[
            start_point:finish_point]
        payload = json.dumps({
            "offer_id": [],
            "product_id": big_info_list,
            "sku": []
        })
        response = requests.request(
            "POST", urls, headers=header, data=payload)
        ozon_data = json.loads(response.text)["result"]["items"]
        if ur_lico == 'ИП Караваев':
            for dat in ozon_data:
                if 'Ночник' in dat["name"]:
                    # article = dat["offer_id"].split('-')[0]
                    article = dat["offer_id"]
                    article_ozon_dict[article] = [
                        dat["offer_id"], dat["barcode"], dat["id"], dat["sku"], dat["fbo_sku"], dat["fbs_sku"]]
                else:
                    article_ozon_dict[dat["offer_id"]] = [
                        dat["offer_id"], dat["barcode"], dat["id"], dat["sku"], dat["fbo_sku"], dat["fbs_sku"]]
        else:
            for dat in ozon_data:
                article_ozon_dict[dat["offer_id"].capitalize()] = [
                    dat["offer_id"], dat["barcode"], dat["id"], dat["sku"], dat["fbo_sku"], dat["fbs_sku"]]
    return article_ozon_dict


@sender_error_to_tg
def yandex_raw_articles_data(ur_lico, nextPageToken='', raw_articles_list=None):
    """Создает и возвращает словарь с данными fbs_sku_data {артикул: остаток_fbs}"""
    header = header_yandex_dict[ur_lico]
    if raw_articles_list == None:
        raw_articles_list = []
    url = f'https://api.partner.market.yandex.ru/businesses/{yandex_business_id_dict[ur_lico]}/offer-mappings?limit=200&page_token={nextPageToken}'
    payload = json.dumps({})

    response = requests.request(
        "POST", url, headers=header, data=payload)
    main_articles_data = json.loads(response.text)['result']
    articles_data = main_articles_data['offerMappings']
    for article in articles_data:
        if article['offer']['barcodes'] and 'marketSku' in article['mapping']:
            inner_list = []
            # print(article['offer']['offerId'])
            inner_list.append(article['offer']['offerId'])
            inner_list.append(article['offer']['barcodes'][0])
            inner_list.append(article['mapping']['marketSku'])
            if 'Ночник' in article['offer']['name']:
                inner_list.append('Ночник')
            else:
                inner_list.append('Грамота')
            raw_articles_list.append(inner_list)

    if main_articles_data['paging']:
        return yandex_raw_articles_data(ur_lico,
                                        main_articles_data['paging']['nextPageToken'], raw_articles_list)
    return raw_articles_list


@sender_error_to_tg
def yandex_articles(ur_lico):
    """Формирует данные для сопоставления"""
    raw_articles_list = yandex_raw_articles_data(ur_lico)
    article_yandex_dict = {}
    for article in raw_articles_list:
        if ur_lico == 'ИП Караваев' and article[-1] == 'Ночник':
            # common_article = article[0].split('-')[0]
            common_article = article[0]
            article_yandex_dict[common_article] = article[:3]
        else:
            article_yandex_dict[article[0].capitalize()] = article[:3]
    return article_yandex_dict


@sender_error_to_tg
def wb_matching_articles(ur_lico):
    """
    WILDBERRIES.
    Функция сопоставляет артикулы с WB с общей базой
    """
    if ur_lico == 'ИП Караваев':
        wb_article_data = wb_ip_article_compare()
    else:
        wb_article_data = wb_ooo_article_compare(ur_lico)
    for common_article, wb_data in wb_article_data.items():
        if Articles.objects.filter(company=ur_lico, wb_nomenclature=wb_data[2]).exists():
            wb_article = Articles.objects.get(company=ur_lico,
                                              wb_nomenclature=wb_data[2])
            if wb_article.wb_seller_article != wb_data[0] or str(wb_article.wb_barcode) != str(wb_data[1]):
                wb_article.status = 'Не сопоставлено'
                wb_article.company = ur_lico
                wb_article.wb_seller_article = wb_data[0]
                wb_article.wb_barcode = wb_data[1]
                wb_article.common_article = common_article
                wb_article.name = wb_data[3]
                wb_article.wb_photo_address = wb_data[4]
                wb_article.save()
                message = (f'{ur_lico} проверьте артикул {common_article} на вб вручную. \
                           Не совпали данные. Артикулы: {wb_article.wb_seller_article} {wb_data[0]}. \
                           Баркоды: {wb_article.wb_barcode} {wb_data[1]}. \
                           common_article: {wb_article.common_article} {common_article}')
                bot.send_message(chat_id=CHAT_ID_ADMIN, text=message)
            else:
                wb_article.status = 'Сопоставлено'
                wb_article.company = ur_lico
                wb_article.name = wb_data[3]
                wb_article.wb_photo_address = wb_data[4]
                wb_article.save()
        else:
            wb = Articles(
                common_article=common_article,
                status='Сопоставлено',
                company=ur_lico,
                name=wb_data[3],
                wb_seller_article=wb_data[0],
                wb_barcode=wb_data[1],
                wb_nomenclature=wb_data[2],
                wb_photo_address=wb_data[4],
                designer_article=None,
                copy_right=None
            )
            wb.save()


@sender_error_to_tg
def ozon_matching_articles(ur_lico):
    """Функция сопоставляет артикулы с Ozon с общей базой"""
    ozon_article_data = ozon_cleaning_articles(ur_lico)
    main_data = Articles.objects.filter(company=ur_lico)
    for common_article, ozon_data in ozon_article_data.items():
        ozon_article = None
        if main_data.filter(wb_barcode=ozon_data[1]).exists():

            ozon_article = main_data.filter(wb_barcode=ozon_data[1])[0]
            if len(main_data.filter(wb_barcode=ozon_data[1])) > 1:
                print(main_data.filter(wb_barcode=ozon_data[1])[
                      0].wb_seller_article, main_data.filter(wb_barcode=ozon_data[1])[
                      0].wb_barcode, len(main_data.filter(wb_barcode=ozon_data[1])))
        elif main_data.filter(common_article=common_article).exists():
            ozon_article = Articles.objects.get(
                company=ur_lico, common_article=common_article)
        if ozon_article != None:
            if (ozon_article.ozon_seller_article != ozon_data[0] and ozon_article.ozon_seller_article != None
                or str(ozon_article.ozon_barcode) != str(ozon_data[1]) and ozon_article.ozon_barcode != None
                or ozon_article.ozon_product_id != ozon_data[2] and ozon_article.ozon_product_id != None
                or ozon_article.ozon_sku != ozon_data[3] and ozon_article.ozon_sku != None
                or ozon_article.ozon_fbo_sku_id != ozon_data[4] and ozon_article.ozon_fbo_sku_id != None
                    or ozon_article.ozon_fbs_sku_id != ozon_data[5] and ozon_article.ozon_fbs_sku_id != None):
                ozon_article.status = 'Не сопоставлено'
                ozon_article.company = ur_lico
                ozon_article.save()
                # print('ozon_article.wb_barcode', ozon_article.wb_barcode, ozon_article.wb_seller_article,
                #      ozon_article.ozon_seller_article, ozon_article.ozon_barcode)
                time.sleep(1)
                message = (f'{ur_lico} проверьте артикул {common_article} на ozon вручную.  \
                           Артикул поставщика {ozon_article.ozon_seller_article} - {ozon_data[0]}. \
                           Barcode {ozon_article.ozon_barcode} - {ozon_data[1]}. \
                           Product_id {ozon_article.ozon_product_id} - {ozon_data[2]}. \
                           SKU {ozon_article.ozon_sku} - {ozon_data[3]}. \
                           SKU_FBO {ozon_article.ozon_fbo_sku_id} - {ozon_data[4]}. \
                           SKU_FBS {ozon_article.ozon_fbs_sku_id} - {ozon_data[5]}. \
                           Не совпали данные')
                ozon_article.status = 'Сопоставлено'
                ozon_article.ozon_seller_article = ozon_data[0]
                ozon_article.ozon_barcode = ozon_data[1]
                ozon_article.ozon_product_id = ozon_data[2]
                ozon_article.ozon_sku = ozon_data[3]
                ozon_article.ozon_fbo_sku_id = ozon_data[4]
                ozon_article.ozon_fbs_sku_id = ozon_data[5]
                ozon_article.save()
                print(message)
                bot.send_message(chat_id=CHAT_ID_ADMIN, text=message)
            elif ozon_article.ozon_barcode == None:
                ozon = ozon_article
                ozon.status = 'Сопоставлено'
                ozon.company = ur_lico
                ozon.ozon_seller_article = ozon_data[0]
                ozon.ozon_barcode = ozon_data[1]
                ozon.ozon_product_id = int(ozon_data[2])
                ozon.ozon_sku = int(ozon_data[3])
                ozon.ozon_fbo_sku_id = int(ozon_data[4])
                ozon.ozon_fbs_sku_id = int(ozon_data[5])
                ozon.save()
            ozon_article = None
        else:
            ozon = Articles(
                common_article=common_article,
                status='Сопоставлено',
                company=ur_lico,
                ozon_seller_article=ozon_data[0],
                ozon_barcode=ozon_data[1],
                ozon_product_id=int(ozon_data[2]),
                ozon_sku=int(ozon_data[3]),
                ozon_fbo_sku_id=int(ozon_data[4]),
                ozon_fbs_sku_id=int(ozon_data[5]),
                designer_article=None,
                copy_right=None
            )
            ozon.save()


@sender_error_to_tg
def yandex_matching_articles(ur_lico):
    """Функция сопоставляет артикулы с Яндекса с общей базой"""
    yandex_article_data = yandex_articles(ur_lico)
    main_data = Articles.objects.filter(company=ur_lico)
    for common_article, yandex_data in yandex_article_data.items():
        yandex_article = None
        if main_data.filter(wb_barcode=yandex_data[1]).exists():
            yandex_article = Articles.objects.filter(company=ur_lico,
                                                     wb_barcode=yandex_data[1])[0]

        elif main_data.filter(common_article=common_article).exists():
            yandex_article = Articles.objects.get(company=ur_lico,
                                                  common_article=common_article)
        elif main_data.filter(ozon_barcode=yandex_data[1]).exists():
            yandex_article = Articles.objects.get(company=ur_lico,
                                                  ozon_barcode=yandex_data[1])
        if yandex_article != None:
            if (yandex_article.yandex_seller_article != yandex_data[0] and yandex_article.yandex_seller_article != None
                or str(yandex_article.yandex_barcode) != str(yandex_data[1]) and yandex_article.yandex_barcode != None
                    or str(yandex_article.yandex_sku) != str(yandex_data[2]) and yandex_article.yandex_sku != None):
                
                print(yandex_article.yandex_seller_article, yandex_data[0])
                print(yandex_article.yandex_barcode, yandex_data[1], type(yandex_article.yandex_barcode), type(yandex_data[1]))
                print(yandex_article.yandex_sku, yandex_data[2], type(yandex_article.yandex_sku), type(yandex_data[2]))
                print('**********************')

                yandex_article.status = 'Cопоставлено'

                yandex_article.yandex_seller_article = yandex_data[0]
                yandex_article.yandex_barcode = yandex_data[1]
                yandex_article.yandex_sku = yandex_data[2]
                yandex_article.company = ur_lico
                yandex_article.save()
                message = (f'Проверьте артикул {common_article} на яндекс вручную. \
                            Barcode {yandex_article.yandex_barcode} - {yandex_data[1]}. \
                            SKU {yandex_article.yandex_sku} - {yandex_data[2]}. \
                            Не совпали данные')
                time.sleep(0.5)
                bot.send_message(chat_id=CHAT_ID_ADMIN, text=message)
            elif yandex_article.yandex_barcode == None:
                yandex_article.status = 'Сопоставлено'
                yandex_article.company = ur_lico
                yandex_article.yandex_seller_article = yandex_data[0]
                yandex_article.yandex_barcode = yandex_data[1]
                yandex_article.yandex_sku = int(yandex_data[2])
                yandex_article.save()
            yandex_article = None
        else:
            yandex = Articles(
                common_article=common_article,
                status='Сопоставлено',
                company=ur_lico,
                yandex_seller_article=yandex_data[0],
                yandex_barcode=yandex_data[1],
                yandex_sku=int(yandex_data[2]),
                designer_article=None,
                copy_right=None
            )
            yandex.save()
        yandex_article = None


def excel_creating_mod(data):
    """Создает и скачивает excel файл"""
    # Создаем DataFrame из данных
    wb = Workbook()
    # Получаем активный лист
    ws = wb.active

    # Заполняем лист данными
    for row, item in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=str(item.common_article))
        if item.group:
            ws.cell(row=row, column=2, value=str(item.group.company))
        ws.cell(row=row, column=3, value=str(item.group))
    # Устанавливаем заголовки столбцов
    ws.cell(row=1, column=1, value='Артикул')
    ws.cell(row=1, column=2, value='Юр. лицо')
    ws.cell(row=1, column=3, value='Ценовая группа артикула')

    al = Alignment(horizontal="center",
                   vertical="center")
    al_left = Alignment(horizontal="left",
                        vertical="center")
    al2 = Alignment(vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")
    pattern = PatternFill('solid', fgColor="fcff52")

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10

    for i in range(len(data)+1):
        for c in ws[f'A{i+1}:C{i+1}']:
            for i in range(3):
                c[i].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[i].alignment = al_left

    # Сохраняем книгу Excel в память
    response = HttpResponse(content_type='application/xlsx')
    name = f'Articles_groups_{datetime.now().strftime("%Y.%m.%d")}.xlsx'
    file_data = 'attachment; filename=' + name
    response['Content-Disposition'] = file_data
    wb.save(response)

    return response


def excel_with_price_groups_creating_mod(data, ur_lico):
    """
    Экспортирует в Excel таблицу с группами.
    Создает и скачивает excel файл с ценовыми группами
    """
    # Создаем DataFrame из данных
    wb = Workbook()
    # Получаем активный лист
    ws = wb.active
    # Заполняем лист данными
    for row, item in enumerate(data, start=2):

        wb_price = round(item.old_price * (100 - item.wb_discount) / 100)
        ws.cell(row=row, column=1, value=str(item.name))
        ws.cell(row=row, column=2, value=str(item.company))
        ws.cell(row=row, column=3, value=str(item.old_price))
        ws.cell(row=row, column=4, value=str(item.wb_discount))
        ws.cell(row=row, column=5, value=str(wb_price))
        ws.cell(row=row, column=6, value=str(item.ozon_price))
        ws.cell(row=row, column=7, value=str(item.yandex_price))
        ws.cell(row=row, column=8, value=str(item.min_price))
        # ws.cell(row=row, column=8, value=str(item.old_price))
    # Устанавливаем заголовки столбцов
    for i in range(len(COLUMN_EXCEL_CREATING_GROUP_LIST)):
        ws.cell(row=1, column=i+1, value=COLUMN_EXCEL_CREATING_GROUP_LIST[i])

    al_left = Alignment(horizontal="left",
                        vertical="center")
    thin = Side(border_style="thin", color="000000")

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    # ws.column_dimensions['H'].width = 12

    for i in range(len(data)+1):
        for c in ws[f'A{i+1}:I{i+1}']:
            for i in range(len(COLUMN_EXCEL_CREATING_GROUP_LIST)):
                c[i].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[i].alignment = al_left

    # Сохраняем книгу Excel в память
    response = HttpResponse(content_type='application/xlsx')
    name = f'Groups_{datetime.now().strftime("%Y.%m.%d")}.xlsx'
    file_data = 'attachment; filename=' + name
    response['Content-Disposition'] = file_data
    wb.save(response)

    return response


def excel_compare_table(data):
    """Экспортирует в excel таблицу сверки"""
    # Создаем DataFrame из данных
    wb = Workbook()
    # Получаем активный лист
    ws = wb.active
    # Заполняем лист данными
    for row, item in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=str(item.common_article))
        ws.cell(row=row, column=2, value=str(item.status))
        ws.cell(row=row, column=3, value=str(item.company))

        ws.cell(row=row, column=4, value=str(item.wb_seller_article))
        ws.cell(row=row, column=5, value=str(item.wb_barcode))
        ws.cell(row=row, column=6, value=str(item.wb_nomenclature))

        ws.cell(row=row, column=7, value=str(item.ozon_seller_article))
        ws.cell(row=row, column=8, value=str(item.ozon_barcode))
        ws.cell(row=row, column=9, value=str(item.ozon_product_id))
        ws.cell(row=row, column=10, value=str(item.ozon_sku))
        ws.cell(row=row, column=11, value=str(item.ozon_fbo_sku_id))
        ws.cell(row=row, column=12, value=str(item.ozon_fbs_sku_id))

        ws.cell(row=row, column=13, value=str(item.yandex_seller_article))
        ws.cell(row=row, column=14, value=str(item.yandex_barcode))
        ws.cell(row=row, column=15, value=str(item.yandex_sku))

    # Устанавливаем заголовки столбцов
    ws.cell(row=1, column=1, value='Общий артикул')
    ws.cell(row=1, column=2, value='Статус')
    ws.cell(row=1, column=3, value='Компания')

    ws.cell(row=1, column=4, value='WB артикул поставщика')
    ws.cell(row=1, column=5, value='WB баркод')
    ws.cell(row=1, column=6, value='WB номенклатура')

    ws.cell(row=1, column=7, value='OZON артикул поставщика')
    ws.cell(row=1, column=8, value='OZON баркод')
    ws.cell(row=1, column=9, value='OZON PRODUCT_ID')
    ws.cell(row=1, column=10, value='OZON SKU')
    ws.cell(row=1, column=11, value='OZON FBO SKU ID')
    ws.cell(row=1, column=12, value='OZON FBS SKU ID')

    ws.cell(row=1, column=13, value='YANDEX артиукл поставщика')
    ws.cell(row=1, column=14, value='YANDEX баркод')
    ws.cell(row=1, column=15, value='YANDEX SKU')

    al = Alignment(horizontal="center",
                   vertical="center")
    al_left = Alignment(horizontal="left",
                        vertical="center")
    al2 = Alignment(vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")
    pattern = PatternFill('solid', fgColor="fcff52")

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15

    for i in range(len(data)+1):
        for c in ws[f'A{i+1}:O{i+1}']:
            for i in range(15):
                c[i].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[i].alignment = al_left

    # Сохраняем книгу Excel в память
    response = HttpResponse(content_type='application/xlsx')
    name = f'Article_table_{datetime.now().strftime("%Y.%m.%d")}.xlsx'
    file_data = 'attachment; filename=' + name
    response['Content-Disposition'] = file_data
    wb.save(response)
    return response


def excel_article_costprice_export(data):
    """Экспортирует в excel таблицу артикулы - себестоимость"""
    # Создаем DataFrame из данных
    wb = Workbook()
    # Получаем активный лист
    ws = wb.active
    # Заполняем лист данными
    for row, item in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=str(item.common_article))
        ws.cell(row=row, column=2, value=str(item.cost_price))

    # Устанавливаем заголовки столбцов
    ws.cell(row=1, column=1, value='Общий артикул')
    ws.cell(row=1, column=2, value='Себестоимость')

    al = Alignment(horizontal="center",
                   vertical="center")
    al_left = Alignment(horizontal="left",
                        vertical="center")
    al2 = Alignment(vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")
    pattern = PatternFill('solid', fgColor="fcff52")

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15

    for i in range(len(data)+1):
        for c in ws[f'A{i+1}:B{i+1}']:
            for i in range(2):
                c[i].border = Border(top=thin, left=thin,
                                     bottom=thin, right=thin)
                c[i].alignment = al_left

    # Сохраняем книгу Excel в память
    response = HttpResponse(content_type='application/xlsx')
    name = f'Article_cost_price_{datetime.now().strftime("%Y.%m.%d")}.xlsx'
    file_data = 'attachment; filename=' + name
    response['Content-Disposition'] = file_data
    wb.save(response)
    return response


def check_ur_lico_in_file(ur_lico_list, ur_lico, xlsx_file):
    """Проверка на совпадения юр. лица в файле и рабочей страницы"""
    checker_list = []
    for company_ind in range(len(ur_lico_list)):
        if ur_lico_list[company_ind] and str(ur_lico_list[company_ind]) != 'nan' and ur_lico != ur_lico_list[company_ind]:
            checker_list.append(ur_lico_list[company_ind])

    if checker_list:
        ur_lico_message = ''
        for ur in checker_list:
            ur_lico_message += f' \n{ur},'
        message_for_user = f'''В файле {xlsx_file} не совпадают юр. лица с выбранной вкладкой.\
            На вкладке юр. лицо {ur_lico}, в файле найдено {ur_lico_message}.\n
            Проверьте верно ли выбран файл?'''
        return json.dumps({"text": message_for_user})


def excel_import_group_create_data(xlsx_file, ur_lico):
    """
    Импортирует данные о группе из Excel
    Создает группы цен на основе полученных из Excel данных.
    """
    workbook = load_workbook(filename=xlsx_file, read_only=True)
    worksheet = workbook.active

    excel_data_common = pd.read_excel(xlsx_file)
    column_list = excel_data_common.columns.tolist()
    error_column_list = []
    for i in COLUMN_EXCEL_CREATING_GROUP_LIST:
        if i not in column_list:
            error_column_list.append(i)
    if not error_column_list:
        excel_data = pd.DataFrame(
            excel_data_common, columns=COLUMN_EXCEL_CREATING_GROUP_LIST)
        name_list = excel_data['Название'].to_list()
        ur_lico_list = excel_data['Юр. лицо'].to_list()
        old_price_list = excel_data['Старая цена'].to_list()
        wb_sale_list = excel_data['WB скидка продавца'].to_list()
        ozon_price_list = excel_data['Цена OZON'].to_list()
        yandex_price_list = excel_data['Цена YANDEX'].to_list()
        min_price_list = excel_data['Мин. цена'].to_list()

        ur_lico_checker = check_ur_lico_in_file(
            ur_lico_list, ur_lico, xlsx_file)
        if ur_lico_checker:
            return ur_lico_checker

        check_dict = {'Название': name_list,
                      'Юр. лицо': ur_lico_list,
                      'Старая цена': old_price_list,
                      'WB скидка продавца': wb_sale_list,
                      'Цена OZON': ozon_price_list,
                      'Цена YANDEX': yandex_price_list,
                      'Мин. цена': min_price_list}
        check_data_list = []
        for column_name, value_list in check_dict.items():
            if any('nan' in str(element) for element in value_list):
                check_data_list.append(column_name)

        if check_data_list:
            column = ''
            for data in check_data_list:
                column += f' {data},'
            message_for_user = f'''В файле {xlsx_file} проверьте длину столбца {column}. В нем не хватает данных'''
            return json.dumps({"text": message_for_user})

        for i in range(len(name_list)):
            if Groups.objects.filter(company=ur_lico_list[i],
                                     name=name_list[i]).exists():
                group_obj = Groups.objects.get(company=ur_lico_list[i],
                                               name=name_list[i])
                group_obj.wb_price = old_price_list[i]
                group_obj.wb_discount = wb_sale_list[i]
                group_obj.ozon_price = ozon_price_list[i]
                group_obj.yandex_price = yandex_price_list[i]
                group_obj.min_price = min_price_list[i]
                group_obj.old_price = old_price_list[i]
                group_obj.save()
            else:
                Groups(
                    name=name_list[i],
                    company=ur_lico_list[i],
                    wb_price=old_price_list[i],
                    wb_discount=wb_sale_list[i],
                    ozon_price=ozon_price_list[i],
                    yandex_price=yandex_price_list[i],
                    min_price=min_price_list[i],
                    old_price=old_price_list[i]
                ).save()

    else:
        message = ''
        for group in error_column_list:
            message += f'\n{group},'
        message_for_user = f'''Вы пытались загрузить ошибочный файл для создания/обновления групп цен для юр. лица {ur_lico}.\
        Название файла: {xlsx_file}.\
        В этом файле нет необходимых названий столбцов: {message}'''
        return json.dumps({"text": message_for_user})


def excel_import_data(xlsx_file, ur_lico):
    """Импортирует данные о группе артикула из Excel"""

    excel_data_common = pd.read_excel(xlsx_file)
    column_list = excel_data_common.columns.tolist()
    if 'Артикул' in column_list and 'Юр. лицо' in column_list and 'Ценовая группа артикула' in column_list:
        excel_data = pd.DataFrame(excel_data_common, columns=[
                                  'Артикул', 'Юр. лицо', 'Ценовая группа артикула'])
        article_list = excel_data['Артикул'].to_list()
        ur_lico_list = excel_data['Юр. лицо'].to_list()
        group_name_list = excel_data['Ценовая группа артикула'].to_list()

        # Словарь {артикул: название_группы}
        article_value_dict = {}
        # Список для обновления строк в БД
        new_objects = []
        error_group_name_list = []

        ur_lico_checker = check_ur_lico_in_file(
            ur_lico_list, ur_lico, xlsx_file)
        if ur_lico_checker:
            return ur_lico_checker

        for i in range(len(article_list)):
            article_value_dict[article_list[i]] = group_name_list[i]

        for row in range(len(article_list)):
            article_obj = ArticleGroup.objects.get(
                common_article=Articles.objects.get(company=ur_lico, common_article=article_list[row]))
            if Groups.objects.filter(name=article_value_dict[article_obj.common_article.common_article]).exists():
                article_obj.group = Groups.objects.get(company=ur_lico,
                                                       name=article_value_dict[article_obj.common_article.common_article])
            elif str(article_value_dict[article_obj.common_article.common_article]) != 'nan':
                error_group_name_list.append(
                    article_value_dict[article_obj.common_article.common_article])
                article_obj.group = None
            else:
                article_obj.group = None
            new_objects.append(article_obj)

        ArticleGroup.objects.bulk_update(new_objects, ['group'])
        if error_group_name_list:
            message = ''
            for group in error_group_name_list:
                message += f'\n{group},'
            message_for_user = f'''В базе данных {ur_lico} не нашел названия групп из файла {xlsx_file}: {message}.\n\nЭти группы не сохранил.'''

            return json.dumps({"text": message_for_user})
    else:
        message_for_user = f'''Вы пытались загрузить ошибочный файл c ценовыми группами  для юр. лица {ur_lico}.\
        Название файла: {xlsx_file}.\
        В этом файле нет необходимых названий столбцов: Артикул, Ценовая группа артикула'''
        return json.dumps({"text": message_for_user})


def excel_import_article_costprice_data(xlsx_file, ur_lico):
    """Импортирует данные о группе артикула из Excel"""

    excel_data_common = pd.read_excel(xlsx_file)
    excel_data = pd.DataFrame(excel_data_common, columns=[
                              'Общий артикул', 'Себестоимость'])
    article_list = excel_data['Общий артикул'].to_list()
    costprice_list = excel_data['Себестоимость'].to_list()

    # Словарь {артикул: название_группы}
    article_value_dict = {}
    # Список для обновления строк в БД
    new_objects = []

    for i in range(len(article_list)):
        article_value_dict[article_list[i]] = costprice_list[i]

    for row in range(len(article_list)):
        article_obj = Articles.objects.get(
            company=ur_lico, common_article=article_list[row])
        article_obj.cost_price = article_value_dict[article_obj.common_article]
        new_objects.append(article_obj)

    Articles.objects.bulk_update(new_objects, ['cost_price'])


def excel_new_group_data(xlsx_file):
    """Создает новые ценовые группы, взяв данные из Excel"""
    workbook = load_workbook(filename=xlsx_file, read_only=True)
    worksheet = workbook.active
    # Читаем файл построчно и создаем объекты.
    for row in range(1, len(list(worksheet.rows))):
        if list(worksheet.rows)[row][1].value == None or list(worksheet.rows)[row][1].value == 'None':
            article = ArticleGroup.objects.get(
                common_article=Articles.objects.get(common_article=list(worksheet.rows)[row][0].value))
            article.group = None
            article.save()
        else:
            new_obj = ArticleGroup.objects.filter(
                common_article=Articles.objects.get(
                    common_article=list(worksheet.rows)[row][0].value)
            ).update(group=Groups.objects.get(name=list(worksheet.rows)[row][1].value))


def wb_price_changer(header, info_list: list):
    """Изменяет цену входящего списка артикулов на WB"""
    time.sleep(1)
    url = 'https://discounts-prices-api.wb.ru/api/v2/upload/task'
    payload = json.dumps({"data": info_list})
    response_data = requests.request(
        "POST", url, headers=header, data=payload)
    print(response_data.status_code)
    print(response_data.text)
    print(info_list)


def articles_price_discount(ur_lico):
    """
    Функция возвращает данные по всем артикулам о текущей цене
    после скидки продавца и текущей скидки продавца
    {nm_id: {'price': price, 'discount': discount}}

    """
    from api_request.wb_requests import get_article_list_data
    header = header_wb_dict[ur_lico]
    main_data = get_article_list_data(header)
    returned_dict = {}
    for data in main_data:
        price = data['sizes'][0]['price']
        discount = data['discount']
        nm_id = data['nmID']

        returned_dict[nm_id] = {'price': price, 'discount': discount}
    return returned_dict



def wilberries_price_change(ur_lico, articles_list: list, price: int, discount: int):
    """
    Изменяет цену на артикулы Wildberries

    current_price - словарь для проверки текущей цены со скидкой продавца.
        Если текущая цена и скидка равна устанавливаемой, то товар не попадает на изменение цены
        {nm_id: {'price': price, 'discount': discount}}
    """
    koef_articles = math.ceil(len(articles_list)/1000)
    header = header_wb_dict[ur_lico]
    current_price = articles_price_discount(ur_lico)
    for i in range(koef_articles):
        data_for_change = []
        start_point = i*1000
        finish_point = (i+1)*1000
        data_articles_list = articles_list[
            start_point:finish_point]
        for article in data_articles_list:
            if article != None and article in current_price:
                if current_price[article]['discount'] != discount:
                    inner_data_dict = {
                        "nmID": article,
                        "price": price,
                        "discount": discount
                    }
                    data_for_change.append(inner_data_dict)
        
        wb_price_changer(header, data_for_change)


def ozon_price_changer(header, info_list: list):
    """Изменяет цену входящего списка артикулов на OZON"""
    url = 'https://api-seller.ozon.ru/v1/product/import/prices'
    payload = json.dumps({"prices": info_list})

    response_data = requests.request(
        "POST", url, headers=header, data=payload)
    # print('ozon_price_changer', response_data.status_code)


def ozon_price_change(ur_lico, articles_list: list, price: float, min_price: float, old_price=0):
    """Изменяет цену на артикулы OZON"""
    koef_articles = math.ceil(len(articles_list)/1000)
    header = header_ozon_dict[ur_lico]
    for i in range(koef_articles):
        data_for_change = []
        start_point = i*1000
        finish_point = (i+1)*1000
        data_articles_list = articles_list[
            start_point:finish_point]
        for article in data_articles_list:
            if article != None:
                inner_data_dict = {
                    "auto_action_enabled": "UNKNOWN",
                    "currency_code": "RUB",
                    "min_price": str(min_price),
                    "offer_id": "",
                    "old_price": str(old_price),
                    "price": str(price),
                    "price_strategy_enabled": "UNKNOWN",
                    "product_id": article
                }
                data_for_change.append(inner_data_dict)
        ozon_price_changer(header, data_for_change)


def yandex_price_changer(header, business_id, info_list: list):
    """Изменяет цену входящего списка артикулов на OZON"""
    url = f'https://api.partner.market.yandex.ru/businesses/{business_id}/offer-prices/updates'
    payload = json.dumps({"offers": info_list})
    response_data = requests.request(
        "POST", url, headers=header, data=payload)


def yandex_price_change(ur_lico, articles_list: list, price: float, old_price=0):
    """Изменяет цену на артикулы YANDEX"""
    koef_articles = math.ceil(len(articles_list)/500)
    header = header_yandex_dict[ur_lico]
    business_id = yandex_business_id_dict[ur_lico]
    for i in range(koef_articles):
        data_for_change = []
        start_point = i*500
        finish_point = (i+1)*500
        data_articles_list = articles_list[
            start_point:finish_point]
        for article in data_articles_list:
            if article != None:
                inner_data_dict = {
                    "offerId": article,
                    "price": {
                        "value": price,
                        "currencyId": "RUR",
                        "discountBase": old_price
                    }
                }
                data_for_change.append(inner_data_dict)
        yandex_price_changer(header, business_id, data_for_change)


@sender_error_to_tg
def wb_articles_list(ur_lico, offset=0, article_price_data=None, iter_numb=0):
    """Получаем массив арткулов с ценами и скидками для ВБ"""
    if article_price_data == None:
        article_price_data = []
    url = f'https://discounts-prices-api.wb.ru/api/v2/list/goods/filter?limit=1000&offset={offset}'
    header = header_wb_dict[ur_lico]
    response = requests.request("GET", url, headers=header)
    iter_numb += 1
    if response.status_code == 200:
        main_data = json.loads(response.text)['data']['listGoods']
        for data in main_data:
            inner_dict = {}
            inner_dict['nmId'] = data['nmID']
            inner_dict['price'] = data['sizes'][0]['price']
            inner_dict['discount'] = data['discount']
            article_price_data.append(inner_dict)
        if len(main_data) == 1000:

            offset = 1000 * iter_numb
            wb_articles_list(ur_lico, offset, article_price_data, iter_numb)
        return article_price_data
    elif response.status_code != 200 and iter_numb < 10:
        time.sleep(10)
        return wb_articles_list(ur_lico, offset, article_price_data, iter_numb)
    else:
        text = f'Приложение price_system. supplyment. Функция: wb_articles_list. Статус код: {response.status_code}'
        bot.send_message(chat_id=CHAT_ID_ADMIN,
                         text=text, parse_mode='HTML')


@sender_error_to_tg
def ozon_articles_list(ur_lico, last_id='', main_price_data=None):
    """Получаем массив арткулов с ценами и скидками для OZON"""
    header = header_ozon_dict[ur_lico]
    if main_price_data is None:
        main_price_data = []
    url = 'https://api-seller.ozon.ru/v4/product/info/prices'
    payload = json.dumps({
        "filter": {
            "offer_id": [],
            "product_id": [],
            "visibility": "ALL"
        },
        "last_id": last_id,
        "limit": 1000
    })
    response = requests.request(
        "POST", url, headers=header, data=payload)
    if response.status_code == 200:
        article_price_data = json.loads(response.text)['result']['items']
        for data in article_price_data:
            main_price_data.append(data)
        if len(article_price_data) == 1000:
            ozon_articles_list(ur_lico, json.loads(
                response.text)['result']['last_id'], main_price_data)
        return main_price_data
    else:
        return ozon_articles_list(ur_lico)


@sender_error_to_tg
def yandex_articles_list(ur_lico, page_token='', main_price_data=None):
    """Получаем массив арткулов с ценами и скидками для YANDEX"""
    header = header_yandex_dict[ur_lico]
    business_id = yandex_business_id_dict[ur_lico]
    if main_price_data is None:
        main_price_data = []

    url = f'https://api.partner.market.yandex.ru/businesses/{business_id}/offer-mappings?limit=200&page_token={page_token}'
    payload = {}
    response = requests.request(
        "POST", url, headers=header, data=payload)
    if response.status_code == 200:
        article_price_data = json.loads(response.text)[
            'result']['offerMappings']

        for data in article_price_data:
            main_price_data.append(data)
        if 'nextPageToken' in json.loads(response.text)[
                'result']['paging']:
            page_toket = json.loads(response.text)[
                'result']['paging']['nextPageToken']
        else:
            page_toket = ''
        if len(article_price_data) == 200:
            yandex_articles_list(ur_lico, page_toket, main_price_data)
        return main_price_data
    else:
        return yandex_articles_list(ur_lico)


def applies_price_for_price_group(group_name, ur_lico_name):
    """Присваивает цены артикулам, которые находятся в группе с названием group_name"""
    from price_system.periodical_tasks import (ozon_add_price_info,
                                               wb_add_price_info,
                                               yandex_add_price_info)
    names = ArticleGroup.objects.filter(
        group__name=group_name)
    if len(names) > 0:
        wb_price = names[0].group.wb_price
        wb_discount = names[0].group.wb_discount
        ozon_price = names[0].group.ozon_price
        yandex_price = names[0].group.yandex_price
        min_price = names[0].group.min_price
        old_price = names[0].group.old_price
        wb_nom_list = []
        oz_nom_list = []
        yandex_nom_list = []
        articles_wb_data_dict = articles_price_discount(ur_lico_name)
        for art in names:
            if art.common_article.wb_nomenclature in articles_wb_data_dict:
                wb_old_price = articles_wb_data_dict[art.common_article.wb_nomenclature]['price']
                wb_discount_from_wb = articles_wb_data_dict[art.common_article.wb_nomenclature]['discount']
                if wb_old_price != wb_price or wb_discount_from_wb != wb_discount:
                    wb_nom_list.append(art.common_article.wb_nomenclature)
            oz_nom_list.append(art.common_article.ozon_product_id)
            yandex_nom_list.append(
                art.common_article.yandex_seller_article)
        wilberries_price_change(
            ur_lico_name, wb_nom_list, wb_price, wb_discount)
        ozon_price_change(ur_lico_name, oz_nom_list,
                          ozon_price, min_price, old_price)
        yandex_price_change(ur_lico_name, yandex_nom_list,
                            yandex_price, old_price)
        # Записываем изененные цены в базу данных
        wb_add_price_info(ur_lico_name)
        ozon_add_price_info(ur_lico_name)
        yandex_add_price_info(ur_lico_name)
