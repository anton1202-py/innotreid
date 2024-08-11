from datetime import datetime

import pandas as pd
from analytika_reklama.models import ArticleCampaignWhiteList
from api_request.wb_requests import (advertisment_campaign_list,
                                     advertisment_campaigns_list_info,
                                     create_auto_advertisment_campaign,
                                     get_budget_adv_campaign)
from create_reklama.models import (AutoReplenish, CpmWbCampaign,
                                   CreatedCampaign, ProcentForAd,
                                   StartPausaCampaign)
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from price_system.models import Articles
from price_system.supplyment import sender_error_to_tg
from reklama.models import UrLico

from web_barcode.constants_file import CHAT_ID_ADMIN, bot, header_wb_dict


def check_data_for_create_adv_campaign(main_data):
    """Обрабатывает и проверяет данные для создания рекламной кампании"""

    raw_articles = main_data['articles']
    ur_lico = UrLico.objects.get(id=main_data['ur_lico'])
    ur_lico_name = ur_lico.ur_lice_name
    campaign_type = int(main_data['select_type'])
    subject_id = int(main_data['select_subject'])
    cpm = int(main_data['cpm'])
    budget = int(main_data['budget'])
    user_chat_id = int(main_data['user_chat_id'])

    header = header_wb_dict[ur_lico_name]
    result_articles = raw_articles.split(', ')

    mns_list = []
    main_articles_list = Articles.objects.filter(
        company=ur_lico).values('wb_nomenclature')
    articles_list = []
    for article in main_articles_list:
        articles_list.append(str(article['wb_nomenclature']))

    # Проверка артикулов на наличие в системе:
    for nmid in result_articles:
        if nmid not in articles_list:
            error = f'Нет артикула {nmid} в на портале у {ur_lico}.'
            bot.send_message(chat_id=user_chat_id,
                             text=error[:4000])
        else:
            mns_list.append(int(nmid))

    for nm_id in mns_list:
        if CreatedCampaign.objects.filter(ur_lico=ur_lico, articles_name=nm_id).exists():
            exist_campaign_obj = CreatedCampaign.objects.filter(
                ur_lico=ur_lico, articles_name=nm_id)[0]
            error = f'Кампания с артикулом {nmid} уже есть у {ur_lico}. Ее номер: {exist_campaign_obj.campaign_number}, название: {exist_campaign_obj.campaign_name}'
            bot.send_message(chat_id=user_chat_id,
                             text=error[:4000])
        else:
            nm_id_for_request = [nm_id]
            article_name = Articles.objects.filter(
                company=ur_lico, wb_nomenclature=nm_id)[0].name
            name_for_request = f'{nm_id} {article_name}'

            response = create_auto_advertisment_campaign(
                header, campaign_type, name_for_request, subject_id, budget, nm_id_for_request, cpm)

            if response.status_code != 200:
                error = f'Не создал кампанию для артикула {nm_id} ({article_name}). Ошибка: {response.text}'
                bot.send_message(chat_id=user_chat_id,
                                 text=error[:4000])
            else:
                error = f'Создал кампанию для артикула {nm_id} ({article_name}). Ее номер: {response.text}'
                bot.send_message(chat_id=user_chat_id,
                                 text=error[:4000])
            saved_data = {
                'ur_lico': UrLico.objects.get(id=main_data['ur_lico']),
                'campaign_name': name_for_request,
                'campaign_number': int(response.text),
                'campaign_type': int(main_data['select_type']),
                'subject_id': int(main_data['select_subject']),
                'cpm': int(main_data['cpm']),
                'budget': int(main_data['budget']),
                'article': nm_id
            }

            add_created_campaign_data_to_database(saved_data)


@sender_error_to_tg
def add_created_campaign_data_to_database(main_data):
    """Обрабатывает и проверяет данные для создания рекламной кампании"""

    article = main_data['article']
    ur_lico = main_data['ur_lico']
    campaign_number = main_data['campaign_number']
    campaign_type = main_data['campaign_type']
    campaign_name = main_data['campaign_name']
    subject_id = int(main_data['subject_id'])

    cpm = main_data['cpm']
    budget = main_data['budget']
    today = datetime.now().strftime("%Y-%m-%d %H:%M")

    CreatedCampaign(
        ur_lico=ur_lico,
        campaign_name=campaign_name,
        campaign_number=campaign_number,
        campaign_type=campaign_type,
        campaign_status=9,
        create_date=today,
        articles_name=article,
        subject_id=subject_id,
        current_cpm=cpm
    ).save()

    adv_obj = CreatedCampaign.objects.get(
        ur_lico=ur_lico, campaign_number=campaign_number, campaign_name=campaign_name)
    ProcentForAd(
        campaign_number=adv_obj,
        koefficient=4,
        koef_date=today,
        virtual_budget=0,
        virtual_budget_date=today,
        campaign_budget_date=today
    ).save()

    AutoReplenish(
        campaign_number=adv_obj
    ).save()

    StartPausaCampaign(
        campaign_number=adv_obj,
        date_status=today,
        campaign_status=9
    ).save()


@sender_error_to_tg
def filter_campaigns_status_type() -> dict:
    """
    Фильтрует рекламные кампании по статусу и типу.
    Возвращает словарь типа:
    {ur_lico:
        {campaign_type:
            {campaign_status: [campaign_number]}
        }
    }
    """
    ur_lico_data = UrLico.objects.all()
    returned_dict = {}
    type_list = [8, 9]
    for ur_lico_obj in ur_lico_data:
        header = header_wb_dict[ur_lico_obj.ur_lice_name]
        campaigns_wb_answer_list = advertisment_campaign_list(header)
        inner_campaigns_data = {}
        for data in campaigns_wb_answer_list['adverts']:

            # Автоматические кампании
            if data['type'] in type_list:

                if data['type'] in inner_campaigns_data:
                    inner_campaign_list = []
                    for campaign in data['advert_list']:
                        inner_campaign_list.append(campaign['advertId'])
                    inner_campaigns_data[data['type']
                                         ][data['status']] = inner_campaign_list
                else:
                    inner_campaign_list = []
                    for campaign in data['advert_list']:
                        inner_campaign_list.append(campaign['advertId'])
                    inner_campaigns_data[data['type']] = {
                        data['status']: inner_campaign_list}
        returned_dict[ur_lico_obj.ur_lice_name] = inner_campaigns_data
    return returned_dict


@sender_error_to_tg
def filter_campaigns_status_only() -> dict:
    """
    Фильтрует рекламные кампании по статусу.
    Возвращает словарь типа:
    {ur_lico:
        {campaign_number: campaign_status}
    }
    """
    ur_lico_data = UrLico.objects.all()
    returned_dict = {}
    type_list = [8, 9]
    for ur_lico_obj in ur_lico_data:
        header = header_wb_dict[ur_lico_obj.ur_lice_name]
        campaigns_wb_answer_list = advertisment_campaign_list(header)
        inner_campaigns_dict = {}
        for data in campaigns_wb_answer_list['adverts']:
            # Автоматические кампании
            if data['type'] in type_list:
                for campaign in data['advert_list']:
                    inner_campaigns_dict[campaign['advertId']] = data['type']

        returned_dict[ur_lico_obj.ur_lice_name] = inner_campaigns_dict
    return returned_dict


@sender_error_to_tg
def update_campaign_cpm(data_adv_list, ur_lico_obj, header):
    """
    Обновляет cpm кампании
    """
    main_data = advertisment_campaigns_list_info(data_adv_list, header)
    date_now = datetime.now()

    for campaign_data in main_data:
        if "autoParams" in campaign_data:
            if 'nmCPM' in campaign_data["autoParams"]:
                cpm = campaign_data["autoParams"]['nmCPM'][0]['cpm']
            elif 'cpm' in campaign_data["autoParams"]:
                cpm = campaign_data["autoParams"]['cpm']
            else:
                cpm = None
        else:
            cpm = None
        if cpm:
            campaign_obj = CreatedCampaign.objects.get(
                ur_lico=ur_lico_obj, campaign_number=str(campaign_data["advertId"]))
            campaign_obj.current_cpm = cpm
            campaign_obj.save()
            cpm_data = CpmWbCampaign.objects.filter(
                campaign_number=campaign_obj).order_by('-id').first()
            if cpm_data:
                if cpm_data.cpm != cpm:
                    CpmWbCampaign(
                        campaign_number=campaign_obj,
                        cpm=cpm,
                        cpm_date=date_now
                    ).save()
            else:
                CpmWbCampaign(
                    campaign_number=campaign_obj,
                    cpm=cpm,
                    cpm_date=date_now
                ).save()


@sender_error_to_tg
def update_campaign_budget(campaign_obj, header):
    """
    Обновляет бюджет кампании
    """
    budget_data = get_budget_adv_campaign(
        header, int(campaign_obj.campaign_number))
    if budget_data:
        balance = budget_data['total']
        campaign_obj.balance = balance
        campaign_obj.save()


@sender_error_to_tg
def white_list_phrase(campaign_obj):
    """
    Возвращает белый список кампании
    """
    if ArticleCampaignWhiteList.objects.filter(campaign=campaign_obj).exists():
        white_phrases = ArticleCampaignWhiteList.objects.filter(
            campaign=campaign_obj)
        main_list_for_return = []
        for phrase in white_phrases:
            white_phrase = phrase.phrase_list
            white_list = white_phrase.split(', ')
            for word in white_list:
                if word not in main_list_for_return:
                    main_list_for_return.append(word)
        return main_list_for_return
    else:
        return []


@sender_error_to_tg
def create_reklama_template_excel_file(ur_lico_data, subject_id):
    """Создает и скачивает excel файл"""
    # Создаем DataFrame из данных
    wb = Workbook()
    # Получаем активный лист
    ws = wb.active

    # Устанавливаем заголовки столбцов
    ws.cell(row=1, column=1, value='Юр. лицо')
    ws.cell(row=1, column=2, value='Тип кампании')
    ws.cell(row=1, column=3, value='Предмет')
    ws.cell(row=1, column=4, value='Артикул WB (nmID)')
    ws.cell(row=1, column=5, value='Бюджет')
    ws.cell(row=1, column=6, value='Ставка')

    ws.cell(row=1, column=9, value='Юр. лица')
    for i in range(len(ur_lico_data)):
        ws.cell(row=2+i, column=9, value=ur_lico_data[i].ur_lice_name)

    ws.cell(row=1, column=10, value='Типы кампаний')
    ws.cell(row=2, column=10, value='Автоматическая')

    ws.cell(row=1, column=11, value='Предметы')
    for j in range(len(subject_id.keys())):
        ws.cell(row=2+j, column=11, value=list(subject_id.keys())[j])

    ws.cell(row=1, column=12, value='Бюджеты')
    ws.cell(row=2, column=12, value='Минимум 1000р')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 16
    ws.column_dimensions['L'].width = 16

    # Сохраняем книгу Excel в память
    response = HttpResponse(content_type='application/xlsx')
    name = f'Template_campaign_reklama_{datetime.now().strftime("%Y.%m.%d")}.xlsx'
    file_data = 'attachment; filename=' + name
    response['Content-Disposition'] = file_data
    wb.save(response)

    return response


@sender_error_to_tg
def create_reklama_excel_with_campaign_data(xlsx_file):
    """Импортирует данные о группе артикула из Excel"""
    excel_data_common = pd.read_excel(xlsx_file)
    column_list = excel_data_common.columns.tolist()
    if 'Юр. лицо' in column_list and 'Тип кампании' in column_list and 'Предмет' in column_list and 'Артикул WB (nmID)' in column_list:
        excel_data = pd.DataFrame(excel_data_common, columns=[
                                  'Юр. лицо', 'Тип кампании', 'Предмет', 'Артикул WB (nmID)', 'Бюджет', 'Ставка'])
        article_list = excel_data['Артикул WB (nmID)'].to_list()
        ur_lico_list = excel_data['Юр. лицо'].to_list()
        type_list = excel_data['Тип кампании'].to_list()
        subject_list = excel_data['Предмет'].to_list()
        budget_list = excel_data['Бюджет'].to_list()
        cpm_list = excel_data['Ставка'].to_list()
        # return article_list, ur_lico_list, type_list, subject_list, budget_list, cpm_list

    else:
        return f'Вы пытались загрузить ошибочный файл {xlsx_file}.'


def check_data_create_adv_campaign_from_excel_file(articles_list, nmid, ur_lico, user_chat_id, campaign_type, subject_id, budget, cpm):
    """Обрабатывает и проверяет данные для создания рекламной кампании, полученные из excel файла"""
    if nmid not in articles_list:
        error = f'Нет артикула {nmid} в на портале у {ur_lico}.'
        bot.send_message(chat_id=user_chat_id,
                         text=error[:4000])
    else:
        if CreatedCampaign.objects.filter(ur_lico__ur_lice_name=ur_lico, articles_name=nmid).exists():
            exist_campaign_obj = CreatedCampaign.objects.filter(
                ur_lico__ur_lice_name=ur_lico, articles_name=nmid)[0]
            error = f'Кампания с артикулом {nmid} уже есть у {ur_lico}. Ее номер: {exist_campaign_obj.campaign_number}, название: {exist_campaign_obj.campaign_name}'
            bot.send_message(chat_id=user_chat_id,
                             text=error[:4000])
        else:
            nm_id_for_request = [int(nmid)]
            article_name = Articles.objects.filter(
                company=ur_lico, wb_nomenclature=nmid)[0].name
            name_for_request = f'{nmid} {article_name}'
            header = header_wb_dict[ur_lico]
            response = create_auto_advertisment_campaign(
                header, campaign_type, name_for_request, subject_id, budget, nm_id_for_request, cpm)
            if response.status_code != 200:
                error = f'Не создал кампанию для артикула {nmid} ({article_name}). Ошибка: {response.text}'
                bot.send_message(chat_id=user_chat_id,
                                 text=error[:4000])
            else:
                error = f'Создал кампанию для артикула {nmid} ({article_name}). Ее номер: {response.text}'
                bot.send_message(chat_id=user_chat_id,
                                 text=error[:4000])
                saved_data = {
                    'ur_lico': UrLico.objects.get(ur_lice_name=ur_lico),
                    'campaign_name': name_for_request,
                    'campaign_number': int(response.text),
                    'campaign_type': campaign_type,
                    'subject_id': subject_id,
                    'cpm': cpm,
                    'budget': budget,
                    'article': nmid
                }
                add_created_campaign_data_to_database(saved_data)
